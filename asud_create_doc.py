"""
Скрипт для пакетного создания Входящих документов в АСУД ИК
из Excel-файла (колонка B — краткое содержание, колонка C — корреспондент).

Установка:
    pip install selenium openpyxl pyinstaller

Сборка exe:
    pyinstaller --onefile --name asud_create_doc asud_create_doc.py

Положи msedgedriver.exe рядом с exe/скриптом.
"""

import time
import sys
import os
from datetime import date
import openpyxl

try:
    import pyautogui
    PYAUTOGUI_AVAILABLE = True
except ImportError:
    PYAUTOGUI_AVAILABLE = False

try:
    from pywinauto.application import Application
    from pywinauto import timings
    PYWINAUTO_AVAILABLE = True
except ImportError:
    PYWINAUTO_AVAILABLE = False
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains


# ===== НАСТРОЙКИ =====
ASUD_URL = "https://asud.interrao.ru/asudik/"
TIMEOUT = 20
AUTO_REGISTER = False  # True — регистрирует автоматически, False — останавливается перед регистрацией


def get_driver_path():
    if getattr(sys, 'frozen', False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))

    driver_path = os.path.join(base_dir, "msedgedriver.exe")

    if not os.path.exists(driver_path):
        print(f"!! msedgedriver.exe не найден в папке: {base_dir}")
        print("   Положи msedgedriver.exe рядом с этим файлом.")
        input("Enter для выхода...")
        sys.exit(1)

    return driver_path


import re as _re

# Маппинг индекса из Excel → название вида в АСУД
DOC_TYPE_MAP = {
    1: "Указы, распоряжения Президента Российской Федерации",
    2: "Документы Администрации Президента",
    3: "Документы Правительства Российской Федерации",
    4: "Документы Федеральных органов исполнительной и законодательной власти",
    5: "Письма юридических лиц",
    6: "Письма компаний Топливно-энергетического комплекса",
    7: "Документы органов законодательной и исполнительной власти субъектов",
    8: "Письма, заявления и жалобы граждан, акционеров",
}


def _clean_body(text):
    """Очищает TextBody:
    - Убирает строку про 'ВНИМАНИЕ! Письмо было отправлено внешним...'
    - Убирает лишние пустые строки
    - Убирает _x000D_
    """
    if not text:
        return ""
    t = str(text).replace('_x000D_', '\n')
    # Удаляем строки начинающиеся с "ВНИМАНИЕ! Письмо было отправлено внешним"
    # (и близкие варианты — регистр не важен)
    lines = t.split('\n')
    cleaned_lines = []
    for line in lines:
        stripped = line.strip()
        # Пропускаем уведомление от почты
        if _re.search(r'внимание!?\s*письмо\s+было\s+отправлено\s+внешним',
                      stripped, _re.IGNORECASE):
            continue
        cleaned_lines.append(line)
    t = '\n'.join(cleaned_lines)
    # Схлопываем множественные пустые строки в одну
    t = _re.sub(r'\n\s*\n\s*\n+', '\n\n', t)
    # Убираем пробелы в начале/конце
    return t.strip()


def _parse_sender(body):
    """Извлекает ФИО отправителя из 'From: ...' в теле письма."""
    if not body:
        return ""
    clean = body.replace('_x000D_', '\n')
    m = _re.search(r'From:\s*([^<\n\r]+?)(?:\s*<[^>]*>)?[\n\r]', clean)
    if not m:
        return ""
    name = m.group(1).strip()
    # Убираем разные "VASILYEVA TATIANA" варианты — только кириллица
    return name


def load_excel(file_path):
    """Читает Excel и возвращает список писем.
    Ожидаемые колонки:
      B: Subject (тема)
      C: TextBody (тело с From:)
      G: Тип (0=пропустить, 1-8 = индекс вида в АСУД)
    Строки с Тип=0 или без темы — пропускаются.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if len(row) < 7:
            skipped += 1
            continue
        subject = row[1]  # B
        body = row[2]     # C
        type_idx = row[6] # G

        if not subject:
            skipped += 1
            continue
        # Парсим индекс
        try:
            type_idx = int(type_idx) if type_idx is not None else 0
        except (ValueError, TypeError):
            type_idx = 0
        if type_idx == 0 or type_idx not in DOC_TYPE_MAP:
            skipped += 1
            continue

        # Чистим "FW: " из темы (используется только для лога)
        clean_subject = str(subject).strip()
        clean_subject = _re.sub(r'^(FW:|RE:|Fwd:)\s*', '', clean_subject, flags=_re.IGNORECASE)

        # Краткое содержание = TextBody (колонка C), очищенное от служебных строк
        body_clean = _clean_body(body) if body else clean_subject

        # Link из колонки A — дата-время как в имени .msg файла.
        # Сохраняем исходный объект (может быть str или datetime) — find_msg_by_link
        # обработает оба варианта и сгенерирует все возможные форматы имени.
        link = row[0]

        # Отправитель из TextBody ("From: ФИО")
        sender = _parse_sender(body) if body else ""

        rows.append({
            "содержание": body_clean,
            "корреспондент": sender or "Не указан",
            "тема": clean_subject,  # только для лога
            "тип_индекс": type_idx,
            "тип_название": DOC_TYPE_MAP[type_idx],
            "link": link,
        })
    wb.close()
    print(f"Загружено писем: {len(rows)}, пропущено (Тип=0 или пустые): {skipped}")
    return rows


def js_click(driver, element, description=""):
    """Надёжный клик с fallback'ами для GWT/GXT-элементов."""
    # Прокручиваем к элементу
    try:
        driver.execute_script(
            "arguments[0].scrollIntoView({block: 'center', inline: 'center'});", element)
        time.sleep(0.2)
    except Exception:
        pass

    # Способ 1: ActionChains — имитирует реальную мышь с hover (лучше всего для GXT dropdown)
    try:
        from selenium.webdriver.common.action_chains import ActionChains as AC
        AC(driver).move_to_element(element).pause(0.2).click().perform()
        print(f"  ОК клик (mouse): {description}")
        time.sleep(0.5)
        return
    except Exception:
        pass

    # Способ 2: обычный Selenium click
    try:
        element.click()
        print(f"  ОК клик: {description}")
        time.sleep(0.5)
        return
    except Exception:
        pass

    # Способ 3: JS .click()
    try:
        driver.execute_script("arguments[0].click();", element)
        print(f"  ОК JS-клик: {description}")
        time.sleep(0.5)
        return
    except Exception:
        pass

    # Способ 4: полный набор mouse-событий (для самых упрямых GWT)
    try:
        driver.execute_script("""
            var el = arguments[0];
            var events = ['mouseover', 'mousedown', 'mouseup', 'click'];
            events.forEach(function(type) {
                var evt = new MouseEvent(type, {bubbles: true, cancelable: true, view: window});
                el.dispatchEvent(evt);
            });
        """, element)
        print(f"  ОК mouse-events: {description}")
        time.sleep(0.5)
    except Exception as e:
        print(f"  !! Клик не удался: {description}: {e}")


def wait_asud_loaded(driver, max_wait=120):
    """Адаптивное ожидание полной загрузки АСУД.
    1) document.readyState === 'complete'
    2) Кнопка создания документа кликабельна
    3) В таблице документов появились строки (данные прогрузились)
    """
    print("  Жду готовности страницы...")
    try:
        WebDriverWait(driver, max_wait).until(
            lambda d: d.execute_script("return document.readyState === 'complete'")
        )
    except Exception:
        print("  ! readyState не complete за отведённое время")

    print("  Жду кнопку создания документа...")
    try:
        WebDriverWait(driver, max_wait).until(
            EC.element_to_be_clickable((By.ID, "mainscreen-create-button"))
        )
    except Exception:
        print("  ! Кнопка создания не появилась")

    print("  Жду загрузку данных в таблице...")
    try:
        WebDriverWait(driver, max_wait).until(
            lambda d: len(d.find_elements(By.CSS_SELECTOR,
                "tr[class*='GridView-row'], tr[class*='grid-row'], "
                "tr[class*='OSHSGridStyle-row'], tr[class*='obj-list-rec']")) > 0
        )
    except Exception:
        print("  ! Данные в таблице не появились — продолжаем")

    # Дополнительная пауза чтобы GWT дорисовал UI
    time.sleep(5)
    print("  ОК АСУД загружен")


def wait_and_click(driver, by, selector, description="", timeout=TIMEOUT):
    print(f"  -> Ожидаю: {description or selector}")
    el = WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, selector))
    )
    time.sleep(0.5)
    try:
        el.click()
    except Exception:
        driver.execute_script("arguments[0].click();", el)
    print(f"  ОК Клик: {description or selector}")
    time.sleep(0.5)
    return el


def fio_to_initials(full_name):
    """Преобразует 'Калганова Тамара Алексеевна' -> 'Калганова Т А'
    для сопоставления с тем, что показывает АСУД."""
    parts = full_name.strip().split()
    if len(parts) >= 3:
        return f"{parts[0]} {parts[1][0]} {parts[2][0]}"
    elif len(parts) == 2:
        return f"{parts[0]} {parts[1][0]}"
    return parts[0] if parts else full_name


def match_correspondent(text, full_name):
    """Проверяет совпадение: текст из АСУД (инициалы) vs полное ФИО из Excel.
    Мягкая версия — fallback на фамилию. Подходит для адресата (Басманов —
    один в базе). НЕ использовать для корреспондента (однофамильцы!)."""
    text_clean = text.strip()
    if full_name in text_clean:
        return True
    initials = fio_to_initials(full_name)
    text_norm = text_clean.replace('.', '').replace(',', '')
    initials_norm = initials.replace('.', '').replace(',', '')
    if initials_norm.lower() in text_norm.lower():
        return True
    # Совпадение только по фамилии (фоллбэк)
    surname = full_name.split()[0]
    if text_clean.lower().startswith(surname.lower()):
        return True
    return False


def match_correspondent_strict(text, full_name):
    """Строгая проверка: ТОЛЬКО полное ФИО или инициалы.
    Без фоллбэка на фамилию — чтобы не выбрать однофамильца."""
    text_clean = text.strip()
    # Прямое совпадение по полному ФИО
    if full_name in text_clean:
        return True
    # Совпадение по инициалам: "Калганова Т А" или "Калганова Т.А."
    initials = fio_to_initials(full_name)
    text_norm = text_clean.replace('.', '').replace(',', '')
    initials_norm = initials.replace('.', '').replace(',', '')
    if initials_norm.lower() in text_norm.lower():
        return True
    return False


def find_input_near_label(driver, label_text):
    """Находит input combobox в той же tr/контейнере что и лейбл."""
    labels = driver.find_elements(By.XPATH,
        f"//*[normalize-space(text())='{label_text}']")
    for label in labels:
        try:
            if not label.is_displayed():
                continue
            # Поднимаемся вверх ища контейнер с input
            for level in range(1, 6):
                parent = label
                for _ in range(level):
                    parent = parent.find_element(By.XPATH, "..")
                inputs = parent.find_elements(By.CSS_SELECTOR,
                    "input[id*='select_combobox-input'], input[type='text']")
                visible = [i for i in inputs
                           if i.is_displayed() and i.get_attribute("readonly") is None]
                if visible:
                    return visible[0]
        except Exception:
            continue
    return None


def create_correspondent(driver, person_name):
    """Создаёт нового корреспондента через диалоги АСУД.
    Flow:
      1. Клик '+' рядом с Корреспондент → 'Поиск корреспондента'
      2. 'Добавить' → 'Редактирование организации'
      3. В поле 'Поиск организации' ввести фамилию → выбрать из выпадающего
      4. Клик 'Добавить' в секции 'Физические лица'
      5. Заполнить карточку физ. лица (ФИО + Должность=ФЛ) → 'Добавить'
      6. 'Выбрать физ. лиц' → возврат в 'Поиск корреспондента'
      7. 'Готово'
    """
    parts = person_name.strip().split()
    if not parts:
        print(f"  !! Пустое ФИО корреспондента")
        return
    # Гибкий разбор: если отчества/имени нет — подставляем букву "Н"
    # (поле обязательное в АСУД, но хоть что-то)
    surname = parts[0]
    first_name = parts[1] if len(parts) >= 2 else "Н"
    middle_name = parts[2] if len(parts) >= 3 else "Н"
    if len(parts) < 3:
        print(f"  ! Неполное ФИО '{person_name}' — недостающие части = 'Н'")
    print(f"  Создаю нового корреспондента: {surname} {first_name} {middle_name}")

    # ШАГ 1: Клик "+" рядом с полем Корреспондент
    print("    [1/7] Клик '+' у Корреспондент...")
    try:
        # Находим лейбл Корреспондент, рядом ищем img с data-marker='select-btn'
        plus_btn = None
        label = driver.find_element(By.XPATH,
            "//*[normalize-space(text())='Корреспондент']")
        parent = label
        for _ in range(1, 7):
            parent = parent.find_element(By.XPATH, "..")
            btns = parent.find_elements(By.CSS_SELECTOR,
                "img[data-marker='select-btn'], img.gwt-Image")
            visible = [b for b in btns if b.is_displayed()]
            if visible:
                plus_btn = visible[-1]
                break
        if not plus_btn:
            print("    !! Кнопка '+' у Корреспондент не найдена")
            return
        js_click(driver, plus_btn, "+ Корреспондент")
        # Ждём появления диалога "Поиск корреспондента" вместо sleep
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.XPATH,
                    "//*[contains(text(),'Поиск корреспондента')]"))
            )
        except Exception:
            print("    ! Диалог 'Поиск корреспондента' не появился за 15 сек")
    except Exception as e:
        print(f"    !! Ошибка шаг 1: {e}")
        close_open_modals(driver)
        return

    # ШАГ 2: В "Поиск корреспондента" нажать "Добавить"
    print("    [2/7] 'Добавить' в Поиске корреспондента...")
    try:
        # Ждём появления кнопки "Добавить"
        add_btn = None
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.XPATH,
                    "//*[normalize-space(text())='Добавить']"))
            )
        except Exception:
            pass
        btns = driver.find_elements(By.XPATH,
            "//*[normalize-space(text())='Добавить']")
        for b in btns:
            if b.is_displayed():
                add_btn = b
                break
        if not add_btn:
            print("    !! Кнопка 'Добавить' не найдена")
            close_open_modals(driver)
            return
        js_click(driver, add_btn, "Добавить (новый корреспондент)")
        # Ждём появления диалога "Редактирование организации"
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.XPATH,
                    "//*[contains(text(),'Редактирование организации') or "
                    "contains(text(),'Поиск организации')]"))
            )
        except Exception:
            print("    ! Диалог 'Редактирование организации' не появился за 15 сек")
    except Exception as e:
        print(f"    !! Ошибка шаг 2: {e}")
        close_open_modals(driver)
        return

    # ШАГ 3: В "Редактирование организации" ввести фамилию в "Поиск организации"
    print("    [3/7] Поиск организации...")

    try:
        # Ждём появления поля "Поиск организации" (до 15 сек)
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.XPATH,
                    "//*[normalize-space(text())='Поиск организации']"))
            )
        except Exception:
            print("    ! Поле 'Поиск организации' не появилось за 15 сек")
        # Ищем input рядом с лейблом "Поиск организации" и ВВОДИМ значение
        # через JS в одной операции — чтобы избежать stale element reference.
        # JS возвращает 'ok' если нашёл и ввёл, 'no-label'/'no-input' если не удалось.
        js_result = driver.execute_script("""
            var surname = arguments[0];
            // Ищем все элементы с текстом "Поиск организации"
            var xpath = "//*[normalize-space(text())='Поиск организации']";
            var iter = document.evaluate(xpath, document, null,
                XPathResult.ORDERED_NODE_ITERATOR_TYPE, null);
            var labels = [];
            var node;
            while ((node = iter.iterateNext()) !== null) {
                if (node.offsetParent !== null) labels.push(node);
            }
            if (labels.length === 0) return 'no-label';

            // Для каждого лейбла ищем input рядом, поднимаясь вверх по DOM
            for (var li = 0; li < labels.length; li++) {
                var parent = labels[li];
                for (var lvl = 0; lvl < 6; lvl++) {
                    parent = parent.parentElement;
                    if (!parent) break;
                    var inputs = parent.querySelectorAll('input[type="text"]');
                    for (var i = 0; i < inputs.length; i++) {
                        var inp = inputs[i];
                        if (inp.offsetParent !== null && !inp.readOnly) {
                            // Нашли input — вводим
                            inp.focus();
                            inp.value = surname;
                            inp.dispatchEvent(new Event('input', {bubbles: true}));
                            inp.dispatchEvent(new Event('keyup', {bubbles: true}));
                            inp.dispatchEvent(new Event('change', {bubbles: true}));
                            return 'ok';
                        }
                    }
                }
            }
            return 'no-input';
        """, surname)
        print(f"    JS ввод в Поиск организации: {js_result}")

        if js_result != 'ok':
            print(f"    !! Не удалось ввести в поле (результат: {js_result})")
            close_open_modals(driver)
            return

        # Ждём появления кнопки "Создать организацию" — она появляется
        # с задержкой ~1 сек после ввода фамилии (autocomplete обрабатывает)
        print("    Жду появления кнопки 'Создать организацию'...")
        create_org_btn = None
        for attempt in range(10):  # до 10 сек
            try:
                # 1) По id
                try:
                    btn = driver.find_element(By.CSS_SELECTOR,
                        "[id*='create_custom_org'], [id*='custom_org_button']")
                    if btn.is_displayed():
                        create_org_btn = btn
                        break
                except Exception:
                    pass

                # 2) По тексту
                btns = driver.find_elements(By.XPATH,
                    "//*[contains(text(),'Создать организацию')]")
                for b in btns:
                    if b.is_displayed():
                        create_org_btn = b
                        break
                if create_org_btn:
                    break
            except Exception:
                pass
            time.sleep(1)

        if create_org_btn:
            js_click(driver, create_org_btn, "Создать организацию")
            print(f"    ОК Организация создана: {surname}")
            time.sleep(1)
        else:
            # Старый fallback — выбор из dropdown
            print("    ! Кнопка 'Создать организацию' не найдена, пробую dropdown...")
            target = None
            candidates = driver.find_elements(By.XPATH,
                f"//*[contains(text(),'{surname}')]")
            for c in candidates:
                try:
                    if not c.is_displayed():
                        continue
                    if c.tag_name.lower() == 'input':
                        continue
                    text = c.text.strip()
                    if text == surname:
                        target = c
                        break
                except Exception:
                    continue
            if not target:
                for c in candidates:
                    try:
                        if c.is_displayed() and c.tag_name.lower() != 'input':
                            target = c
                            break
                    except Exception:
                        continue
            if target:
                driver.execute_script(
                    "arguments[0].scrollIntoView({block: 'center'});", target)
                time.sleep(0.3)
                ActionChains(driver).move_to_element(target).pause(0.3).click().perform()
                time.sleep(1)
                print(f"    ОК Организация выбрана: {target.text.strip()[:60]}")
            else:
                print("    !! Ни кнопки 'Создать', ни варианта в списке не найдено")
                close_open_modals(driver)
                return
    except Exception as e:
        print(f"    !! Ошибка шаг 3: {e}")
        close_open_modals(driver)
        return

    # ШАГ 4: В секции "Физические лица" нажать "Добавить"
    print("    [4/7] 'Добавить' в Физические лица...")
    try:
        # Ждём пока секция "Физические лица" появится
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.XPATH,
                    "//*[contains(text(),'Физические лица')]"))
            )
        except Exception:
            pass

        # Даём серверу время обработать создание организации
        time.sleep(3)

        # Ждём пока кнопка "Добавить" в физ.лицах станет КЛИКАБЕЛЬНОЙ
        # (после создания/выбора организации — не просто появится в DOM)
        print("    Жду активации кнопки 'Добавить'...")
        add_user_btn = None
        for attempt in range(20):  # до 20 сек ожидания
            try:
                # Сначала ищем по id
                try:
                    btn = driver.find_element(By.CSS_SELECTOR,
                        "[id*='header-organization-dialog-add-a-user-button']")
                except Exception:
                    btn = None

                if not btn:
                    # Фоллбэк: ищем в секции "Физические лица"
                    try:
                        section = driver.find_element(By.XPATH,
                            "//*[contains(text(),'Физические лица')]")
                        parent = section
                        for _ in range(1, 6):
                            parent = parent.find_element(By.XPATH, "..")
                            btns = parent.find_elements(By.XPATH,
                                ".//*[normalize-space(text())='Добавить']")
                            visible = [b for b in btns if b.is_displayed()]
                            if visible:
                                btn = visible[0]
                                break
                    except Exception:
                        pass

                if btn:
                    # Проверяем кликабельность: не disabled, видим, в окне
                    is_enabled = driver.execute_script("""
                        var el = arguments[0];
                        if (!el.offsetParent) return false;  // скрыт
                        if (el.getAttribute('aria-disabled') === 'true') return false;
                        if (el.classList.contains('x-disabled')) return false;
                        if (el.classList.contains('disabled')) return false;
                        var style = window.getComputedStyle(el);
                        if (style.pointerEvents === 'none') return false;
                        if (parseFloat(style.opacity) < 0.5) return false;
                        return true;
                    """, btn)
                    if is_enabled:
                        add_user_btn = btn
                        break
                    else:
                        print(f"    ! Кнопка пока disabled (попытка {attempt + 1})...")
            except Exception:
                pass
            time.sleep(1)

        if not add_user_btn:
            print("    !! Кнопка 'Добавить' не активировалась за 20 сек")
            close_open_modals(driver)
            return

        js_click(driver, add_user_btn, "Добавить физ. лицо")
        # Ждём открытия карточки физ. лица (поле "Фамилия")
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.XPATH,
                    "//*[normalize-space(text())='Фамилия']"))
            )
        except Exception:
            print("    ! Карточка физ. лица не появилась за 15 сек")
    except Exception as e:
        print(f"    !! Ошибка шаг 4: {e}")
        close_open_modals(driver)
        return

    # ШАГ 5: Заполнить карточку физ. лица и нажать "Добавить"
    print("    [5/7] Заполнение карточки физ. лица...")
    try:
        # Точные id из DevTools:
        # outer_person_dialog-last_name / last_name-input
        # outer_person_dialog-first_name / first_name-input
        # outer_person_dialog-middle_name / middle_name-input
        # outer_person_dialog-position / position-input
        fields = [
            ("Фамилия", surname, "outer_person_dialog-last_name-input"),
            ("Имя", first_name, "outer_person_dialog-first_name-input"),
            ("Отчество", middle_name, "outer_person_dialog-middle_name-input"),
            ("Должность", "ФЛ", "outer_person_dialog-position-input"),
        ]
        for label_text, value, input_id in fields:
            result = driver.execute_script("""
                var inputId = arguments[0];
                var value = arguments[1];

                var el = document.getElementById(inputId);
                if (!el) {
                    // Фоллбэк: ищем по id-подстроке (на случай суффикса)
                    var base = inputId.replace('-input', '');
                    var container = document.getElementById(base);
                    if (container) {
                        var inputs = container.querySelectorAll('input[type="text"]');
                        for (var i = 0; i < inputs.length; i++) {
                            if (inputs[i].offsetParent !== null && !inputs[i].readOnly) {
                                el = inputs[i];
                                break;
                            }
                        }
                    }
                }
                if (!el) return 'no-element:' + inputId;
                if (el.offsetParent === null) return 'hidden:' + inputId;

                el.focus();
                el.value = value;
                el.dispatchEvent(new Event('input', {bubbles: true}));
                el.dispatchEvent(new Event('change', {bubbles: true}));
                return 'ok:' + el.id;
            """, input_id, value)

            if result.startswith('ok'):
                print(f"      ОК {label_text}: {value}  [{result}]")
            else:
                print(f"      !! {label_text} не заполнено ({result})")
            time.sleep(0.3)

        # Нажимаем "Добавить" в карточке физ. лица
        save_btn = None
        try:
            save_btn = driver.find_element(By.CSS_SELECTOR,
                "[id*='Parton_person_dialog_save_button']")
        except Exception:
            pass
        if not save_btn:
            # Фоллбэк по тексту
            btns = driver.find_elements(By.XPATH,
                "//*[normalize-space(text())='Добавить']")
            visible = [b for b in btns if b.is_displayed()]
            if visible:
                save_btn = visible[-1]
        if save_btn:
            js_click(driver, save_btn, "Сохранить карточку")
            # Ждём закрытия карточки — возврат в "Редактирование организации"
            # (кнопка "Выбрать физ. лиц" должна стать активной)
            try:
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.XPATH,
                        "//*[contains(text(),'Выбрать физ')]"))
                )
            except Exception:
                print("    ! Кнопка 'Выбрать физ. лиц' не появилась за 15 сек")
    except Exception as e:
        print(f"    !! Ошибка шаг 5: {e}")
        close_open_modals(driver)
        return

    # ШАГ 6: Нажать "Выбрать физ. лиц"
    print("    [6/7] Выбрать физ. лиц...")
    try:
        select_btn = None
        try:
            select_btn = driver.find_element(By.CSS_SELECTOR,
                "[id*='Parton_organization_dialog_select_persons_button']")
        except Exception:
            pass
        if not select_btn:
            btns = driver.find_elements(By.XPATH,
                "//*[contains(text(),'Выбрать физ')]")
            visible = [b for b in btns if b.is_displayed()]
            if visible:
                select_btn = visible[0]
        if select_btn:
            js_click(driver, select_btn, "Выбрать физ. лиц")
            # Ждём возврата в диалог "Поиск корреспондента" — появления кнопки "Готово"
            try:
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.ID, "oshs-select-button"))
                )
            except Exception:
                print("    ! Кнопка 'Готово' не появилась за 15 сек")
        else:
            print("    !! Кнопка 'Выбрать физ. лиц' не найдена")
            close_open_modals(driver)
            return
    except Exception as e:
        print(f"    !! Ошибка шаг 6: {e}")
        close_open_modals(driver)
        return

    # ШАГ 7: Нажать "Готово" в "Поиск корреспондента"
    print("    [7/7] Готово...")
    try:
        done_btn = None
        try:
            done_btn = driver.find_element(By.ID, "oshs-select-button")
        except Exception:
            pass
        if not done_btn:
            btns = driver.find_elements(By.XPATH,
                "//*[normalize-space(text())='Готово']")
            visible = [b for b in btns if b.is_displayed()]
            if visible:
                done_btn = visible[0]
        if done_btn:
            js_click(driver, done_btn, "Готово")
            # Ждём закрытия модалки
            wait_modal_closed(driver)
            print(f"  ОК Корреспондент создан: {person_name}")
        else:
            print("    !! Кнопка 'Готово' не найдена")
            close_open_modals(driver)
    except Exception as e:
        print(f"    !! Ошибка шаг 7: {e}")
        close_open_modals(driver)


def fill_correspondent(driver, person_name):
    """Заполняет поле Корреспондент через combobox."""
    print(f"  Корреспондент: {person_name}")
    time.sleep(1)

    # Ищем поле строго рядом с лейблом "Корреспондент" (не "Номер у корреспондента"!)
    corr_input = find_input_near_label(driver, "Корреспондент")

    if not corr_input:
        print("  !! Поле корреспондента не найдено!")
        return

    # Вводим фамилию для поиска
    surname = person_name.split()[0]
    initials = fio_to_initials(person_name)

    corr_input.click()
    time.sleep(0.3)
    corr_input.clear()
    time.sleep(0.3)
    for char in surname:
        corr_input.send_keys(char)
        time.sleep(0.1)
    print(f"  ОК Введена фамилия: {surname}")
    time.sleep(2)

    # Кликаем через ActionChains — так работало в Служебной записке
    def find_all_with_surname():
        results = driver.find_elements(By.XPATH,
            f"//*[contains(text(),'{surname}')]")
        found = []
        for r in results:
            try:
                if not r.is_displayed() or r == corr_input:
                    continue
                if r.tag_name.lower() == 'input':
                    continue
                found.append(r)
            except Exception:
                continue
        return found

    all_results = find_all_with_surname()
    if not all_results:
        corr_input.send_keys(Keys.ENTER)
        time.sleep(2)
        all_results = find_all_with_surname()

    print(f"  Найдено кандидатов с фамилией '{surname}': {len(all_results)}")

    # Ищем СТРОГОЕ совпадение по инициалам (без фоллбэка на фамилию,
    # иначе выберем однофамильца вместо создания нового)
    target = None
    for r in all_results:
        try:
            if match_correspondent_strict(r.text, person_name):
                target = r
                print(f"  Совпадение по инициалам: {r.text.strip()[:80]}")
                break
        except Exception:
            continue

    if target:
        # Клик через ActionChains (как в Служебной записке — это работало)
        try:
            driver.execute_script(
                "arguments[0].scrollIntoView({block: 'center'});", target)
            time.sleep(0.3)
            ActionChains(driver).move_to_element(target).pause(0.3).click().perform()
            time.sleep(1)
            print(f"  ОК Корреспондент выбран: {person_name}")
            return
        except Exception as e:
            print(f"  !! Ошибка выбора корреспондента: {e}")
            return

    # Нет точного совпадения по инициалам — создаём нового корреспондента
    print(f"  По инициалам '{initials}' совпадений нет — создаю нового корреспондента")
    # Закрываем dropdown (если открыт) через Escape
    try:
        corr_input.send_keys(Keys.ESCAPE)
        time.sleep(0.5)
    except Exception:
        pass
    create_correspondent(driver, person_name)


def fill_corr_number(driver, index=None):
    """Заполняет поле 'Номер у корреспондента' значением 'б/н (N)'."""
    value = f"б/н ({index})" if index else "б/н"
    print(f"  Номер у корреспондента: {value}")

    inp = find_input_near_label(driver, "Номер у корреспондента")

    if not inp:
        print("  !! Поле 'Номер у корреспондента' не найдено")
        return

    try:
        driver.execute_script(
            "arguments[0].scrollIntoView({block: 'center'});", inp)
        time.sleep(0.3)
        inp.click()
        time.sleep(0.3)
        inp.clear()
        time.sleep(0.2)
        inp.send_keys(value)
        time.sleep(0.3)
        inp.send_keys(Keys.TAB)
        print("  ОК Номер заполнен")
    except Exception as e:
        # Fallback через JS
        try:
            driver.execute_script("""
                arguments[0].value = arguments[1];
                arguments[0].dispatchEvent(new Event('input', {bubbles: true}));
                arguments[0].dispatchEvent(new Event('change', {bubbles: true}));
                arguments[0].dispatchEvent(new Event('blur', {bubbles: true}));
            """, inp, value)
            print("  ОК Номер заполнен через JS")
        except Exception as e2:
            print(f"  !! Ошибка заполнения номера: {e2}")


def fill_corr_date(driver):
    """Заполняет поле 'Дата у корреспондента' сегодняшней датой.
    Использует точный поиск по лейблу, чтобы не попасть в 'Дата помещения в архив'
    или 'Дата регистрации'."""
    today = date.today().strftime("%d.%m.%Y")
    print(f"  Дата у корреспондента: {today}")

    # Ищем лейбл с ТОЧНЫМ совпадением 'Дата у корреспондента'
    labels = driver.find_elements(By.XPATH,
        "//*[normalize-space(text())='Дата у корреспондента']")

    inp = None
    for label in labels:
        try:
            if not label.is_displayed():
                continue
            # Поднимаемся вверх, ищем input в том же контейнере
            for level in range(1, 6):
                parent = label
                for _ in range(level):
                    parent = parent.find_element(By.XPATH, "..")
                inputs = parent.find_elements(By.CSS_SELECTOR, "input[type='text']")
                visible = [i for i in inputs
                           if i.is_displayed() and i.get_attribute("readonly") is None]
                if visible:
                    inp = visible[0]
                    break
            if inp:
                break
        except Exception:
            continue

    if not inp:
        print("  !! Поле 'Дата у корреспондента' не найдено")
        return

    try:
        # Прокручиваем к элементу чтобы он был в центре (и триггер не перекрывал)
        driver.execute_script(
            "arguments[0].scrollIntoView({block: 'center', inline: 'center'});", inp)
        time.sleep(0.5)

        # Клик через JS — обходит перекрытие другими элементами
        driver.execute_script("arguments[0].focus(); arguments[0].click();", inp)
        time.sleep(0.3)

        # Очистка: выделяем всё и удаляем
        inp.send_keys(Keys.CONTROL + "a")
        time.sleep(0.2)
        inp.send_keys(Keys.DELETE)
        time.sleep(0.2)

        # Вводим дату
        inp.send_keys(today)
        time.sleep(0.3)
        inp.send_keys(Keys.TAB)
        print(f"  ОК Дата заполнена: {today}")
    except Exception as e:
        # Финальный fallback: напрямую через JS set value + trigger change
        try:
            driver.execute_script("""
                arguments[0].value = arguments[1];
                arguments[0].dispatchEvent(new Event('input', {bubbles: true}));
                arguments[0].dispatchEvent(new Event('change', {bubbles: true}));
                arguments[0].dispatchEvent(new Event('blur', {bubbles: true}));
            """, inp, today)
            print(f"  ОК Дата заполнена через JS: {today}")
        except Exception as e2:
            print(f"  !! Ошибка заполнения даты: {e2}")


def fill_delivery_method(driver):
    """Выбирает 'Электронная почта' в поле 'Способ получения'.
    Это triggerfield: клик открывает контекстное меню со списком опций."""
    print("  Способ получения: Электронная почта")

    # Находим лейбл по точному тексту
    labels = driver.find_elements(By.XPATH,
        "//*[normalize-space(text())='Способ получения']")

    trigger = None
    for label in labels:
        try:
            if not label.is_displayed():
                continue
            # Поднимаемся вверх по DOM и ищем любой кликабельный элемент формы
            for level in range(1, 8):
                parent = label
                for _ in range(level):
                    parent = parent.find_element(By.XPATH, "..")

                # 1) Любой input (readonly или нет — не важно)
                inputs = parent.find_elements(By.CSS_SELECTOR, "input[type='text']")
                for i in inputs:
                    if i.is_displayed():
                        trigger = i
                        break
                if trigger:
                    break

                # 2) Триггер-элементы (div/img с классами trigger/combobox)
                for sel in ["div[class*='trigger']", "img[class*='trigger']",
                            "[class*='ComboBox']", "[class*='combobox']",
                            "[data-marker*='trigger']"]:
                    try:
                        el = parent.find_element(By.CSS_SELECTOR, sel)
                        if el.is_displayed():
                            trigger = el
                            break
                    except Exception:
                        continue
                if trigger:
                    break
            if trigger:
                break
        except Exception:
            continue

    if not trigger:
        print("  !! Поле 'Способ получения' не найдено")
        return
    print(f"  Найден элемент triggerfield: <{trigger.tag_name}>")

    # 2. Прокручиваем и кликаем чтобы открыть dropdown
    try:
        driver.execute_script(
            "arguments[0].scrollIntoView({block: 'center'});", trigger)
        time.sleep(0.3)
    except Exception:
        pass

    js_click(driver, trigger, "Открыть список 'Способ получения'")
    time.sleep(1.5)  # ждём появления dropdown

    # 3. Ищем "Электронная почта" в появившемся меню
    #    GXT BoundList рендерит опции в отдельном layer (z-index),
    #    поэтому ищем глобально по всему документу.
    target_text = "Электронная почта"
    option = None

    # Ждём появления опций в dropdown
    for attempt in range(3):
        # Ищем через find_elements (не WebDriverWait — dropdown может быть уже виден)
        candidates = driver.find_elements(By.XPATH,
            f"//*[contains(text(),'{target_text}')]")
        print(f"  Попытка {attempt + 1}: найдено {len(candidates)} элементов '{target_text}'")

        for c in candidates:
            try:
                if not c.is_displayed():
                    continue
                if c.tag_name.lower() == 'input':
                    continue
                # Пропускаем очевидно не-опции (слишком большие контейнеры)
                option = c
                break
            except Exception:
                continue

        if option:
            break
        time.sleep(1)

    if option:
        try:
            driver.execute_script(
                "arguments[0].scrollIntoView({block: 'center'});", option)
            time.sleep(0.3)
            # Клик через ActionChains — тот же способ что работает для dropdown адресатов
            ActionChains(driver).move_to_element(option).pause(0.3).click().perform()
            time.sleep(0.5)
            print("  ОК Способ получения выбран: Электронная почта")
            return
        except Exception as e:
            print(f"  ! ActionChains не сработал: {e}, пробую обычный клик...")
            try:
                option.click()
                time.sleep(0.5)
                print("  ОК Способ получения выбран (click)")
                return
            except Exception:
                # JS click как последний вариант
                driver.execute_script("arguments[0].click();", option)
                time.sleep(0.5)
                print("  ОК Способ получения выбран (JS)")
                return

    # Фоллбэк: <select>
    try:
        selects = driver.find_elements(By.TAG_NAME, "select")
        for sel in selects:
            if not sel.is_displayed():
                continue
            for opt in sel.find_elements(By.TAG_NAME, "option"):
                if target_text in opt.text:
                    opt.click()
                    print("  ОК Способ получения выбран (select)")
                    return
    except Exception:
        pass

    print("  !! 'Электронная почта' не найдена в списке")


OUTLOOK_SUBJECTS_DIR = r"D:\OutlookSubjects"


def get_attachment_path():
    """Ищет .msg-пустышку рядом с exe/скриптом (fallback)."""
    if getattr(sys, 'frozen', False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))

    msg_files = [f for f in os.listdir(base_dir) if f.lower().endswith('.msg')]
    if not msg_files:
        return None
    if len(msg_files) == 1:
        return os.path.join(base_dir, msg_files[0])
    print(f"  ! Найдено {len(msg_files)} .msg файлов-пустышек, берём: {msg_files[0]}")
    return os.path.join(base_dir, msg_files[0])


def _link_to_variants(link):
    """Возвращает список возможных вариантов имени файла для link.
    Обрабатывает datetime → строку и варианты с/без ведущих нулей."""
    from datetime import datetime, date as _date
    variants = set()

    # Если datetime/date — форматируем в разные паттерны
    if isinstance(link, (datetime, _date)):
        try:
            # С ведущим нулём: "16.04.2026 09-53-27"
            variants.add(link.strftime("%d.%m.%Y %H-%M-%S"))
            # Без ведущего нуля: "16.04.2026 9-53-27"
            # (кроссплатформенно — через замену)
            with_lead = link.strftime("%d.%m.%Y %H-%M-%S")
            # Если час начинается с 0 и не 00 — убираем лидирующий 0
            m = _re.match(r'^(\d{2}\.\d{2}\.\d{4})\s+0(\d)-(\d{2})-(\d{2})$', with_lead)
            if m:
                variants.add(f"{m.group(1)} {m.group(2)}-{m.group(3)}-{m.group(4)}")
            # ISO-формат на всякий: "2026-04-16 09:53:27"
            variants.add(link.strftime("%Y-%m-%d %H:%M:%S"))
            variants.add(link.strftime("%Y-%m-%d %H-%M-%S"))
        except Exception:
            pass
    else:
        s = str(link).strip()
        if s:
            variants.add(s)
            # Попытаемся с/без ведущего 0 в часе
            m = _re.match(r'^(\d{2}\.\d{2}\.\d{4})\s+(\d{1,2})-(\d{2})-(\d{2})$', s)
            if m:
                d, h, mn, sec = m.groups()
                # Без ведущего нуля
                variants.add(f"{d} {int(h)}-{mn}-{sec}")
                # С ведущим нулём
                variants.add(f"{d} {int(h):02d}-{mn}-{sec}")

    variants.discard('')
    return list(variants)


def find_msg_by_link(link, fallback_path=None):
    """Ищет .msg файл в D:\\OutlookSubjects по имени = link.
    Если не нашёл — возвращает fallback_path (пустышку).
    """
    print(f"  [attach] link = {link!r} (тип: {type(link).__name__})")

    if link is None or (isinstance(link, str) and not link.strip()):
        print(f"  ! link пустой — беру пустышку")
        return fallback_path

    if not os.path.isdir(OUTLOOK_SUBJECTS_DIR):
        print(f"  ! Папка {OUTLOOK_SUBJECTS_DIR} не найдена — беру пустышку")
        return fallback_path

    # Получаем все возможные варианты имени
    variants = _link_to_variants(link)
    print(f"  Ищу файл, варианты имени: {variants}")

    # Попытка 1: точное совпадение по любому варианту (+ .MSG / .msg)
    for v in variants:
        for ext in ('.msg', '.MSG'):
            path = os.path.join(OUTLOOK_SUBJECTS_DIR, v + ext)
            if os.path.isfile(path):
                print(f"  ОК Нашёл: {os.path.basename(path)}")
                return path

    # Попытка 2: ищем по подстроке (любой вариант в имени файла)
    try:
        all_files = os.listdir(OUTLOOK_SUBJECTS_DIR)
        msg_files = [f for f in all_files if f.lower().endswith('.msg')]
        for f in msg_files:
            for v in variants:
                if v in f:
                    full = os.path.join(OUTLOOK_SUBJECTS_DIR, f)
                    print(f"  ОК Нашёл (подстрока): {f}")
                    return full
    except Exception as e:
        print(f"  ! Ошибка при сканировании папки: {e}")

    print(f"  ! Файл для link={link!r} не найден в {OUTLOOK_SUBJECTS_DIR} — беру пустышку")
    return fallback_path


def wait_modal_closed(driver, timeout=15):
    """Ждёт пока закроется модальное окно GXT ModalPanel."""
    print("  Жду закрытия модального окна...")
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: not any(
                m.is_displayed() for m in d.find_elements(
                    By.CSS_SELECTOR, "div[class*='ModalPanel'][class*='panel']"
                )
            )
        )
        print("  ОК Модалка закрыта")
    except Exception:
        print("  ! Модалка всё ещё открыта — пробую закрыть через Esc")
        try:
            from selenium.webdriver.common.action_chains import ActionChains as AC
            AC(driver).send_keys(Keys.ESCAPE).perform()
            time.sleep(1)
        except Exception:
            pass


def close_open_modals(driver, max_escapes=5):
    """Закрывает все открытые модалки/диалоги через Escape.
    Используется при сбое в create_correspondent, чтобы последующие
    шаги заполнения формы (адресат и т.д.) не блокировались."""
    print("  Закрываю открытые диалоги через Escape...")
    try:
        from selenium.webdriver.common.action_chains import ActionChains as AC
        for i in range(max_escapes):
            # Проверяем есть ли видимые модалки
            modals = driver.find_elements(By.CSS_SELECTOR,
                "div[class*='ModalPanel'][class*='panel']")
            visible_modals = [m for m in modals if m.is_displayed()]
            if not visible_modals:
                print(f"  ОК Все модалки закрыты (попыток: {i})")
                return
            AC(driver).send_keys(Keys.ESCAPE).perform()
            time.sleep(1)
        print(f"  ! Не все модалки закрылись после {max_escapes} Escape")
    except Exception as e:
        print(f"  ! Ошибка закрытия модалок: {e}")


def attach_content(driver, file_path):
    """Присоединяет файл: пробует сначала через input[type=file],
    если не получается — использует pyautogui для ввода пути в нативный Explorer."""
    print(f"  Присоединение файла: {os.path.basename(file_path)}")

    # Стратегия 1: input[type=file] уже в DOM без кликов
    inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='file']")
    print(f"  input[type=file] до клика: {len(inputs)}")

    file_attached = False

    if inputs:
        try:
            driver.execute_script("""
                var el = arguments[0];
                el.style.display = 'block';
                el.style.visibility = 'visible';
                el.style.opacity = '1';
                el.removeAttribute('hidden');
            """, inputs[0])
            time.sleep(0.3)
            inputs[0].send_keys(file_path)
            time.sleep(1)
            driver.execute_script(
                "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));",
                inputs[0])
            print(f"  ОК Файл отправлен через input[type=file]")
            file_attached = True
        except Exception as e:
            print(f"  ! Не удалось через input: {e}")

    # Стратегия 2: кликаем кнопку — открывается нативный Explorer,
    # pywinauto находит окно по заголовку и шлёт клавиши именно ему
    if not file_attached:
        if not PYWINAUTO_AVAILABLE:
            print("  !! pywinauto не установлен — пропускаю прикрепление")
            return

        try:
            btn = WebDriverWait(driver, TIMEOUT).until(
                EC.presence_of_element_located((By.XPATH,
                    "//div[contains(text(),'Присоединить содержимое')]"))
            )
            js_click(driver, btn, "Присоединить содержимое")
        except Exception as e:
            print(f"  !! Кнопка 'Присоединить содержимое' не найдена: {e}")
            return

        print("  Жду открытия нативного Explorer...")
        time.sleep(2)

        # Подключаемся к окну Explorer — это может быть "Открыть", "Open",
        # "Выбор файла" и т.п. Пробуем несколько вариантов заголовков.
        try:
            app = None
            for title_re in [".*Открыт.*", ".*Open.*", ".*Выбор.*", ".*Choose.*"]:
                try:
                    app = Application(backend='win32').connect(
                        title_re=title_re, timeout=10
                    )
                    print(f"  Найдено окно по шаблону: {title_re}")
                    break
                except Exception:
                    continue

            if not app:
                print("  !! Окно Explorer не найдено — закрываю через Escape")
                if PYAUTOGUI_AVAILABLE:
                    pyautogui.press('escape')
                return

            dlg = app.top_window()
            # Делаем окно активным, чтобы ввод пошёл в него
            dlg.set_focus()
            time.sleep(0.5)

            # Набираем путь к файлу через type_keys (поддерживает кириллицу
            # если указать with_spaces=True)
            dlg.type_keys(file_path, with_spaces=True, pause=0.02)
            time.sleep(0.5)
            dlg.type_keys("{ENTER}")
            time.sleep(2)
            print(f"  ОК Файл выбран через Explorer: {os.path.basename(file_path)}")
            file_attached = True
        except Exception as e:
            print(f"  !! Ошибка pywinauto: {e}")
            # Пробуем закрыть Explorer через pyautogui
            if PYAUTOGUI_AVAILABLE:
                try:
                    pyautogui.press('escape')
                    time.sleep(1)
                except Exception:
                    pass
            return

    if not file_attached:
        print("  !! Не удалось прикрепить файл")
        return

    # Небольшая пауза на обработку загрузки
    time.sleep(2)

    # Подтверждаем загрузку в модалке
    try:
        confirm_btn = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,
                "#SetContentDialogBtnSend, [id*='SetContentDialogBtnSend']"))
        )
        js_click(driver, confirm_btn, "Подтвердить присоединение")
        time.sleep(3)
        print("  ОК Файл присоединён!")
    except Exception:
        try:
            btns = driver.find_elements(By.XPATH,
                "//button[contains(text(),'Присоединить')] | //div[contains(text(),'Присоединить')]")
            visible = [b for b in btns if b.is_displayed()]
            if visible:
                js_click(driver, visible[-1], "Подтвердить (fallback)")
                time.sleep(3)
                print("  ОК Файл присоединён!")
        except Exception as e:
            print(f"  !! Ошибка подтверждения: {e}")


def add_addressee(driver, person_name):
    """Добавляет адресата через combobox рядом с разделом 'Адресаты'.
    Использует ActionChains как в рабочей версии Служебной записки."""
    print(f"  Адресат: {person_name}")
    try:
        addr_input = find_input_near_label(driver, "Адресаты")

        if not addr_input:
            print("  !! Поле адресата не найдено")
            return

        surname = person_name.split()[0]
        initials = fio_to_initials(person_name)

        addr_input.click()
        time.sleep(0.5)
        addr_input.clear()
        time.sleep(0.3)
        for char in surname:
            addr_input.send_keys(char)
            time.sleep(0.1)
        print(f"  ОК Введена фамилия: {surname}")
        time.sleep(2)

        def find_all_with_surname():
            results = driver.find_elements(By.XPATH,
                f"//*[contains(text(),'{surname}')]")
            found = []
            for r in results:
                try:
                    if not r.is_displayed() or r == addr_input:
                        continue
                    if r.tag_name.lower() == 'input':
                        continue
                    found.append(r)
                except Exception:
                    continue
            return found

        all_results = find_all_with_surname()
        if not all_results:
            addr_input.send_keys(Keys.ENTER)
            time.sleep(2)
            all_results = find_all_with_surname()

        print(f"  Найдено кандидатов с фамилией '{surname}': {len(all_results)}")

        target = None
        for r in all_results:
            try:
                if match_correspondent(r.text, person_name):
                    target = r
                    print(f"  Совпадение по инициалам: {r.text.strip()[:80]}")
                    break
            except Exception:
                continue

        if not target and all_results:
            target = all_results[0]
            print(f"  ! По инициалам '{initials}' не нашли, беру первого: {target.text.strip()[:80]}")

        if not target:
            print(f"  !! Адресат не найден в списке: {person_name}")
            return

        # ActionChains click (как в Служебной записке)
        driver.execute_script(
            "arguments[0].scrollIntoView({block: 'center'});", target)
        time.sleep(0.3)
        ActionChains(driver).move_to_element(target).pause(0.3).click().perform()
        time.sleep(1)
        print(f"  ОК Адресат добавлен: {person_name}")
    except Exception as e:
        print(f"  !! Ошибка добавления адресата: {e}")


def go_to_distribution_tab(driver):
    """Переходит на вкладку 'Рассылка'."""
    print("  Переход на вкладку Рассылка...")
    try:
        tab = WebDriverWait(driver, TIMEOUT).until(
            EC.presence_of_element_located((By.XPATH,
                "//div[contains(text(),'Рассылка')] | //span[contains(text(),'Рассылка')]"))
        )
        js_click(driver, tab, "Vkladka Rassylka")
        time.sleep(2)
        print("  ОК Вкладка Рассылка открыта")
    except Exception as e:
        print(f"  !! Вкладка Рассылка не найдена: {e}")


def add_distribution_addressee(driver, person_name):
    """Добавляет адресата на вкладке 'Рассылка' через поле 'Добавить адресатов'."""
    print(f"  Рассылка адресат: {person_name}")
    try:
        # Ищем поле "Добавить адресатов" — combobox внизу
        corr_input = None

        # По id combobox
        inputs = driver.find_elements(By.CSS_SELECTOR, "input[id*='select_combobox-input']")
        visible = [i for i in inputs if i.is_displayed()]
        if visible:
            corr_input = visible[-1]  # берём последний видимый (внизу страницы)

        if not corr_input:
            # По лейблу
            try:
                label = driver.find_element(By.XPATH,
                    "//*[contains(text(),'Добавить адресатов')]")
                parent = label.find_element(By.XPATH, "./ancestor::div[contains(@class,'field')] | ./ancestor::tr")
                inp = parent.find_element(By.CSS_SELECTOR, "input[type='text']")
                if inp.is_displayed():
                    corr_input = inp
            except Exception:
                pass

        if not corr_input:
            print("  !! Поле 'Добавить адресатов' не найдено!")
            return

        # Вводим фамилию
        surname = person_name.split()[0]
        corr_input.click()
        time.sleep(0.3)
        corr_input.clear()
        corr_input.send_keys(surname)
        print(f"  ОК Введена фамилия: {surname}")
        time.sleep(2)

        # Ищем в выпадающем списке
        results = driver.find_elements(By.XPATH,
            f"//*[contains(text(),'{surname}')]")
        visible_results = [r for r in results if r.is_displayed() and r != corr_input]
        if visible_results:
            for r in visible_results:
                if match_correspondent(r.text, person_name):
                    js_click(driver, r, f"Vybor: {r.text.strip()}")
                    time.sleep(1)
                    print(f"  ОК Адресат добавлен: {person_name}")
                    return
            js_click(driver, visible_results[0], f"Vybor (pervyj): {visible_results[0].text.strip()}")
            time.sleep(1)
            print(f"  ОК Адресат добавлен: {person_name}")
        else:
            print(f"  !! Адресат не найден в списке: {person_name}")
    except Exception as e:
        print(f"  !! Ошибка: {e}")


def create_one_document(driver, doc_data, index, total):
    """Создаёт один входящий документ."""
    print(f"\n{'='*60}")
    print(f"ДОКУМЕНТ {index}/{total}")
    print(f"  Содержание: {doc_data['содержание'][:80]}...")
    print(f"  Корреспондент: {doc_data['корреспондент']}")
    print(f"{'='*60}")

    # ШАГ 1: Кнопка создания документа
    print("\n[1/5] Кнопка создания документа...")
    el = WebDriverWait(driver, TIMEOUT).until(
        EC.presence_of_element_located((By.ID, "mainscreen-create-button"))
    )
    time.sleep(1)
    js_click(driver, el, "Кнопка создания документа")
    time.sleep(3)

    # ШАГ 2: Тип — Входящий документ
    print("\n[2/5] Тип: Входящий документ...")
    wait_and_click(driver, By.XPATH,
        "//div[contains(text(),'Входящий документ')]",
        "Входящий документ")
    time.sleep(1)

    # ШАГ 3: Вид — берём из doc_data (определён типом в Excel)
    doc_subtype = doc_data.get("тип_название", "Письма, заявления и жалобы граждан, акционеров")
    # Для поиска берём первые 30 символов — достаточно уникально
    subtype_short = doc_subtype[:30]
    print(f"\n[3/5] Вид: {doc_subtype}...")
    wait_and_click(driver, By.XPATH,
        f"//div[contains(text(),'{subtype_short}')] | //td[contains(text(),'{subtype_short}')]",
        doc_subtype)
    time.sleep(0.5)

    # Кнопка "Создать документ"
    print("  Создать документ...")
    wait_and_click(driver, By.XPATH,
        "//button[contains(text(),'Создать документ')] | //div[contains(text(),'Создать документ')]",
        "Создать документ")
    print("  Жду загрузку формы (5 сек)...")
    time.sleep(5)

    # ШАГ 4: Заполнение формы
    print("\n[4/5] Заполняю форму...")

    # --- Краткое содержание ---
    print("\n  Краткое содержание:")
    try:
        textareas = driver.find_elements(By.TAG_NAME, "textarea")
        visible_ta = [ta for ta in textareas if ta.is_displayed()]
        if visible_ta:
            visible_ta[0].click()
            time.sleep(0.3)
            visible_ta[0].clear()
            visible_ta[0].send_keys(doc_data["содержание"])
            print(f"  ОК Заполнено")
        else:
            print("  !! Textarea не найдена")
    except Exception as e:
        print(f"  !! Ошибка: {e}")
    time.sleep(0.5)

    # --- Корреспондент ---
    print("\n  Корреспондент:")
    fill_correspondent(driver, doc_data["корреспондент"])
    time.sleep(1)

    # --- Номер у корреспондента ---
    print("\n  Номер:")
    fill_corr_number(driver, index)
    time.sleep(0.5)

    # --- Дата у корреспондента ---
    print("\n  Дата:")
    fill_corr_date(driver)
    time.sleep(0.5)

    # --- Адресат (Басманов) ---
    print("\n  Адресат:")
    add_addressee(driver, "Басманов Александр Владимирович")
    time.sleep(0.5)

    # --- Способ получения ---
    print("\n  Способ получения:")
    fill_delivery_method(driver)
    time.sleep(0.5)

    # ШАГ 5: Сохранение
    print("\n[5/8] Сохранение...")
    try:
        # Сначала пытаемся по id="header-save-btn"
        save_btn = None
        try:
            save_btn = WebDriverWait(driver, TIMEOUT).until(
                EC.element_to_be_clickable((By.ID, "header-save-btn"))
            )
        except Exception:
            # Фоллбэк: кнопка Сохранить в шапке по тексту
            btns = driver.find_elements(By.XPATH,
                "//*[normalize-space(text())='Сохранить']")
            for b in btns:
                if b.is_displayed():
                    save_btn = b
                    break

        if save_btn:
            js_click(driver, save_btn, "Сохранить")
            time.sleep(3)
            print(f"  ОК Документ {index}/{total} сохранён!")
        else:
            print("  !! Кнопка 'Сохранить' не найдена")
    except Exception as e:
        print(f"  !! Ошибка сохранения: {e}")

    # ШАГ 6: Присоединить содержимое
    # Ищем файл в D:\OutlookSubjects по имени = link, иначе пустышка
    attach_path = find_msg_by_link(doc_data.get("link"), doc_data.get("файл"))
    if attach_path:
        print(f"\n[6/7] Присоединение содержимого: {os.path.basename(attach_path)}")
        attach_content(driver, attach_path)
        wait_modal_closed(driver)
    else:
        print("\n[6/7] Нет файла для присоединения (пропускаю)")

    # ШАГ 7: Документ сохранён, оставляем в черновиках — закрываем карточку
    print(f"\n[7/7] Документ {index}/{total} — сохранён в черновики.")
    print(f"  Содержание: {doc_data['содержание'][:60]}...")
    print(f"  Корреспондент: {doc_data['корреспондент']}")

    # Закрываем карточку крестиком (если она ещё открыта)
    print("  Закрываю карточку...")
    time.sleep(2)
    try:
        close_btn = driver.find_element(By.ID, "header-close-btn")
        if close_btn.is_displayed():
            ActionChains(driver).move_to_element(close_btn).pause(0.3).click().perform()
            time.sleep(2)
            print("  ОК Карточка закрыта")
        else:
            print("  Карточка уже закрыта")
    except Exception:
        print("  Карточка уже закрыта")

    # Проверяем что кнопка создания документа доступна (мы на главной)
    try:
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "mainscreen-create-button"))
        )
    except Exception:
        # Если не нашли — полная перезагрузка как fallback
        print("  ! Главная не загрузилась — перезагружаю АСУД...")
        driver.get(ASUD_URL)
        wait_asud_loaded(driver)


def main():
    print("=" * 60)
    print("АСУД ИК - Пакетное создание Входящих документов")
    print("=" * 60)

    # --- Поиск Excel-файла рядом с exe/скриптом ---
    if getattr(sys, 'frozen', False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))

    xlsx_files = [f for f in os.listdir(base_dir) if f.lower().endswith('.xlsx')]

    if not xlsx_files:
        print(f"!! Не найден .xlsx файл в папке: {base_dir}")
        print("   Положи Excel-файл рядом с exe.")
        input("Enter...")
        sys.exit(1)
    elif len(xlsx_files) == 1:
        excel_path = os.path.join(base_dir, xlsx_files[0])
        print(f"\nНайден файл: {xlsx_files[0]}")
    else:
        print(f"\nНайдено {len(xlsx_files)} xlsx-файлов:")
        for i, f in enumerate(xlsx_files, 1):
            print(f"  {i}. {f}")
        choice = input("Выбери номер файла: ").strip()
        try:
            excel_path = os.path.join(base_dir, xlsx_files[int(choice) - 1])
        except (ValueError, IndexError):
            print("!! Неверный выбор")
            input("Enter...")
            sys.exit(1)

    # --- Поиск .msg файла ---
    msg_path = get_attachment_path()
    if msg_path:
        print(f"Файл для присоединения: {os.path.basename(msg_path)}")
    else:
        print("! .msg файл не найден — документы будут без вложения")

    # --- Загрузка данных ---
    print(f"\nЧтение файла: {excel_path}")
    docs = load_excel(excel_path)
    # Добавляем путь к файлу в каждый документ
    for doc in docs:
        doc["файл"] = msg_path
    print(f"Найдено документов: {len(docs)}")

    if not docs:
        print("!! Нет данных для создания!")
        input("Enter...")
        sys.exit(1)

    # Показать превью
    print("\nПервые 5 записей:")
    for i, d in enumerate(docs[:5], 1):
        print(f"  {i}. {d['корреспондент']} | {d['содержание'][:60]}...")

    print(f"\nВсего: {len(docs)} документов")
    confirm = input("Начать создание? (да/нет): ").strip().lower()
    if confirm not in ("da", "yes", "y", "д", "да"):
        print("Отменено.")
        sys.exit(0)

    # --- Запуск браузера ---
    driver_path = get_driver_path()
    print(f"\nEdgeDriver: {driver_path}")

    options = EdgeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--auth-server-whitelist=*.interrao.ru")
    options.add_argument("--auth-negotiate-delegate-whitelist=*.interrao.ru")
    options.add_argument("--log-level=3")
    options.add_experimental_option('excludeSwitches', ['enable-logging'])

    service = EdgeService(executable_path=driver_path)
    driver = webdriver.Edge(service=service, options=options)

    try:
        # Открываем АСУД
        print("\nОткрываю АСУД...")
        driver.get(ASUD_URL)
        wait_asud_loaded(driver)

        # Создаём документы в цикле
        for i, doc in enumerate(docs, 1):
            try:
                create_one_document(driver, doc, i, len(docs))
            except Exception as e:
                print(f"\n!! ОШИБКА при создании документа {i}: {e}")
                print("  Пробую следующий...")
                driver.get(ASUD_URL)
                wait_asud_loaded(driver)
                continue

        print(f"\n{'='*60}")
        print(f"ГОТОВО! Создано документов: {len(docs)}")
        print(f"{'='*60}")

        input("\nEnter для закрытия браузера...")

    except Exception as e:
        print(f"\n!! Ошибка: {e}")
        input("Enter для закрытия...")

    finally:
        driver.quit()
        print("\nОК Браузер закрыт.")


if __name__ == "__main__":
    main()
