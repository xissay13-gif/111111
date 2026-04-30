"""
mix_routing.py — Пакетное создание Входящих документов (auto-create + smart-routing).

Читает Excel (Лист2), извлекает ФИО корреспондента из TextBody.
Для каждой строки:
  - Определяет тип документа по индексу (колонка D)
  - Создаёт карточку, заполняет поля
  - Прикрепляет .msg из D:\\OutlookSubjects по Link
  - Если ФИО найдено → регистрирует + На резолюцию + Да
  - Если ФИО не найдено → корреспондент="Неизвестный...", оставляет в ЧЕРНОВИКАХ
    + WARNING в логе (чтобы вручную доработать после прогона)

Excel Лист2: A=Link, B=Subject, C=TextBody, D=Тип, E=To (игнорируем — пересыльщик).

Модули:
  config.py        — настройки (+ config.json)
  ui.py            — Selenium UI-хелперы
  correspondent.py — выбор/создание корреспондента + extract_fio_from_text
  attachments.py   — поиск и прикрепление файлов
"""

import os
import re
import sys
import json
import time
import logging
from datetime import date, datetime, timedelta

import openpyxl
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import config as cfg
from ui import (click, wait_and_click, find_input_near_label,
                wait_asud_loaded, wait_modal_closed, close_open_modals,
                js_set_value, js_type_combobox, find_dropdown_options)
from correspondent import (fill_correspondent_field, match_strict, fio_to_initials,
                           extract_fio_from_text)
from attachments import find_msg_by_link, get_dummy_msg, attach_content, move_to_done


# ================= LOGGING =================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
    handlers=[logging.StreamHandler()],
)
log = logging.getLogger("asud")
start_time = time.monotonic()


# ================= EXCEL =================

def _clean_body(text):
    """Очищает TextBody от служебных строк (Original Message, ВНИМАНИЕ…)."""
    if not text:
        return ""
    t = str(text).replace('_x000D_', '\n')
    lines = t.split('\n')
    cleaned = []
    for line in lines:
        s = line.strip()
        if re.search(r'внимание!?\s*письмо\s+было\s+отправлено\s+внешним', s, re.IGNORECASE):
            continue
        if re.match(r'^-{3,}\s*Original\s*Message\s*-{3,}$', s, re.IGNORECASE):
            continue
        cleaned.append(line)
    t = '\n'.join(cleaned)
    t = re.sub(r'\n\s*\n\s*\n+', '\n\n', t)
    return t.strip()


# ================= STATE (resume после крэша) =================

def _link_key(link):
    """Стабильный строковый ключ из Link (для state-файла)."""
    if link is None:
        return ""
    if isinstance(link, (datetime, date)):
        try:
            return link.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(link)
    return str(link).strip()


def _state_path(xlsx_path):
    """Путь к state-файлу рядом с exe, привязан к имени Excel."""
    name = os.path.splitext(os.path.basename(xlsx_path))[0]
    # Безопасное имя файла
    safe = re.sub(r'[^\w.\-]+', '_', name)
    return os.path.join(cfg.get_base_dir(), f"mix_state_{safe}.json")


def load_state(xlsx_path):
    """Загружает set обработанных Link-ключей."""
    path = _state_path(xlsx_path)
    if not os.path.isfile(path):
        return set()
    try:
        with open(path, encoding='utf-8') as f:
            data = json.load(f)
        return set(data.get('processed', []))
    except Exception as e:
        log.warning(f"Не удалось прочитать state {path}: {e}")
        return set()


def save_state(xlsx_path, processed_set):
    """Атомарно перезаписывает state-файл."""
    path = _state_path(xlsx_path)
    try:
        tmp = path + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump({'processed': sorted(processed_set)}, f,
                      ensure_ascii=False, indent=2)
        os.replace(tmp, path)
    except Exception as e:
        log.warning(f"Не удалось сохранить state {path}: {e}")


# ================= OUTPUT XLSX (для clean-resolutions) =================

# Округ → ФИО начальника абонентского отдела
OKRUG_TO_FIO = {
    "САО": "Гренц Екатерина Александровна",
    "ЦАО": "Емельянова Татьяна Николаевна",
    "ОАО": "Рендюк Юлия Павловна",
    "ЛАО": "Вырва Елена Анатольевна",
    "КАО": "Кравец Татьяна Александровна",
}


def _output_xlsx_path(excel_path):
    """<реестр>_резолюции.xlsx рядом с реестром."""
    base, _ext = os.path.splitext(excel_path)
    return base + "_резолюции.xlsx"


def _ensure_output_xlsx(path):
    """Создаёт output xlsx с шапкой если не существует."""
    if os.path.isfile(path):
        return
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Резолюции"
        ws.append(["ОПТС", "Округ", "ФИО", "Link"])
        for c in range(1, 5):
            ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
        widths = {1: 30, 2: 8, 3: 35, 4: 22}
        for col, w in widths.items():
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)].width = w
        ws.freeze_panes = "A2"
        wb.save(path)
    except Exception as e:
        log.warning(f"Не удалось создать {path}: {e}")


def _append_output_row(path, doc_data, asud_id):
    """Дописывает строку в output xlsx после успешной регистрации.
    Колонки: ОПТС | Округ | ФИО | Link
    """
    # Округ — пытаемся определить автоматически из TextBody
    okrug = None
    try:
        from okrug_parser import okrug_from_textbody
        okrug = okrug_from_textbody(doc_data.get("содержание"),
                                     base_dir_fn=cfg.get_base_dir)
    except Exception as e:
        log.warning(f"okrug_parser упал: {e}")
    fio = OKRUG_TO_FIO.get(okrug) if okrug else None
    link = doc_data.get("link")
    link_str = ""
    if link:
        if isinstance(link, (datetime, date)):
            link_str = link.strftime("%d.%m.%Y %H-%M-%S")
        else:
            link_str = str(link)
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        ws.append([asud_id or "", okrug or "", fio or "", link_str])
        wb.save(path)
        wb.close()
        log.info(f"  → {os.path.basename(path)}: "
                 f"{asud_id or '—'} | {okrug or '—'} | {fio or '—'}")
    except Exception as e:
        log.warning(f"Не удалось записать в {path}: {e}")


def load_excel(file_path):
    """Читает Лист2. Колонки: A=Link, B=Subject, C=TextBody, D=Тип (index).

    ФИО корреспондента извлекается из TextBody.
    Если не найдено — ставится заглушка (unknown_correspondent) и флаг corr_found=False
    → такой документ потом останется в черновиках.
    """
    sheet_name = settings.get("sheet_name", cfg.DEFAULTS["sheet_name"])
    unknown = settings.get("unknown_correspondent", cfg.DEFAULTS["unknown_correspondent"])

    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        log.warning(f"Лист '{sheet_name}' не найден, использую активный: {wb.active.title}")
        ws = wb.active

    rows = []
    skipped = 0
    unknown_rows = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
        if not row or len(row) < 4:
            skipped += 1
            continue
        link = row[0]
        subject = row[1]
        body = row[2]
        type_idx = row[3]

        if not subject:
            skipped += 1
            continue
        try:
            type_idx = int(type_idx) if type_idx is not None else 0
        except (ValueError, TypeError):
            type_idx = 0
        if type_idx == 0 or type_idx not in cfg.DOC_TYPE_MAP:
            skipped += 1
            continue

        clean_subject = re.sub(r'^(FW:|RE:|Fwd:)\s*', '',
                               str(subject).strip(), flags=re.IGNORECASE)
        body_clean = _clean_body(body) if body else clean_subject

        fio, fio_src = extract_fio_from_text(body)
        if fio:
            correspondent = fio
            corr_found = True
        else:
            correspondent = unknown
            corr_found = False
            unknown_rows.append((row_idx, clean_subject))

        rows.append({
            "row_idx": row_idx,
            "содержание": body_clean,
            "корреспондент": correspondent,
            "корр_найден": corr_found,
            "корр_источник": fio_src,
            "тема": clean_subject,
            "тип_индекс": type_idx,
            "тип_название": cfg.DOC_TYPE_MAP[type_idx],
            "link": link,
        })
    wb.close()

    log.info(f"Загружено: {len(rows)}, пропущено: {skipped}")
    log.info(f"  ФИО найдено: {sum(1 for r in rows if r['корр_найден'])}")
    log.info(f"  ФИО НЕ найдено: {len(unknown_rows)} (уйдут в черновики)")
    for ri, subj in unknown_rows:
        log.info(f"    → Row {ri}: {subj[:60]}")

    return rows


# ================= FORM FILLING =================

def fill_text(driver, text):
    """Заполняет краткое содержание (textarea) — JS-set, plain-поле."""
    try:
        areas = driver.find_elements(By.TAG_NAME, "textarea")
        visible = [a for a in areas if a.is_displayed()]
        if visible:
            js_set_value(driver, visible[0], text)
            log.info("Краткое содержание заполнено (JS)")
        else:
            log.warning("Textarea не найдена")
    except Exception as e:
        log.error(f"Ошибка заполнения содержания: {e}")


def fill_corr_number(driver, link=None):
    """Заполняет 'Номер у корреспондента' = 'б/н <link>'."""
    if isinstance(link, (datetime, date)):
        link_str = link.strftime("%d.%m.%Y %H-%M-%S")
    elif link:
        link_str = str(link).strip()
    else:
        link_str = ""
    value = f"б/н {link_str}" if link_str else "б/н"

    inp = find_input_near_label(driver, "Номер у корреспондента")
    if not inp:
        log.warning("Поле 'Номер у корреспондента' не найдено")
        return
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
        js_set_value(driver, inp, value)
        log.info(f"Номер (JS): {value}")
    except Exception as e:
        log.warning(f"Номер: ошибка {e}")


def fill_corr_date(driver):
    """Заполняет 'Дата у корреспондента' = сегодня."""
    today = date.today().strftime("%d.%m.%Y")
    labels = driver.find_elements(By.XPATH,
        "//*[normalize-space(text())='Дата у корреспондента']")
    inp = None
    for label in labels:
        try:
            if not label.is_displayed():
                continue
            for level in range(1, 6):
                parent = label
                for _ in range(level):
                    parent = parent.find_element(By.XPATH, "..")
                inputs = parent.find_elements(By.CSS_SELECTOR, "input[type='text']")
                visible = [i for i in inputs
                           if i.is_displayed() and i.get_attribute("readonly") is None]
                if visible:
                    inp = visible[0]
                    break
            if inp:
                break
        except Exception:
            continue

    if not inp:
        log.warning("Поле 'Дата у корреспондента' не найдено")
        return
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
        js_set_value(driver, inp, today)
        log.info(f"Дата (JS): {today}")
    except Exception as e:
        log.warning(f"Дата: ошибка {e}")


def fill_delivery_method(driver):
    """Выбирает 'Электронная почта' в 'Способ получения'."""
    target_text = settings.get("delivery_method", "Электронная почта")
    labels = driver.find_elements(By.XPATH,
        "//*[normalize-space(text())='Способ получения']")
    trigger = None
    for label in labels:
        try:
            if not label.is_displayed():
                continue
            for level in range(1, 8):
                parent = label
                for _ in range(level):
                    parent = parent.find_element(By.XPATH, "..")
                inputs = parent.find_elements(By.CSS_SELECTOR, "input[type='text']")
                for i in inputs:
                    if i.is_displayed():
                        trigger = i
                        break
                if trigger:
                    break
                for sel in ["div[class*='trigger']", "img[class*='trigger']"]:
                    try:
                        el = parent.find_element(By.CSS_SELECTOR, sel)
                        if el.is_displayed():
                            trigger = el
                            break
                    except Exception:
                        continue
                if trigger:
                    break
            if trigger:
                break
        except Exception:
            continue

    if not trigger:
        log.warning("Поле 'Способ получения' не найдено")
        return

    click(driver, trigger, "Способ получения")
    time.sleep(1.5)

    option = None
    for _ in range(3):
        candidates = driver.find_elements(By.XPATH,
            f"//*[contains(text(),'{target_text}')]")
        for c in candidates:
            try:
                if c.is_displayed() and c.tag_name.lower() != 'input':
                    option = c
                    break
            except Exception:
                continue
        if option:
            break
        time.sleep(1)

    if option:
        click(driver, option, target_text)
        log.info(f"Способ получения: {target_text}")
    else:
        log.warning(f"'{target_text}' не найдена в списке")


def add_addressee(driver, person_name):
    """Добавляет одного адресата через combobox."""
    inp = find_input_near_label(driver, "Адресаты")
    if not inp:
        log.warning("Поле адресата не найдено")
        return

    surname = person_name.split()[0]
    inp.click()
    # Combobox-autocomplete — JS-set + dispatch input/keyup, чтобы GXT
    # перерисовал выпадашку без посимвольного ввода.
    js_type_combobox(driver, inp, surname)

    from correspondent import match_correspondent

    all_results = []
    try:
        WebDriverWait(driver, 5).until(
            lambda d: len(find_dropdown_options(d, surname, inp)) > 0)
        all_results = find_dropdown_options(driver, surname, inp)
    except Exception:
        try:
            inp.send_keys(Keys.ENTER)
            WebDriverWait(driver, 3).until(
                lambda d: len(find_dropdown_options(d, surname, inp)) > 0)
            all_results = find_dropdown_options(driver, surname, inp)
        except Exception:
            pass

    target = None
    for r in all_results:
        try:
            if match_correspondent(r.text, person_name):
                target = r
                break
        except Exception:
            continue
    if not target and all_results:
        target = all_results[0]

    if target:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", target)
        ActionChains(driver).move_to_element(target).pause(0.2).click().perform()
        log.info(f"Адресат добавлен: {person_name}")
    else:
        log.warning(f"Адресат не найден: {person_name}")


# ================= REGISTRATION =================

_CAPTURE_ASUD_ID_JS = r"""
// Один проход по DOM на стороне браузера: ищем регистрационный номер.
// Возвращает строку или null. Стратегии по убыванию точности:
//   1. <b> внутри [data-marker='ScreenHeader1']
//   2. любой <b> с подходящим текстом
//   3. regex по тексту страницы
const RE = /\b([А-Я]{2,5}(?:\/[А-Я0-9.\-]+){2,})\b/u;
function looksLike(t) {
    if (!t) return false;
    t = t.trim();
    if (!t.includes('/') || t.length < 6) return false;
    if (/^\d{2}\.\d{2}\.\d{4}/.test(t)) return false;
    return true;
}
const header = document.querySelector("[data-marker='ScreenHeader1']");
if (header) {
    for (const b of header.querySelectorAll('b')) {
        const t = (b.textContent || '').trim();
        if (looksLike(t)) return t;
    }
}
for (const b of document.querySelectorAll('b')) {
    const t = (b.textContent || '').trim();
    if (looksLike(t)) return t;
}
const m = (document.body.innerText || '').match(RE);
return m ? m[1] : null;
"""


def capture_asud_id(driver, timeout=15):
    """Читает регистрационный номер документа после регистрации.
    Один JS-вызов на итерацию (быстрый поллинг 100ms).
    Возвращает 'ОПТС/8/19892' / 'ОРТС/СЗ/ТЭС-03-УПО-02/176' или None."""
    end = time.monotonic() + timeout
    while time.monotonic() < end:
        try:
            asud_id = driver.execute_script(_CAPTURE_ASUD_ID_JS)
            if asud_id:
                log.info(f"  asud_id: {asud_id!r}")
                return asud_id
        except Exception:
            pass
        time.sleep(0.1)

    log.warning("Регистрационный номер не захватили — пуст в output")
    return None


def register_and_resolve(driver, index, total):
    """Регистрирует + На резолюцию + Да. Возвращает asud_id (если захватили) или None."""
    log.info("Регистрирую...")
    registered = False
    asud_id = None
    try:
        btn = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,
                "#header-action-btn-register, [id*='header-action-btn-register']")))
        click(driver, btn, "Зарегистрировать")
        # capture_asud_id внутри сам ждёт через WebDriverWait — sleep не нужен
        asud_id = capture_asud_id(driver, timeout=10)
        if asud_id:
            log.info(f"Документ {index}/{total} ЗАРЕГИСТРИРОВАН: {asud_id}")
        else:
            log.warning(f"Документ {index}/{total} ЗАРЕГИСТРИРОВАН (номер не захватили)")
        registered = True
    except Exception:
        try:
            btn = driver.find_element(By.XPATH, "//div[contains(text(),'Зарегистрировать')]")
            click(driver, btn, "Зарегистрировать (fallback)")
            registered = True
        except Exception as e:
            log.error(f"'Зарегистрировать' не найдена: {e}")

    if not registered:
        return None

    # На резолюцию
    res_btn = None
    for _ in range(10):
        try:
            btn = driver.find_element(By.ID, "header-action-btn-send_on_resolution")
            if btn.is_displayed():
                res_btn = btn
                break
        except Exception:
            pass
        try:
            for b in driver.find_elements(By.XPATH, "//*[contains(text(),'На резолюцию')]"):
                if b.is_displayed():
                    res_btn = b
                    break
        except Exception:
            pass
        if res_btn:
            break
        time.sleep(1)

    if not res_btn:
        log.warning("'На резолюцию' не появилась")
        return asud_id

    click(driver, res_btn, "На резолюцию")

    # Да — сразу опрашиваем DOM, sleep не нужен (внутри цикла уже есть time.sleep(1) между попытками)
    yes_btn = None
    for _ in range(10):
        # 1) Точный id
        try:
            btn = driver.find_element(By.ID, "confirm_dialog_btn_yes")
            if btn.is_displayed():
                yes_btn = btn
                break
        except Exception:
            pass
        # 2) Substring id (GWT может добавлять префиксы/суффиксы)
        try:
            btn = driver.find_element(By.CSS_SELECTOR,
                "[id*='confirm_dialog_btn_yes'], [id*='confirm'][id*='yes']")
            if btn.is_displayed():
                yes_btn = btn
                break
        except Exception:
            pass
        # 3) По тексту "Да" в любом видимом элементе
        try:
            for b in driver.find_elements(By.XPATH, "//*[normalize-space(text())='Да']"):
                if b.is_displayed():
                    yes_btn = b
                    break
        except Exception:
            pass
        if yes_btn:
            break
        time.sleep(1)

    if yes_btn:
        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", yes_btn)
            time.sleep(0.3)
        except Exception:
            pass
        clicked = False
        # 1) ActionChains
        try:
            ActionChains(driver).move_to_element(yes_btn).pause(0.3).click().perform()
            log.info(f"Клик 'Да' (ActionChains): id={yes_btn.get_attribute('id')}")
            clicked = True
        except Exception as e:
            log.info(f"ActionChains 'Да' не сработал: {e}")
        # 2) JS
        if not clicked:
            try:
                driver.execute_script("arguments[0].click();", yes_btn)
                log.info("Клик 'Да' (JS)")
                clicked = True
            except Exception as e:
                log.info(f"JS 'Да' не сработал: {e}")
        # 3) Нативный
        if not clicked:
            try:
                yes_btn.click()
                log.info("Клик 'Да' (native)")
                clicked = True
            except Exception as e:
                log.info(f"Native 'Да' не сработал: {e}")
        # 4) Enter — GWT-диалоги часто принимают его
        if clicked:
            try:
                ActionChains(driver).send_keys(Keys.ENTER).perform()
            except Exception:
                pass
        # Ждём пока модалка закроется (диалог 'Да' уйдёт из DOM)
        try:
            WebDriverWait(driver, 5).until_not(
                EC.visibility_of(yes_btn))
        except Exception:
            pass
        log.info(f"Документ {index}/{total} НА РЕЗОЛЮЦИИ")
    else:
        log.warning("Диалог 'Да' не появился за 10 сек")
    return asud_id


def close_card_and_wait_main(driver):
    """Закрывает карточку через header-close-btn и ждёт главную страницу."""
    try:
        close_btn = driver.find_element(By.ID, "header-close-btn")
        if close_btn.is_displayed():
            ActionChains(driver).move_to_element(close_btn).pause(0.2).click().perform()
            log.info("Карточка закрыта")
        else:
            log.info("Карточка уже закрыта")
    except Exception:
        log.info("Карточка уже закрыта")

    # Ждём главную напрямую — это и есть «карточка ушла + список загружен»
    try:
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "mainscreen-create-button")))
    except Exception:
        log.warning("Главная не загрузилась — перезагружаю")
        driver.get(settings.get("asud_url", cfg.DEFAULTS["asud_url"]))
        wait_asud_loaded(driver)


# ================= DOCUMENT FLOW =================

def create_one_document(driver, doc_data, index, total):
    """Создаёт один входящий документ."""
    log.info(f"{'='*50}")
    log.info(f"ДОКУМЕНТ {index}/{total}: {doc_data['тема'][:60]}")
    log.info(f"Корреспондент: {doc_data['корреспондент']} "
             f"({'найден ' + (doc_data['корр_источник'] or '') if doc_data['корр_найден'] else 'ЗАГЛУШКА'})")
    log.info(f"Тип: {doc_data['тип_название']}")

    # [1/7] Кнопка создания
    el = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
        EC.presence_of_element_located((By.ID, "mainscreen-create-button")))
    click(driver, el, "Создать документ")

    # [2/7] Входящий документ
    wait_and_click(driver, By.XPATH,
        "//div[contains(text(),'Входящий документ')]", "Входящий документ")

    # [3/7] Вид
    subtype = doc_data.get("тип_название", "Письма, заявления и жалобы граждан, акционеров")
    short = subtype[:30]
    wait_and_click(driver, By.XPATH,
        f"//div[contains(text(),'{short}')] | //td[contains(text(),'{short}')]", subtype)

    wait_and_click(driver, By.XPATH,
        "//button[contains(text(),'Создать документ')] | //div[contains(text(),'Создать документ')]",
        "Создать документ")

    # Ждём пока форма документа отрендерится — появится textarea (краткое содержание)
    try:
        WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            lambda d: any(t.is_displayed() for t in d.find_elements(By.TAG_NAME, "textarea")))
    except Exception:
        log.warning("Textarea формы не появилась — продолжаю по таймауту")

    # [4/7] Заполнение формы
    fill_text(driver, doc_data["содержание"])
    fill_correspondent_field(driver, doc_data["корреспондент"])
    fill_corr_number(driver, doc_data.get("link"))
    fill_corr_date(driver)

    for addr in settings.get("addressees", cfg.DEFAULTS["addressees"]):
        add_addressee(driver, addr)

    fill_delivery_method(driver)

    # [5/7] Сохранение
    try:
        save_btn = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            EC.element_to_be_clickable((By.ID, "header-save-btn")))
        click(driver, save_btn, "Сохранить")
        # Ждём кнопку 'Зарегистрировать' — признак что save прошёл и форма ушла в режим регистрации
        try:
            WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
                EC.presence_of_element_located((By.CSS_SELECTOR,
                    "#header-action-btn-register, [id*='header-action-btn-register']")))
        except Exception:
            log.warning("После Сохранить кнопка 'Зарегистрировать' не появилась")
        log.info(f"Документ {index}/{total} сохранён")
    except Exception as e:
        log.error(f"Ошибка сохранения: {e}")

    # [6/7] Прикрепление
    outlook_dir = settings.get("outlook_dir", cfg.DEFAULTS["outlook_dir"])
    dummy_path = doc_data.get("файл")
    attach_path = find_msg_by_link(doc_data.get("link"), outlook_dir, dummy_path)
    if attach_path:
        log.info(f"Прикрепляю: {os.path.basename(attach_path)}")
        attach_content(driver, attach_path)
        wait_modal_closed(driver)
    else:
        log.info("Нет файла — пропускаю")

    # [7/7] Регистрация (если ФИО найдено) или черновик
    asud_id = None
    if doc_data["корр_найден"]:
        asud_id = register_and_resolve(driver, index, total)
        # После успешной регистрации — реальный (не dummy) .msg → Завершено/
        # Черновики НЕ переносим: файл нужен для ручной доработки.
        if attach_path and attach_path != dummy_path:
            move_to_done(attach_path, outlook_dir)
    else:
        log.warning(f"Row {doc_data['row_idx']}: ФИО НЕ найдено — "
                    f"оставляю в ЧЕРНОВИКАХ для ручной доработки "
                    f"(тема: {doc_data['тема'][:60]}). "
                    f"Файл НЕ перемещаю — лежит на месте.")

    close_card_and_wait_main(driver)
    return asud_id


# ================= MAIN =================

settings = {}


def main():
    global settings
    settings = cfg.load()

    log.info("=" * 50)
    log.info("АСУД ИК — MIX (auto-create + smart-routing)")
    log.info("=" * 50)

    base_dir = cfg.get_base_dir()

    # Excel
    xlsx_files = [f for f in os.listdir(base_dir) if f.lower().endswith('.xlsx')]
    if not xlsx_files:
        log.error(f"Нет .xlsx в {base_dir}")
        input("Enter...")
        sys.exit(1)
    elif len(xlsx_files) == 1:
        excel_path = os.path.join(base_dir, xlsx_files[0])
        log.info(f"Файл: {xlsx_files[0]}")
    else:
        print(f"\nНайдено {len(xlsx_files)} xlsx-файлов:")
        for i, f in enumerate(xlsx_files, 1):
            print(f"  {i}. {f}")
        choice = input("Выбери номер: ").strip()
        try:
            excel_path = os.path.join(base_dir, xlsx_files[int(choice) - 1])
        except (ValueError, IndexError):
            log.error("Неверный выбор")
            sys.exit(1)

    # Папка с .msg — интерактивный ввод (Enter = дефолт из config)
    default_outlook = settings.get("outlook_dir", cfg.DEFAULTS["outlook_dir"])
    print(f"\nПапка с .msg-файлами (поиск рекурсивно по подпапкам).")
    print(f"Нажми Enter, чтобы использовать: {default_outlook}")
    user_dir = input("Путь: ").strip().strip('"').strip("'")
    if user_dir:
        settings["outlook_dir"] = user_dir
    outlook_dir = settings["outlook_dir"]
    if not os.path.isdir(outlook_dir):
        log.warning(f"Папка '{outlook_dir}' не существует — "
                    f"все вложения уйдут как пустышки")
    else:
        log.info(f"Папка вложений: {outlook_dir}")

    # Пустышка (для случаев когда .msg по link не найден)
    msg_path = get_dummy_msg(base_dir)
    if msg_path:
        log.info(f"Пустышка: {os.path.basename(msg_path)}")

    # Данные
    docs = load_excel(excel_path)
    for doc in docs:
        doc["файл"] = msg_path

    if not docs:
        log.error("Нет данных!")
        input("Enter...")
        sys.exit(1)

    # Resume: проверяем state-файл
    processed = load_state(excel_path)
    if processed:
        done_in_current = [d for d in docs
                           if _link_key(d.get("link")) in processed]
        if done_in_current:
            print(f"\nВ state-файле {len(done_in_current)} ранее обработанных документов.")
            print("  Enter / 'да'  — ПРОПУСТИТЬ их, продолжить с остальных (по умолчанию)")
            print("  'нет'         — обработать ВСЁ заново (дубли в АСУД! не рекомендуется)")
            print("  'сброс'       — обнулить state и обработать всё (для полного старта)")
            ans = input("Что делаем? [да]: ").strip().lower()
            if ans in ("сброс", "reset"):
                save_state(excel_path, set())
                processed = set()
                log.info("State обнулён — обрабатываю всё заново")
            elif ans in ("нет", "н", "n", "no"):
                log.info("Обрабатываю всё заново (state будет дополнен)")
            else:
                before = len(docs)
                docs = [d for d in docs
                        if _link_key(d.get("link")) not in processed]
                log.info(f"Пропускаю {before - len(docs)} обработанных, "
                         f"осталось {len(docs)}")
                if not docs:
                    log.info("Все строки уже обработаны — нечего делать.")
                    input("Enter...")
                    sys.exit(0)

    known = sum(1 for d in docs if d["корр_найден"])
    unknown = len(docs) - known
    print(f"\nПервые 5:")
    for i, d in enumerate(docs[:5], 1):
        flag = 'OK' if d["корр_найден"] else '!!'
        print(f"  {i}. [{d['тип_индекс']}] {flag} {d['корреспондент'][:30]} | {d['тема'][:50]}")
    print(f"\nВсего к обработке: {len(docs)}  (ФИО: {known}, заглушка: {unknown})")

    confirm = input("Начать? (да/нет): ").strip().lower()
    if confirm not in ("да", "д", "y", "yes", ""):
        print("Отменено.")
        sys.exit(0)

    # Браузер
    driver_path = os.path.join(base_dir, "msedgedriver.exe")
    if not os.path.exists(driver_path):
        log.error(f"msedgedriver.exe не найден в {base_dir}")
        input("Enter...")
        sys.exit(1)

    options = EdgeOptions()
    # eager: driver.get() разблокируется на DOMContentLoaded, не ждёт картинки/iframes.
    # Все важные клики идут через WebDriverWait — ранний контроль безопасен.
    options.page_load_strategy = "eager"
    options.add_argument("--start-maximized")
    options.add_argument("--auth-server-whitelist=*.interrao.ru")
    options.add_argument("--auth-negotiate-delegate-whitelist=*.interrao.ru")
    options.add_argument("--log-level=3")
    options.add_experimental_option('excludeSwitches', ['enable-logging'])

    service = EdgeService(executable_path=driver_path)
    driver = webdriver.Edge(service=service, options=options)

    try:
        url = settings.get("asud_url", cfg.DEFAULTS["asud_url"])
        log.info(f"Открываю {url}")
        driver.get(url)
        wait_asud_loaded(driver)

        # Output xlsx для clean-resolutions: ОПТС | Округ | ФИО | Link
        output_path = _output_xlsx_path(excel_path)
        _ensure_output_xlsx(output_path)
        log.info(f"Output xlsx: {output_path}")

        done_count, err_count = 0, 0
        for i, doc in enumerate(docs, 1):
            try:
                asud_id = create_one_document(driver, doc, i, len(docs))
                # Помечаем как обработанный сразу после успеха (до следующей итерации)
                key = _link_key(doc.get("link"))
                if key:
                    processed.add(key)
                    save_state(excel_path, processed)
                # Запись в output xlsx (даже если asud_id не захватили — ставим Link)
                _append_output_row(output_path, doc, asud_id)
                done_count += 1
            except Exception as e:
                log.error(f"ОШИБКА документ {i}: {e}")
                err_count += 1
                driver.get(url)
                wait_asud_loaded(driver)
                continue

        elapsed_seconds = time.monotonic() - start_time
        elapsed = timedelta(seconds=int(elapsed_seconds))
        avg = timedelta(seconds=int(elapsed_seconds / done_count)) if done_count else None

        summary = [
            "",
            "=" * 60,
            f"ГОТОВО!",
            f"  Обработано:   {done_count} / {len(docs)}",
            f"  Ошибок:       {err_count}",
            f"  В черновиках: {unknown}",
            f"  Затрачено:    {elapsed}" + (f"  (в среднем {avg}/док)" if avg else ""),
            "=" * 60,
        ]
        for line in summary:
            log.info(line)
            print(line)

        if unknown:
            log.warning(f"Проверьте {unknown} документов в черновиках — "
                        f"ФИО не извлечено автоматически")
        if err_count:
            log.warning(f"{err_count} документов упали с ошибкой. "
                        f"Перезапуск скрипта продолжит с них "
                        f"(уже обработанные запомнены в state-файле).")
        input("\nEnter для закрытия...")

    except Exception as e:
        log.error(f"Ошибка: {e}")
        input("Enter...")
    finally:
        driver.quit()
        log.info("Браузер закрыт")


if __name__ == "__main__":
    main()
