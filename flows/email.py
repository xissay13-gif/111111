"""
flows/email.py — Email-direct: создание Входящих документов прямо из .msg файлов
в указанной папке (без xlsx-реестра).

Два режима:
  • main()        — однопроходный: пробежать все .msg в корне папки и выйти
  • daemon_main() — непрерывный мониторинг: опрос папки раз в N сек,
                     обработка новых .msg, автоперемещение по результату
                     (Завершено / Ошибки / Черновики), Ctrl+C для остановки

Запуск через app.py с --mode=email (одноразовый) или --mode=email --watch (daemon).
"""

import os
import re
import signal
import sys
import time
import logging
from datetime import date, datetime, timedelta

import openpyxl
import extract_msg
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService

from shared import config as cfg
from shared.ui import wait_asud_loaded
from shared.correspondent import extract_fio_from_text
from shared.okrug_parser import okrug_from_textbody
from shared.attachments import move_to_done, move_to_errors, move_to_drafts
from shared.colors import green, yellow, red, status_colored
from shared.classifier import classify_doc_type

# Переиспользуем создание/регистрацию/output из mix
from flows import mix as mix_flow

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
    handlers=[logging.StreamHandler()],
)
log = logging.getLogger("asud")
start_time = time.monotonic()

settings = {}


# ================= EMAIL → DOC_DATA =================

# Имя .msg-файла начинается с даты-времени: '2026-05-06 10-58-43.msg'
_FILENAME_DATE_RE = re.compile(r'^(\d{4}-\d{2}-\d{2})')


def _msg_date_prefix(file_path):
    """Извлекает 'YYYY-MM-DD' из имени .msg. None если не распарсилось."""
    m = _FILENAME_DATE_RE.match(os.path.basename(file_path))
    return m.group(1) if m else None


def _msg_link(msg, file_path):
    """Генерирует Link для doc_data.

    Приоритет:
      1) Дата письма из .msg (msg.date)
      2) Имя файла без расширения
    Формат — как в mix-flow реестре: 'DD.MM.YYYY HH-MM-SS'.
    """
    try:
        if msg.date:
            return msg.date.strftime("%d.%m.%Y %H-%M-%S")
    except Exception:
        pass
    return os.path.splitext(os.path.basename(file_path))[0]


# ================= PER-DATE XLSX REGISTRY =================

# Колонки реестра (под per-date).
_REGISTRY_HEADERS = ["Номер", "Link", "Округ", "Subject", "Body"]
_REGISTRY_WIDTHS = {1: 18, 2: 22, 3: 8, 4: 50, 5: 80}


def _dated_xlsx_path(base_dir, date_prefix, suffix=None):
    """Путь к per-date реестру.
    Без suffix: Registered/YYYY-MM-DD_резолюции.xlsx
    С suffix:   Registered/YYYY-MM-DD_<suffix>_резолюции.xlsx
    Если date_prefix=None → _unknown_<...>.xlsx (fallback)."""
    registered_dir = os.path.join(base_dir, "Registered")
    os.makedirs(registered_dir, exist_ok=True)
    name = date_prefix or "_unknown"
    if suffix:
        name += f"_{suffix}"
    name += "_резолюции.xlsx"
    return os.path.join(registered_dir, name)


def _ensure_dated_xlsx(path):
    """Создаёт per-date xlsx с шапкой если не существует."""
    if os.path.isfile(path):
        return
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Резолюции"
        ws.append(_REGISTRY_HEADERS)
        for c in range(1, len(_REGISTRY_HEADERS) + 1):
            ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
        for col, w in _REGISTRY_WIDTHS.items():
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)].width = w
        ws.freeze_panes = "A2"
        wb.save(path)
    except Exception as e:
        log.warning(f"Не удалось создать {path}: {e}")


def _append_dated_row(path, doc, asud_id):
    """Дописывает строку в per-date xlsx.
    Колонки: Номер | Link | Округ | Subject | Body
    """
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        ws.append([
            asud_id or "",
            doc.get("link") or "",
            doc.get("округ_прогноз") or "",
            doc.get("тема") or "",
            doc.get("содержание") or "",  # уже _clean_body
        ])
        wb.save(path)
    except Exception as e:
        log.warning(f"Не удалось записать строку в {path}: {e}")


def _list_root_msgs(folder_path):
    """Возвращает sorted list абсолютных путей к .msg в корне folder_path
    (без рекурсии). Пустой list если папки нет / I/O ошибка."""
    try:
        out = []
        for f in os.listdir(folder_path):
            full = os.path.join(folder_path, f)
            if os.path.isfile(full) and f.lower().endswith('.msg'):
                out.append(full)
        return sorted(out)
    except OSError as e:
        log.error(f"Не могу прочитать папку {folder_path}: {e}")
        return []


def _parse_one_msg(msg_path):
    """Парсит один .msg в doc_data dict. None если не получилось/пустое/брак.
    Использует module-global settings."""
    unknown = settings.get("unknown_correspondent",
                            cfg.DEFAULTS["unknown_correspondent"])

    try:
        msg = extract_msg.openMsg(msg_path)
        subject = msg.subject or ""
        body = msg.body or ""
        link = _msg_link(msg, msg_path)
        try:
            msg.close()
        except Exception:
            pass
    except Exception as e:
        log.warning(f"Не удалось прочитать {os.path.basename(msg_path)}: {e}")
        return None

    if not body and not subject:
        log.warning(f"Пустое письмо {os.path.basename(msg_path)} — пропускаю")
        return None

    # Классификатор работает на ИСХОДНОМ body (до _clean_body) — так
    # сохраняется точная реплика правил начальника и доступен From-header
    # для fallback по домену.
    type_idx = classify_doc_type(body)
    force_draft = False
    if type_idx == -1:
        log.warning(f"{os.path.basename(msg_path)}: помечено классификатором "
                    f"как 'случайно отправил' — пропускаю")
        return None
    if type_idx == 0:
        # Классификатор не разобрался — тип 8 + всегда черновик
        type_idx = 8
        force_draft = True
        log.info(f"{os.path.basename(msg_path)}: тип не определился → 8, "
                 f"оставляю в черновике для ручной проверки")
    type_name = cfg.DOC_TYPE_MAP.get(
        type_idx, "Письма, заявления и жалобы граждан, акционеров")

    clean_subject = re.sub(r'^(FW:|RE:|Fwd:)\s*', '',
                            str(subject).strip(), flags=re.IGNORECASE)
    body_clean = mix_flow._clean_body(body) if body else clean_subject

    if force_draft:
        # cat-0 fallback: принудительно черновик с фикс. корреспондентом,
        # чтобы письмо точно ушло в DRAFT-ветку mix.create_one_document.
        corr_found = False
        correspondent = unknown
        fio_src = ""
    else:
        fio, fio_src = extract_fio_from_text(body_clean)
        correspondent = fio if fio else unknown
        corr_found = bool(fio)

    try:
        okrug = okrug_from_textbody(body_clean, base_dir_fn=cfg.get_base_dir)
    except Exception as e:
        log.warning(f"okrug_parser упал для {os.path.basename(msg_path)}: {e}")
        okrug = None

    return {
        "row_idx": 1,
        "содержание": body_clean,
        "корреспондент": correspondent,
        "корр_найден": corr_found,
        "корр_источник": fio_src,
        "тема": clean_subject,
        "тип_индекс": type_idx,
        "тип_название": type_name,
        "link": link,
        "файл": msg_path,
        "округ_прогноз": okrug,
        "msg_date_prefix": _msg_date_prefix(msg_path),
        "force_draft": force_draft,
    }


def load_emails(folder_path):
    """Парсит все .msg в корне folder_path. Возвращает list of doc dicts."""
    msg_files = _list_root_msgs(folder_path)
    log.info(f"Найдено .msg файлов: {len(msg_files)}")

    rows, skipped = [], 0
    for idx, msg_path in enumerate(msg_files, 1):
        doc = _parse_one_msg(msg_path)
        if doc is None:
            skipped += 1
            continue
        doc["row_idx"] = idx
        rows.append(doc)

    log.info(f"Загружено: {len(rows)} писем, пропущено: {skipped}")
    return rows


# ================= PROCESSING ONE DOC =================

def _print_doc_line(index, total, status, info=""):
    """Одна строка результата по документу — с цветным статусом в консоль.
    OK — зелёный, DRAFT — жёлтый, FAILED — красный, DUPLICATE — без цвета."""
    label = status_colored(status)
    suffix = f" — {info}" if info else ""
    print(f"  [{index}/{total}] {label}{suffix}")


def _process_doc(driver, doc, base_dir, folder, index, total, in_daemon,
                 process_mode="mix", output_suffix=None):
    """Обрабатывает один doc: create_one_document → ветвление по статусу
    → запись в xlsx + перенос .msg.

    process_mode:
      'mix'   — текущая логика: ФИО найдено → register, нет → DRAFT
      'smart' — всегда черновик: forсируем корр_найден=False + фикс. корреспондент;
                 DRAFT считается УСПЕХОМ → пишем в реестр и переносим в Завершено/

    output_suffix — суффикс в имени per-date xlsx (для разделения реестров
    при параллельных запусках двух пресетов).

    Возвращает финальный статус: 'OK' | 'DUPLICATE' | 'DRAFT' | 'FAILED'.
    in_daemon=True (mix-режим): DRAFT → перенос в Черновики/.
    """
    msg_path = doc.get("файл")

    if process_mode == "smart":
        # Smart-пресет: каждый .msg создаётся как черновик с фикс. корреспондентом
        doc["корр_найден"] = False
        doc["корреспондент"] = settings.get("unknown_correspondent",
                                             cfg.DEFAULTS["unknown_correspondent"])

    asud_id = mix_flow.create_one_document(driver, doc, index, total)
    status = mix_flow._last_result.get("status", "FAILED")

    if status == "OK":
        xlsx_path = _dated_xlsx_path(base_dir, doc.get("msg_date_prefix"),
                                      output_suffix)
        _ensure_dated_xlsx(xlsx_path)
        _append_dated_row(xlsx_path, doc, asud_id)
        move_to_done(msg_path, folder)
    elif status == "DUPLICATE":
        log.info(f"Документ {index}: уже зарегистрирован — .msg в Завершено/")
        move_to_done(msg_path, folder)
    elif status == "DRAFT":
        if process_mode == "smart":
            # Smart: черновик — это нормальный исход. Пишем в реестр (без АСУД-ID)
            # и переносим .msg в Завершено/.
            xlsx_path = _dated_xlsx_path(base_dir, doc.get("msg_date_prefix"),
                                          output_suffix)
            _ensure_dated_xlsx(xlsx_path)
            _append_dated_row(xlsx_path, doc, asud_id or "")
            move_to_done(msg_path, folder)
            log.info(f"Документ {index}: создан как черновик (smart) — .msg в Завершено/")
        elif in_daemon:
            log.info(f"Документ {index}: ФИО не найдено — .msg в Черновики/")
            move_to_drafts(msg_path, folder)
        # one-shot mix: оставляем в корне как и было
    else:  # FAILED — caller сам решает что делать (retry / move-to-errors)
        pass

    return status


# ================= MAIN =================

def main():
    global settings
    settings = cfg.load()
    cfg.setup_file_logger("email")
    cfg.keep_system_awake(True)

    log.info("=" * 50)
    log.info("АСУД ИК — Email-direct (создание из .msg-писем)")
    log.info("=" * 50)

    base_dir = cfg.get_base_dir()

    # Запрос пути к папке с .msg
    folder = os.environ.get('ASUD_EMAIL_FOLDER')
    if not folder:
        default = settings.get("email_folder", "")
        print(f"\nПапка с .msg-письмами (только из корня папки, подпапки игнорируются).")
        if default:
            print(f"Enter — использовать: {default}")
        user_dir = input("Путь: ").strip().strip('"').strip("'")
        folder = user_dir or default

    if not folder or not os.path.isdir(folder):
        log.error(f"Папка не найдена: {folder!r}")
        input("Enter...")
        sys.exit(1)

    log.info(f"Папка писем: {folder}")

    # Парсим письма
    docs = load_emails(folder)
    if not docs:
        log.error("Нет .msg файлов или все пропущены")
        input("Enter...")
        sys.exit(1)

    # Превью
    known = sum(1 for d in docs if d["корр_найден"])
    unknown_n = len(docs) - known
    print(f"\nПервые 5:")
    for i, d in enumerate(docs[:5], 1):
        flag = 'OK' if d["корр_найден"] else '!!'
        print(f"  {i}. [{d['тип_индекс']}] {flag} {d['корреспондент'][:30]} | {d['тема'][:50]}")
    print(f"\nВсего: {len(docs)}  (ФИО: {known}, заглушка: {unknown_n})")
    print("режим: EMAIL  —  создание + регистрация + На резолюцию + сам .msg как вложение")

    if input("Начать? (да/нет): ").strip().lower() not in ("да", "д", "y", "yes", ""):
        sys.exit(0)

    # === Запуск браузера и обработки (повторяем mix-loop, но с нашими docs)
    driver_path = os.path.join(base_dir, "msedgedriver.exe")
    if not os.path.exists(driver_path):
        log.error(f"msedgedriver.exe не найден в {base_dir}")
        input("Enter...")
        sys.exit(1)

    options = cfg.build_edge_options()
    service = EdgeService(executable_path=driver_path)
    driver = webdriver.Edge(service=service, options=options)

    # Настраиваем mix_flow.settings — он использует module-level global
    mix_flow.settings = settings

    process_mode = os.environ.get('ASUD_EMAIL_PROCESS_MODE', 'mix')
    output_suffix = os.environ.get('ASUD_OUTPUT_SUFFIX') or None
    log.info(f"Логика обработки: {process_mode}"
             + (f", суффикс реестра: {output_suffix}" if output_suffix else ""))

    try:
        url = settings.get("asud_url", cfg.DEFAULTS["asud_url"])
        log.info(f"Открываю {url}")
        driver.get(url)
        wait_asud_loaded(driver)

        # Per-date реестры: Registered/YYYY-MM-DD[_<suffix>]_резолюции.xlsx.
        # Каждый doc пишется в xlsx своей даты (из имени .msg).
        log.info(f"Per-date реестры в: {os.path.join(base_dir, 'Registered')}")

        done_count, dup_count, draft_count, err_count = 0, 0, 0, 0
        for i, doc in enumerate(docs, 1):
            msg_path = doc.get("файл")
            try:
                status = _process_doc(driver, doc, base_dir, folder,
                                       i, len(docs), in_daemon=False,
                                       process_mode=process_mode,
                                       output_suffix=output_suffix)
                if status == "OK":
                    done_count += 1
                elif status == "DUPLICATE":
                    dup_count += 1
                elif status == "DRAFT":
                    draft_count += 1
                else:  # FAILED
                    move_to_errors(msg_path, folder,
                                   f"Регистрация не удалась (status={status})")
                    err_count += 1
                _print_doc_line(i, len(docs), status,
                                 doc.get("тема", "")[:60])
            except Exception as e:
                log.error(f"ОШИБКА документ {i}: {e}")
                move_to_errors(msg_path, folder, f"Exception: {e}")
                err_count += 1
                _print_doc_line(i, len(docs), "FAILED", str(e)[:80])
                try:
                    driver.get(url)
                    wait_asud_loaded(driver)
                except Exception:
                    pass
                continue

        elapsed_seconds = time.monotonic() - start_time
        elapsed = timedelta(seconds=int(elapsed_seconds))
        avg = (timedelta(seconds=int(elapsed_seconds / done_count))
               if done_count else None)
        # Лог-файл — плоский (без ANSI), консоль — цветной
        plain = [
            "",
            "=" * 60,
            "ГОТОВО!",
            f"  Обработано:   {done_count} / {len(docs)}  (→ Завершено/)",
            f"  Дубликаты:    {dup_count}  (уже были в АСУД, → Завершено/)",
            f"  В черновиках: {draft_count}  (ФИО не найдено, .msg остался в корне)",
            f"  Ошибок:       {err_count}  (→ Ошибки/)",
            f"  Затрачено:    {elapsed}" + (f"  (в среднем {avg}/док)" if avg else ""),
            "=" * 60,
        ]
        for line in plain:
            log.info(line)
        print("")
        print("=" * 60)
        print("ГОТОВО!")
        print(f"  Обработано:   {green(str(done_count))} / {len(docs)}  (→ Завершено/)")
        print(f"  Дубликаты:    {dup_count}  (уже были в АСУД, → Завершено/)")
        print(f"  В черновиках: {yellow(str(draft_count))}  (ФИО не найдено, .msg в корне)")
        print(f"  Ошибок:       {red(str(err_count))}  (→ Ошибки/)")
        print(f"  Затрачено:    {elapsed}" + (f"  (в среднем {avg}/док)" if avg else ""))
        print("=" * 60)
        input("\nEnter для закрытия...")
    except Exception as e:
        log.error(f"Ошибка: {e}")
        input("Enter...")
    finally:
        try:
            driver.quit()
        except Exception:
            pass
        cfg.keep_system_awake(False)


# ================= DAEMON MODE =================

# Sigint handler — устанавливает флаг, текущий документ доделывается, потом выход.
_stop_flag = False


def _on_sigint(signum, frame):
    global _stop_flag
    if _stop_flag:
        log.warning("Повторный Ctrl+C — выход немедленно")
        sys.exit(130)
    _stop_flag = True
    log.info("Ctrl+C получен — закончу текущий документ и выйду из мониторинга")


def _interruptible_sleep(seconds):
    """Sleep с пробуждением по _stop_flag (проверка раз в секунду)."""
    for _ in range(int(seconds)):
        if _stop_flag:
            return
        time.sleep(1)


def daemon_main():
    """Непрерывный мониторинг папки: опрос раз в N сек, обработка новых .msg.
    Ctrl+C для остановки (graceful — после текущего документа)."""
    global settings
    settings = cfg.load()
    cfg.setup_file_logger("email_daemon")
    cfg.keep_system_awake(True)

    log.info("=" * 50)
    log.info("АСУД ИК — Email-DAEMON (непрерывный мониторинг)")
    log.info("=" * 50)

    base_dir = cfg.get_base_dir()
    interval = int(settings.get("email_watch_interval_sec",
                                cfg.DEFAULTS["email_watch_interval_sec"]))
    max_retries = int(settings.get("email_max_retries",
                                    cfg.DEFAULTS["email_max_retries"]))
    process_mode = os.environ.get('ASUD_EMAIL_PROCESS_MODE', 'mix')
    output_suffix = os.environ.get('ASUD_OUTPUT_SUFFIX') or None
    log.info(f"Логика обработки: {process_mode}"
             + (f", суффикс реестра: {output_suffix}" if output_suffix else ""))

    # Папка
    folder = os.environ.get('ASUD_EMAIL_FOLDER')
    if not folder:
        default = settings.get("email_folder", "")
        print(f"\nПапка с .msg-письмами для непрерывного мониторинга.")
        if default:
            print(f"Enter — использовать: {default}")
        user_dir = input("Путь: ").strip().strip('"').strip("'")
        folder = user_dir or default
    if not folder or not os.path.isdir(folder):
        log.error(f"Папка не найдена: {folder!r}")
        input("Enter...")
        sys.exit(1)

    log.info(f"Папка: {folder}")
    log.info(f"Опрос: каждые {interval} сек, макс retry: {max_retries}")
    print(f"\nМониторинг включён. Ctrl+C для остановки.")

    # Browser
    driver_path = os.path.join(base_dir, "msedgedriver.exe")
    if not os.path.exists(driver_path):
        log.error(f"msedgedriver.exe не найден в {base_dir}")
        input("Enter...")
        sys.exit(1)

    options = cfg.build_edge_options()
    service = EdgeService(executable_path=driver_path)
    driver = webdriver.Edge(service=service, options=options)
    mix_flow.settings = settings

    signal.signal(signal.SIGINT, _on_sigint)

    url = settings.get("asud_url", cfg.DEFAULTS["asud_url"])
    log.info(f"Открываю {url}")
    driver.get(url)
    wait_asud_loaded(driver)

    # Счётчики и retry-state
    retry_count = {}  # basename → int (фейлов подряд)
    totals = {"OK": 0, "DUPLICATE": 0, "DRAFT": 0, "FAILED": 0, "ITER": 0}

    try:
        while not _stop_flag:
            totals["ITER"] += 1
            queue = _list_root_msgs(folder)
            if not queue:
                log.info(f"[итер. {totals['ITER']}] очередь пуста — sleep {interval}s")
                _interruptible_sleep(interval)
                continue

            log.info(f"[итер. {totals['ITER']}] в очереди: {len(queue)}")
            for idx, msg_path in enumerate(queue, 1):
                if _stop_flag:
                    break
                name = os.path.basename(msg_path)

                doc = _parse_one_msg(msg_path)
                if doc is None:
                    # битый/пустой .msg — сразу в Ошибки чтобы не зацикливаться
                    move_to_errors(msg_path, folder,
                                   "Не удалось распарсить или пустое")
                    totals["FAILED"] += 1
                    retry_count.pop(name, None)
                    _print_doc_line(idx, len(queue), "FAILED",
                                     "не распарсилось / пустое")
                    continue

                try:
                    status = _process_doc(driver, doc, base_dir, folder,
                                           idx, len(queue), in_daemon=True,
                                           process_mode=process_mode,
                                           output_suffix=output_suffix)
                    if status == "FAILED":
                        retry_count[name] = retry_count.get(name, 0) + 1
                        if retry_count[name] >= max_retries:
                            move_to_errors(msg_path, folder,
                                f"Регистрация не удалась за {max_retries} попыток")
                            retry_count.pop(name, None)
                            totals["FAILED"] += 1
                            _print_doc_line(idx, len(queue), "FAILED",
                                             f"max_retries ({max_retries}) → Ошибки/")
                        else:
                            log.warning(f"{name}: фейл {retry_count[name]}/{max_retries} "
                                        f"— оставляю в корне на следующую итерацию")
                            _print_doc_line(idx, len(queue), "FAILED",
                                             f"retry {retry_count[name]}/{max_retries}")
                        try:
                            driver.get(url)
                            wait_asud_loaded(driver)
                        except Exception:
                            pass
                    else:
                        totals[status] = totals.get(status, 0) + 1
                        retry_count.pop(name, None)
                        _print_doc_line(idx, len(queue), status,
                                         doc.get("тема", "")[:60])
                except Exception as e:
                    log.error(f"Exception на {name}: {e}")
                    retry_count[name] = retry_count.get(name, 0) + 1
                    if retry_count[name] >= max_retries:
                        move_to_errors(msg_path, folder, f"Exception: {e}")
                        retry_count.pop(name, None)
                        totals["FAILED"] += 1
                    _print_doc_line(idx, len(queue), "FAILED", str(e)[:80])
                    try:
                        driver.get(url)
                        wait_asud_loaded(driver)
                    except Exception:
                        pass

            if _stop_flag:
                break
            log.info(f"  итог итер. {totals['ITER']}: "
                     f"OK={totals['OK']} DUP={totals['DUPLICATE']} "
                     f"DRAFT={totals['DRAFT']} FAIL={totals['FAILED']}")
            print(f"  итог итер. {totals['ITER']}: "
                  f"OK={green(str(totals['OK']))} DUP={totals['DUPLICATE']} "
                  f"DRAFT={yellow(str(totals['DRAFT']))} "
                  f"FAIL={red(str(totals['FAILED']))}")
            _interruptible_sleep(interval)

        log.info("=" * 60)
        log.info("МОНИТОРИНГ ОСТАНОВЛЕН")
        log.info(f"  Итераций:   {totals['ITER']}")
        log.info(f"  Обработано: {totals['OK']}")
        log.info(f"  Дубликаты:  {totals['DUPLICATE']}")
        log.info(f"  Черновики:  {totals['DRAFT']}")
        log.info(f"  Ошибки:     {totals['FAILED']}")
        log.info("=" * 60)
        print("=" * 60)
        print("МОНИТОРИНГ ОСТАНОВЛЕН")
        print(f"  Итераций:   {totals['ITER']}")
        print(f"  Обработано: {green(str(totals['OK']))}")
        print(f"  Дубликаты:  {totals['DUPLICATE']}")
        print(f"  Черновики:  {yellow(str(totals['DRAFT']))}")
        print(f"  Ошибки:     {red(str(totals['FAILED']))}")
        print("=" * 60)

    finally:
        try:
            driver.quit()
        except Exception:
            pass
        cfg.keep_system_awake(False)


if __name__ == "__main__":
    if os.environ.get("ASUD_WATCH") == "1":
        daemon_main()
    else:
        main()
