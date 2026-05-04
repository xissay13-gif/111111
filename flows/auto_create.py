"""
auto_create_correspondent.py — Пакетное создание Входящих документов в АСУД ИК.

Читает Excel (B=Содержание, C=Корреспондент), автоматически создаёт
корреспондента если не найден. Регистрирует + На резолюцию.

Модули:
  config.py       — настройки (+ config.json)
  ui.py           — Selenium UI-хелперы
  correspondent.py — создание корреспондентов
  attachments.py  — прикрепление файлов (пустышка)
"""

import os
import re
import sys
import time
import logging
from datetime import date, datetime, timedelta

import openpyxl
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from shared import config as cfg
from shared.ui import (click, wait_and_click, find_input_near_label,
                wait_asud_loaded, wait_modal_closed, close_open_modals,
                js_set_value, js_type_combobox, find_dropdown_options)
from shared.correspondent import (fill_correspondent_field, match_correspondent)
from shared.attachments import get_dummy_msg, attach_content


# ================= LOGGING =================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
)
log = logging.getLogger("asud")
start_time = time.monotonic()


# ================= EXCEL =================

def load_excel(file_path):
    """Читает реестр обращений.

    Поддерживаемые форматы (определяются по заголовкам):
      • НОВЫЙ: лист 'результат', колонки № | LS | фио | обращение | ao | fio | тема
        — фио (cyr) = корреспондент, fio (lat) = исполнитель, тема = индекс из DOC_TYPE_MAP
      • СТАРЫЙ: B = содержание, C = корреспондент (первый лист)
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    # Лист 'результат' (новый) или active (старый)
    ws = wb['результат'] if 'результат' in wb.sheetnames else wb.active
    log.info(f"Лист: '{ws.title}'")

    headers = [str(c.value or '').strip() for c in next(ws.iter_rows(max_row=1))]
    log.info(f"Заголовки: {headers}")

    def find_idx(*names):
        for name in names:
            for i, h in enumerate(headers):
                if h.lower() == name.lower():
                    return i
        return None

    # Новый формат: 'фио' (cyr) дважды — первое корреспондент, второе исполнитель
    fio_indices = [i for i, h in enumerate(headers) if h.lower() == 'фио']
    idx_content = find_idx('обращение', 'содержание', 'textbody')
    idx_corr = fio_indices[0] if fio_indices else find_idx('корреспондент')
    idx_executor = find_idx('fio')  # латиница
    idx_ao = find_idx('ao', 'округ')
    idx_type = find_idx('тема', 'тип')

    # Старый формат fallback
    if idx_content is None:
        idx_content = 1  # B
    if idx_corr is None:
        idx_corr = 2  # C

    log.info(f"Колонки: содержание={idx_content}, корреспондент={idx_corr}, "
             f"ao={idx_ao}, исполнитель={idx_executor}, тема={idx_type}")

    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row:
            continue
        content = row[idx_content] if len(row) > idx_content else None
        corr = row[idx_corr] if len(row) > idx_corr else None
        if not (content and corr):
            continue
        item = {
            "содержание": str(content).strip(),
            "корреспондент": str(corr).strip(),
        }
        if idx_ao is not None and len(row) > idx_ao and row[idx_ao]:
            item["ao"] = str(row[idx_ao]).strip()
        if idx_executor is not None and len(row) > idx_executor and row[idx_executor]:
            item["исполнитель"] = str(row[idx_executor]).strip()
        if idx_type is not None and len(row) > idx_type and row[idx_type] is not None:
            try:
                item["тема_индекс"] = int(row[idx_type])
            except Exception:
                pass
        rows.append(item)
    wb.close()
    log.info(f"Загружено: {len(rows)} документов")
    return rows


# ================= FORM FILLING =================

def fill_text(driver, text):
    """Заполняет краткое содержание (textarea) — JS-set, plain-поле."""
    try:
        areas = driver.find_elements(By.TAG_NAME, "textarea")
        visible = [a for a in areas if a.is_displayed()]
        if visible:
            js_set_value(driver, visible[0], text)
            log.info("Краткое содержание заполнено (JS)")
        else:
            log.warning("Textarea не найдена")
    except Exception as e:
        log.error(f"Ошибка содержания: {e}")


def fill_corr_number(driver, index):
    """Заполняет 'Номер у корреспондента' = 'б/н (N)' — JS-set."""
    value = f"б/н ({index})"
    inp = find_input_near_label(driver, "Номер у корреспондента")
    if not inp:
        log.warning("Поле 'Номер у корреспондента' не найдено")
        return
    try:
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", inp)
        js_set_value(driver, inp, value)
        log.info(f"Номер (JS): {value}")
    except Exception as e:
        log.warning(f"Номер: ошибка {e}")


def fill_corr_date(driver):
    """Заполняет 'Дата у корреспондента' = сегодня."""
    from datetime import date
    today = date.today().strftime("%d.%m.%Y")
    labels = driver.find_elements(By.XPATH,
        "//*[normalize-space(text())='Дата у корреспондента']")
    inp = None
    for label in labels:
        try:
            if not label.is_displayed():
                continue
            for level in range(1, 6):
                parent = label
                for _ in range(level):
                    parent = parent.find_element(By.XPATH, "..")
                inputs = parent.find_elements(By.CSS_SELECTOR, "input[type='text']")
                visible = [i for i in inputs
                           if i.is_displayed() and i.get_attribute("readonly") is None]
                if visible:
                    inp = visible[0]
                    break
            if inp:
                break
        except Exception:
            continue
    if not inp:
        log.warning("Поле 'Дата у корреспондента' не найдено")
        return
    try:
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", inp)
        js_set_value(driver, inp, today)
        log.info(f"Дата (JS): {today}")
    except Exception as e:
        log.warning(f"Дата: ошибка {e}")


def fill_delivery_method(driver):
    """Выбирает 'Электронная почта'."""
    target = settings.get("delivery_method", "Электронная почта")
    labels = driver.find_elements(By.XPATH,
        "//*[normalize-space(text())='Способ получения']")
    trigger = None
    for label in labels:
        try:
            if not label.is_displayed():
                continue
            for level in range(1, 8):
                parent = label
                for _ in range(level):
                    parent = parent.find_element(By.XPATH, "..")
                for sel in ["input[type='text']", "div[class*='trigger']", "img[class*='trigger']"]:
                    try:
                        el = parent.find_element(By.CSS_SELECTOR, sel)
                        if el.is_displayed():
                            trigger = el
                            break
                    except Exception:
                        continue
                if trigger:
                    break
            if trigger:
                break
        except Exception:
            continue
    if not trigger:
        log.warning("Поле 'Способ получения' не найдено")
        return
    click(driver, trigger, "Способ получения")

    # Ждём появления нужного пункта в дропдауне (вместо sleep(1.5)+цикла)
    def _option_visible(d):
        for c in d.find_elements(By.XPATH, f"//*[contains(text(),'{target}')]"):
            try:
                if c.is_displayed() and c.tag_name.lower() != 'input':
                    return c
            except Exception:
                continue
        return False
    try:
        opt = WebDriverWait(driver, 5).until(_option_visible)
        click(driver, opt, target)
        log.info(f"Способ получения: {target}")
    except Exception:
        log.warning(f"'{target}' не найдена в выпадашке")


def add_addressee(driver, person_name):
    """Добавляет адресата через combobox (JS-typing + WebDriverWait)."""
    inp = find_input_near_label(driver, "Адресаты")
    if not inp:
        log.warning("Поле адресата не найдено")
        return
    surname = person_name.split()[0]
    inp.click()
    js_type_combobox(driver, inp, surname)

    all_r = []
    try:
        WebDriverWait(driver, 5).until(
            lambda d: len(find_dropdown_options(d, surname, inp)) > 0)
        all_r = find_dropdown_options(driver, surname, inp)
    except Exception:
        try:
            inp.send_keys(Keys.ENTER)
            WebDriverWait(driver, 3).until(
                lambda d: len(find_dropdown_options(d, surname, inp)) > 0)
            all_r = find_dropdown_options(driver, surname, inp)
        except Exception:
            pass

    target = None
    for r in all_r:
        try:
            if match_correspondent(r.text, person_name):
                target = r
                break
        except Exception:
            continue
    if not target and all_r:
        target = all_r[0]
    if target:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", target)
        ActionChains(driver).move_to_element(target).pause(0.2).click().perform()
        log.info(f"Адресат: {person_name}")
    else:
        log.warning(f"Адресат не найден: {person_name}")


# ================= REGISTRATION =================

_CAPTURE_ASUD_ID_JS = r"""
// Один проход по DOM на стороне браузера. Возвращает регистрационный
// номер или null. Строго: только из ScreenHeader1, плюс body-regex
// как last-resort. Не сканируем "любой <b>" чтобы случайно не подобрать
// другой id-подобный элемент на странице (например, id типа служебки).
const RE = /\b([А-Я]{2,5}(?:\/[А-Я0-9.\-]+){2,})\b/u;
function looksLike(t) {
    if (!t) return false;
    t = t.trim();
    if (!t.includes('/') || t.length < 6) return false;
    // НЕ дата вида "01.05.2026"
    if (/^\d{2}\.\d{2}\.\d{4}/.test(t)) return false;
    // Должна быть хотя бы одна цифра — отсекаем чисто-буквенные совпадения
    // (например, если бы где-то нашлось "АБВ/ГД/ЕЖ")
    if (!/\d/.test(t)) return false;
    return true;
}
// Строгий путь: только ScreenHeader1
const header = document.querySelector("[data-marker='ScreenHeader1']");
if (header) {
    for (const b of header.querySelectorAll('b')) {
        const t = (b.textContent || '').trim();
        if (looksLike(t)) return t;
    }
    // Иногда номер не в <b>, а в самом тексте header'а — regex по нему
    const m = (header.textContent || '').match(RE);
    if (m && looksLike(m[1])) return m[1];
}
// Last resort: regex по всему телу страницы
const m2 = (document.body.innerText || '').match(RE);
if (m2 && looksLike(m2[1])) return m2[1];
return null;
"""


def capture_asud_id(driver, timeout=15):
    """Читает регистрационный номер документа после регистрации.
    Один JS-вызов на итерацию (быстрый поллинг 100ms).
    Строгий поиск ТОЛЬКО в ScreenHeader1, чтобы не зацепить тип служебки
    или другой id-подобный текст на странице."""
    end = time.monotonic() + timeout
    while time.monotonic() < end:
        try:
            asud_id = driver.execute_script(_CAPTURE_ASUD_ID_JS)
            if asud_id:
                log.info(f"  asud_id: {asud_id!r}")
                return asud_id
        except Exception:
            pass
        time.sleep(0.1)
    log.warning("Регистрационный номер не захватили — пуст в output")
    return None


def _wait_button_enabled(driver, css_selector, timeout=20):
    """Ждёт пока кнопка станет clickable (data-disabled != '1', visible).
    После attach АСУД на 1-3s держит register-кнопку в disabled пока обрабатывает upload."""
    end = time.monotonic() + timeout
    last_state = None
    while time.monotonic() < end:
        try:
            btn = driver.find_element(By.CSS_SELECTOR, css_selector)
            disabled = btn.get_attribute('data-disabled')
            if btn.is_displayed() and disabled != '1':
                return btn
            if last_state != disabled:
                log.debug(f"  кнопка {css_selector} disabled={disabled}, жду...")
                last_state = disabled
        except Exception as e:
            if last_state != 'missing':
                log.debug(f"  кнопка {css_selector} ещё не в DOM ({e})")
                last_state = 'missing'
        time.sleep(0.3)
    return None


def register_and_resolve(driver, index, total):
    """Регистрирует + На резолюцию + Да. Возвращает asud_id (или None)."""
    log.info("Регистрирую...")
    registered = False
    asud_id = None
    try:
        # Ждём именно ЕНАБЛЕНУЮ кнопку — после attach она может быть
        # data-disabled='1' пока АСУД обрабатывает upload файла
        btn = _wait_button_enabled(driver,
            "#header-action-btn-register, [id*='header-action-btn-register']",
            timeout=cfg.DEFAULTS["timeout"])
        if not btn:
            raise Exception("'Зарегистрировать' не активировалась за timeout")
        click(driver, btn, "Зарегистрировать")
        registered = True
    except Exception:
        try:
            btn = driver.find_element(By.XPATH, "//div[contains(text(),'Зарегистрировать')]")
            click(driver, btn, "Зарегистрировать (fallback)")
            registered = True
        except Exception as e:
            log.error(f"'Зарегистрировать' не найдена: {e}")

    if not registered:
        return None

    # Ждём появления "На резолюцию" — это маркер что регистрация прошла и DOM
    # полностью обновился. asud_id появляется одновременно — захватим после.
    res_btn = None
    try:
        res_btn = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.ID, "header-action-btn-send_on_resolution")))
    except Exception:
        # Fallback по тексту
        try:
            for b in driver.find_elements(By.XPATH, "//*[contains(text(),'На резолюцию')]"):
                if b.is_displayed():
                    res_btn = b
                    break
        except Exception:
            pass

    # Сейчас DOM готов — asud_id ловится за 100-200ms
    asud_id = capture_asud_id(driver, timeout=3)
    if asud_id:
        log.info(f"Документ {index}/{total} ЗАРЕГИСТРИРОВАН: {asud_id}")
    else:
        log.warning(f"Документ {index}/{total} ЗАРЕГИСТРИРОВАН (номер не захватили)")

    if res_btn:
        click(driver, res_btn, "На резолюцию")

        # Ждём confirm-кнопку "Да" одним WebDriverWait — три селектора в lambda,
        # поллинг 500ms (vs прежний цикл с sleep(1))
        def _find_yes(d):
            for sel in [(By.ID, "confirm_dialog_btn_yes"),
                        (By.CSS_SELECTOR, "[id*='confirm_dialog_btn_yes'], [id*='confirm'][id*='yes']")]:
                try:
                    b = d.find_element(*sel)
                    if b.is_displayed():
                        return b
                except Exception:
                    continue
            for b in d.find_elements(By.XPATH, "//*[normalize-space(text())='Да']"):
                try:
                    if b.is_displayed():
                        return b
                except Exception:
                    continue
            return False
        try:
            yes_btn = WebDriverWait(driver, 10).until(_find_yes)
        except Exception:
            yes_btn = None

        if yes_btn:
            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", yes_btn)
                time.sleep(0.3)
            except Exception:
                pass
            clicked = False
            # 1) ActionChains
            try:
                ActionChains(driver).move_to_element(yes_btn).pause(0.3).click().perform()
                log.info(f"Клик 'Да' (ActionChains): id={yes_btn.get_attribute('id')}")
                clicked = True
            except Exception as e:
                log.info(f"ActionChains 'Да' не сработал: {e}")
            # 2) JS
            if not clicked:
                try:
                    driver.execute_script("arguments[0].click();", yes_btn)
                    log.info("Клик 'Да' (JS)")
                    clicked = True
                except Exception as e:
                    log.info(f"JS 'Да' не сработал: {e}")
            # 3) Нативный
            if not clicked:
                try:
                    yes_btn.click()
                    log.info("Клик 'Да' (native)")
                    clicked = True
                except Exception as e:
                    log.info(f"Native 'Да' не сработал: {e}")
            # 4) Enter — GWT-диалоги часто принимают его
            if clicked:
                try:
                    ActionChains(driver).send_keys(Keys.ENTER).perform()
                except Exception:
                    pass
            # Ждём пока модалка закроется (диалог "Да" уйдёт из DOM)
            try:
                WebDriverWait(driver, 5).until_not(EC.visibility_of(yes_btn))
            except Exception:
                pass
            log.info(f"Документ {index}/{total} НА РЕЗОЛЮЦИИ")
        else:
            log.warning("Диалог 'Да' не появился за 10 сек")
    else:
        log.warning("'На резолюцию' не появилась")
    return asud_id


# ================= DOCUMENT FLOW =================

def create_one_document(driver, doc_data, index, total):
    """Создаёт один входящий документ."""
    log.info(f"{'='*50}")
    log.info(f"ДОКУМЕНТ {index}/{total}")
    log.info(f"Содержание: {doc_data['содержание'][:60]}...")
    log.info(f"Корреспондент: {doc_data['корреспондент']}")

    el = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
        EC.presence_of_element_located((By.ID, "mainscreen-create-button")))
    click(driver, el, "Создать документ")

    wait_and_click(driver, By.XPATH,
        "//div[contains(text(),'Входящий документ')]", "Входящий документ")

    # Тип/вид документа: если в реестре есть 'тема' (индекс) — используем
    # DOC_TYPE_MAP, иначе берём дефолт из настроек
    type_idx = doc_data.get("тема_индекс")
    if type_idx and type_idx in cfg.DOC_TYPE_MAP:
        subtype = cfg.DOC_TYPE_MAP[type_idx]
        log.info(f"Тип из реестра: {type_idx} → {subtype}")
    else:
        subtype = settings.get("doc_subtype", "Письма, заявления и жалобы граждан, акционеров")
    short = subtype[:30]
    wait_and_click(driver, By.XPATH,
        f"//div[contains(text(),'{short}')] | //td[contains(text(),'{short}')]", subtype)

    wait_and_click(driver, By.XPATH,
        "//button[contains(text(),'Создать документ')] | //div[contains(text(),'Создать документ')]",
        "Создать документ")

    # Ждём пока форма документа отрендерится — появится textarea (краткое содержание)
    try:
        WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            lambda d: any(t.is_displayed() for t in d.find_elements(By.TAG_NAME, "textarea")))
    except Exception:
        log.warning("Textarea формы не появилась")

    fill_text(driver, doc_data["содержание"])
    fill_correspondent_field(driver, doc_data["корреспондент"])
    fill_corr_number(driver, index)
    fill_corr_date(driver)

    for person in settings.get("addressees", ["Басманов Александр Владимирович"]):
        add_addressee(driver, person)

    fill_delivery_method(driver)

    try:
        save_btn = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            EC.element_to_be_clickable((By.ID, "header-save-btn")))
        click(driver, save_btn, "Сохранить")
        # Ждём появления "Зарегистрировать" — признак что save прошёл
        try:
            WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
                EC.presence_of_element_located((By.CSS_SELECTOR,
                    "#header-action-btn-register, [id*='header-action-btn-register']")))
        except Exception:
            log.warning("После Сохранить кнопка 'Зарегистрировать' не появилась")
        log.info(f"Документ {index}/{total} сохранён")
    except Exception as e:
        log.error(f"Ошибка сохранения: {e}")

    if doc_data.get("файл"):
        attach_content(driver, doc_data["файл"])
        wait_modal_closed(driver)

    asud_id = register_and_resolve(driver, index, total)

    # Ждём что появится первым: либо главный экран (карточка авто-закрылась),
    # либо close-btn (карточка ещё открыта). До 10s, но обычно <1s.
    def _state(d):
        try:
            main_btn = d.find_element(By.ID, "mainscreen-create-button")
            if main_btn.is_displayed() and main_btn.is_enabled():
                return ('main', main_btn)
        except Exception:
            pass
        try:
            close_btn = d.find_element(By.ID, "header-close-btn")
            if close_btn.is_displayed():
                return ('close', close_btn)
        except Exception:
            pass
        return False

    state = None
    try:
        state = WebDriverWait(driver, 10).until(_state)
    except Exception:
        pass

    if state and state[0] == 'close':
        try:
            ActionChains(driver).move_to_element(state[1]).pause(0.2).click().perform()
        except Exception:
            pass
        # После закрытия ждём главный экран
        try:
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "mainscreen-create-button")))
        except Exception:
            driver.get(settings.get("asud_url", cfg.DEFAULTS["asud_url"]))
            wait_asud_loaded(driver)
    elif not state:
        # Не нашли ни одной из кнопок — перезагружаемся
        driver.get(settings.get("asud_url", cfg.DEFAULTS["asud_url"]))
        wait_asud_loaded(driver)
    return asud_id


# ================= OUTPUT XLSX (для clean-resolutions) =================

OKRUG_TO_FIO = {
    "САО": "Гренц Екатерина Александровна",
    "ЦАО": "Емельянова Татьяна Николаевна",
    "ОАО": "Рендюк Юлия Павловна",
    "ЛАО": "Вырва Елена Анатольевна",
    "КАО": "Кравец Татьяна Александровна",
}


def _output_xlsx_path(excel_path):
    base, _ext = os.path.splitext(excel_path)
    return base + "_резолюции.xlsx"


def _ensure_output_xlsx(path):
    if os.path.isfile(path):
        return
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Резолюции"
        ws.append(["ОПТС", "Округ", "ФИО", "Link"])
        for c in range(1, 5):
            ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
        widths = {1: 30, 2: 8, 3: 35, 4: 22}
        for col, w in widths.items():
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)].width = w
        ws.freeze_panes = "A2"
        wb.save(path)
    except Exception as e:
        log.warning(f"Не удалось создать {path}: {e}")


def _append_output_row(path, doc_data, asud_id):
    """Дописывает в output xlsx: ОПТС | Округ | ФИО | Link.

    Округ/ФИО берутся из реестра (новый формат), при отсутствии —
    парсятся из 'Содержание' и/или маппятся через OKRUG_TO_FIO.
    """
    okrug = doc_data.get("ao") or None
    if not okrug:
        try:
            from shared.okrug_parser import okrug_from_textbody
            okrug = okrug_from_textbody(doc_data.get("содержание"),
                                         base_dir_fn=cfg.get_base_dir)
        except Exception as e:
            log.warning(f"okrug_parser упал: {e}")
    fio = doc_data.get("исполнитель") or (OKRUG_TO_FIO.get(okrug) if okrug else None)
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        ws.append([asud_id or "", okrug or "", fio or "", ""])
        wb.save(path)
        wb.close()
        log.info(f"  → {os.path.basename(path)}: "
                 f"{asud_id or '—'} | {okrug or '—'} | {fio or '—'}")
    except Exception as e:
        log.warning(f"Не удалось записать в {path}: {e}")


# ================= MAIN =================

settings = {}


def main():
    global settings
    settings = cfg.load()

    log.info("=" * 50)
    log.info("АСУД ИК — Входящие + автосоздание корреспондентов")
    log.info("=" * 50)

    base_dir = cfg.get_base_dir()

    # Если xlsx уже выбран извне (через app.py) — используем его
    excel_path = os.environ.get('ASUD_XLSX')
    if not excel_path:
        xlsx_files = [f for f in os.listdir(base_dir) if f.lower().endswith('.xlsx')]
        if not xlsx_files:
            log.error(f"Нет .xlsx в {base_dir}")
            input("Enter...")
            sys.exit(1)
        elif len(xlsx_files) == 1:
            excel_path = os.path.join(base_dir, xlsx_files[0])
        else:
            print(f"\nНайдено {len(xlsx_files)} xlsx:")
            for i, f in enumerate(xlsx_files, 1):
                print(f"  {i}. {f}")
            choice = input("Номер: ").strip()
            try:
                excel_path = os.path.join(base_dir, xlsx_files[int(choice) - 1])
            except (ValueError, IndexError):
                log.error("Неверный выбор")
                sys.exit(1)

    msg_path = get_dummy_msg(base_dir)
    docs = load_excel(excel_path)
    for doc in docs:
        doc["файл"] = msg_path

    if not docs:
        log.error("Нет данных!")
        input("Enter...")
        sys.exit(1)

    print(f"\nПервые 5:")
    for i, d in enumerate(docs[:5], 1):
        ao = d.get("ao") or ""
        ti = d.get("тема_индекс") or ""
        print(f"  {i}. [{ao:3} тема={ti}] {d['корреспондент']} | "
              f"{d['содержание'][:50]}...")
    mode_label = os.environ.get('ASUD_MODE', 'auto-create')
    print(f"\nВсего: {len(docs)}  •  режим: {mode_label.upper()}")
    print("auto-create: создание + регистрация + На резолюцию (без .msg)")

    if input("Начать? (да/нет): ").strip().lower() not in ("да", "д", "y", "yes", ""):
        sys.exit(0)

    driver_path = os.path.join(base_dir, "msedgedriver.exe")
    if not os.path.exists(driver_path):
        log.error(f"msedgedriver.exe не найден")
        input("Enter...")
        sys.exit(1)

    options = EdgeOptions()
    # eager: driver.get() разблокируется на DOMContentLoaded, не ждёт картинки/iframes.
    # Все важные клики идут через WebDriverWait — ранний контроль безопасен.
    options.page_load_strategy = "eager"
    options.add_argument("--start-maximized")
    options.add_argument("--auth-server-whitelist=*.interrao.ru")
    options.add_argument("--auth-negotiate-delegate-whitelist=*.interrao.ru")
    options.add_argument("--log-level=3")
    options.add_experimental_option('excludeSwitches', ['enable-logging'])

    driver = webdriver.Edge(service=EdgeService(executable_path=driver_path), options=options)

    try:
        url = settings.get("asud_url", cfg.DEFAULTS["asud_url"])
        driver.get(url)
        wait_asud_loaded(driver)

        # Output xlsx: ОПТС|Округ|ФИО|Link рядом с реестром
        output_path = _output_xlsx_path(excel_path)
        _ensure_output_xlsx(output_path)
        log.info(f"Output xlsx: {output_path}")

        done_count, err_count = 0, 0
        for i, doc in enumerate(docs, 1):
            try:
                asud_id = create_one_document(driver, doc, i, len(docs))
                _append_output_row(output_path, doc, asud_id)
                done_count += 1
            except Exception as e:
                log.error(f"ОШИБКА документ {i}: {e}")
                err_count += 1
                driver.get(url)
                wait_asud_loaded(driver)

        elapsed_seconds = time.monotonic() - start_time
        elapsed = timedelta(seconds=int(elapsed_seconds))
        avg = timedelta(seconds=int(elapsed_seconds / done_count)) if done_count else None

        summary = [
            "",
            "=" * 60,
            f"ГОТОВО!",
            f"  Обработано: {done_count} / {len(docs)}",
            f"  Ошибок:     {err_count}",
            f"  Затрачено:  {elapsed}" + (f"  (в среднем {avg}/док)" if avg else ""),
            "=" * 60,
        ]
        for line in summary:
            log.info(line)
            print(line)

        input("\nEnter для закрытия...")
    except Exception as e:
        log.error(f"Ошибка: {e}")
        input("Enter...")
    finally:
        driver.quit()


if __name__ == "__main__":
    main()
