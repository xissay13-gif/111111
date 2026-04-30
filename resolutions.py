"""
resolutions.py — Выдача резолюций по обращениям из реестра.

Запускается под учёткой Есиной М.В., автоматически переключается
на учётку Халецкой Ю.В. (через выпадашку в шапке профиля).
Заходит в раздел "На резолюцию", для каждой строки реестра находит
соответствующий документ в АСУД и выдаёт резолюцию начальнику
абонентского отдела (по округу из колонки ao/fio в реестре).

Реестр: Лист2 с колонками Link, Subject, TextBody, Тема, To, LS, ao, fio.
Матч документа в АСУД: по номеру обращения из TextBody (regex
"обращение № NNNNNN"), fallback на подстроку первых 60 символов TextBody.
"""

import os
import re
import sys
import time
import logging
from datetime import date, datetime, timedelta

import openpyxl
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import config as cfg
from ui import (click, wait_and_click, find_input_near_label,
                wait_asud_loaded, wait_modal_closed, close_open_modals, js_set_value)
from correspondent import match_correspondent

_log_console = logging.StreamHandler()
_log_console.setLevel(logging.INFO)
_log_console.setFormatter(logging.Formatter(
    '%(asctime)s [%(levelname)s] %(message)s', datefmt='%H:%M:%S'))

logging.basicConfig(level=logging.DEBUG, handlers=[_log_console])
log = logging.getLogger("asud.res")
start_time = time.monotonic()


def _attach_file_logger(base_dir):
    """Подключает FileHandler с DEBUG: <base_dir>/resolutions_<timestamp>.log."""
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(base_dir, f"resolutions_{ts}.log")
        fh = logging.FileHandler(path, encoding='utf-8')
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(logging.Formatter(
            '%(asctime)s.%(msecs)03d [%(levelname)s] %(name)s: %(message)s',
            datefmt='%H:%M:%S'))
        logging.getLogger().addHandler(fh)
        log.info(f"Подробный лог пишется в: {path}")
        return path
    except Exception as e:
        log.warning(f"Не удалось создать файл лога: {e}")
        return None

settings = {}


# ============================================================
# EXCEL
# ============================================================

APPEAL_RE = re.compile(r'обращени[ея]\s*№?\s*(\d{4,10})', re.IGNORECASE)


# Lookup street_name → set of (house, okrug_short) — индекс для матчинга
# адресов из TextBody с адресной БД. Загружается лениво.
_STREET_INDEX = None
_ALL_STREETS_SORTED = None  # отсортированы по длине (длинные первыми)


def _addresses_csv_path():
    """Возвращает путь к addresses.csv: внутри exe (через _MEIPASS)
    или рядом с .py-скриптом в dev-режиме."""
    # PyInstaller --onefile распаковывает данные в sys._MEIPASS
    meipass = getattr(sys, '_MEIPASS', None)
    if meipass:
        path = os.path.join(meipass, 'addresses.csv')
        if os.path.exists(path):
            return path
    # Fallback: рядом с exe / скриптом
    base = cfg.get_base_dir()
    path = os.path.join(base, 'addresses.csv')
    if os.path.exists(path):
        return path
    return None


# --- Нормализация для парсинга адресов ---

_PREFIX_RE = re.compile(
    r'^(?:г\s*омск\s*,?\s*)?'
    r'(?:улица|ул\.?|проспект|пр[-]?т\.?|пр[-]?кт\.?|пр\.?|переулок|пер\.?|'
    r'бульвар|б[-]?р\.?|площадь|пл\.?|шоссе|ш\.?|набережная|наб\.?|'
    r'линия|тупик|проезд|пр[-]?д\.?|микрорайон|мкр\.?)\s*',
    re.IGNORECASE)

_HOUSE_RE = re.compile(r'(\d+[а-я]?)(?:[/\\-](\d+[а-я]?))?', re.IGNORECASE)


def _norm_text_for_match(s):
    """Нормализация для сопоставления: lower, ё→е, без пунктуации, single space.
    Также '3-я Молодежная' → '3 Молодежная'."""
    if not s:
        return ''
    s = str(s).lower().replace('ё', 'е')
    s = re.sub(r'[«»"\'`]', '', s)
    s = re.sub(r'[.,;:()\\/]+', ' ', s)
    s = re.sub(r'(\d+)[\s-]*я(?=\s)', r'\1', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _norm_street_name(s):
    """Только название улицы без префиксов."""
    return _PREFIX_RE.sub('', _norm_text_for_match(s)).strip()


def _norm_house_main(s):
    """Возвращает основной номер дома (первое число)."""
    if not s: return ''
    m = _HOUSE_RE.search(str(s).lower().replace('ё', 'е'))
    return m.group(1) if m else ''


def _build_street_index():
    """Парсит addresses.csv в индекс {street_norm: set((house, okrug_short))}.
    Используется один раз (кешируется в _STREET_INDEX)."""
    global _STREET_INDEX, _ALL_STREETS_SORTED
    if _STREET_INDEX is not None:
        return _STREET_INDEX, _ALL_STREETS_SORTED
    path = _addresses_csv_path()
    if not path:
        log.warning("addresses.csv не найден — авто-определение округа отключено")
        _STREET_INDEX = {}
        _ALL_STREETS_SORTED = []
        return _STREET_INDEX, _ALL_STREETS_SORTED

    # CSV: LS;okrug
    # Старая версия CSV содержит только LS+okrug, но для адрес-парсинга
    # нужны улицы из исходного xlsx. Если у нас только CSV без улиц —
    # парсер не сработает; идём через колонку ao реестра.
    # В будущем addresses.csv можно расширить колонкой улицы.
    try:
        import csv
        from collections import defaultdict
        idx = defaultdict(set)
        with open(path, encoding='utf-8') as f:
            reader = csv.reader(f, delimiter=';')
            header = next(reader, None)
            # Пытаемся определить колонки
            if not header or 'okrug' not in [h.lower() for h in header]:
                _STREET_INDEX = {}
                _ALL_STREETS_SORTED = []
                return _STREET_INDEX, _ALL_STREETS_SORTED
            # Если в CSV есть street/house — используем
            cols = {h.lower(): i for i, h in enumerate(header)}
            if 'street' in cols and 'house' in cols:
                for row in reader:
                    if not row or len(row) <= max(cols['street'], cols['house'], cols['okrug']):
                        continue
                    street = _norm_street_name(row[cols['street']])
                    house = _norm_house_main(row[cols['house']])
                    okrug = row[cols['okrug']].strip()
                    if street and house and okrug:
                        idx[street].add((house, okrug))
                        # Алиасы: 'молодежная 3-я' → '3-я молодежная',
                        # '3 молодежная', 'молодежная 3'
                        m = re.match(r'^(.+?)\s+(\d+)[\s-]*я$', street)
                        if m:
                            base, num = m.group(1).strip(), m.group(2)
                            idx[f"{base} {num}"].add((house, okrug))
                            idx[f"{num}-я {base}"].add((house, okrug))
                            idx[f"{num} {base}"].add((house, okrug))
        log.info(f"Street index: {len(idx)} улиц")
        _STREET_INDEX = idx
        _ALL_STREETS_SORTED = sorted(idx.keys(), key=lambda s: -len(s))
    except Exception as e:
        log.warning(f"Ошибка построения street index: {e}")
        _STREET_INDEX = {}
        _ALL_STREETS_SORTED = []
    return _STREET_INDEX, _ALL_STREETS_SORTED


def _find_street_house(text, idx, sorted_streets):
    """Ищет известную улицу в нормализованном тексте, потом дом рядом."""
    norm = _norm_text_for_match(text)
    for street in sorted_streets:
        if len(street) < 3:
            continue
        # Поиск как целое слово
        pos = 0
        while True:
            i = norm.find(street, pos)
            if i < 0: break
            left_ok = i == 0 or not norm[i-1].isalnum()
            end = i + len(street)
            right_ok = end == len(norm) or not norm[end].isalnum()
            if left_ok and right_ok:
                # Дом в окне 50 символов после улицы
                tail = norm[end:end+50]
                m = re.search(r'\b(\d+[а-я]?)', tail)
                if m:
                    return (street, m.group(1))
            pos = i + 1
    return (None, None)


def _okrug_from_textbody(textbody):
    """Извлекает округ из TextBody.

    Стратегия: суть обращения → почтовый адрес → весь текст.
    Возвращает короткий код округа ('КАО', 'ЦАО', ...) или None.
    """
    if not textbody:
        return None
    idx, sorted_streets = _build_street_index()
    if not idx:
        return None
    text = str(textbody)
    fragments = []
    # 1) Суть обращения
    m = re.search(r'суть\s+обращени[яе]\s*:?\s*([\s\S]+?)(?:\n\s*\n|$)',
                  text, re.IGNORECASE)
    if m:
        fragments.append(('суть', m.group(1)))
    # 2) Почтовый адрес
    m = re.search(r'почтов[а-я]+\s+адрес[а-я]*\s*:\s*([^\n]+)',
                  text, re.IGNORECASE)
    if m:
        fragments.append(('почт', m.group(1)))
    # 3) Весь текст как fallback
    fragments.append(('весь', text))

    for name, frag in fragments:
        street, house = _find_street_house(frag, idx, sorted_streets)
        if street and house:
            # '15г' → '15' для сравнения (литерные суффиксы — те же дома)
            house_digits = re.match(r'\d+', house)
            house_num = house_digits.group(0) if house_digits else house
            for h, o in idx[street]:
                h_digits = re.match(r'\d+', h)
                h_num = h_digits.group(0) if h_digits else h
                if h == house or h_num == house_num:
                    log.info(f"  адрес [{name}]: {street} {house} → {o}")
                    return o
    return None


def _resolve_executor(ao, fio, ls=None, textbody=None):
    """Возвращает ФИО начальника по приоритету:
    1. fio из реестра (если заполнено)
    2. ao из реестра → DEFAULT_OKRUG_MAP
    3. адрес из TextBody → addresses.csv → DEFAULT_OKRUG_MAP

    Параметр ls пока не используется — формат LS в реестре не совпадает
    с адресной БД (11 цифр vs 6), нужно правило конвертации.
    """
    # 1. fio напрямую (вручную проставленный)
    if fio and str(fio).strip():
        return str(fio).strip()
    # 2. ao из реестра (вручную проставленный)
    if ao:
        key = str(ao).strip()
        v = cfg.DEFAULT_OKRUG_MAP.get(key)
        if v:
            return v
    # 3. Парсинг адреса из TextBody
    if textbody:
        ao_short = _okrug_from_textbody(textbody)
        if ao_short:
            v = cfg.DEFAULT_OKRUG_MAP.get(ao_short)
            if v:
                return v
    return None


def load_excel(file_path):
    """Читает таблицу резолюций. Ожидаемые колонки (любой порядок,
    распознаются по заголовку): ОПТС, Округ, ФИО, Link.

    Если файл — старый формат (Лист2 с TextBody) — fallback на адрес-парсер.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Заголовки
    header = [str(c.value or '').strip() for c in next(ws.iter_rows(max_row=1))]
    header_lower = [h.lower() for h in header]
    log.info(f"Заголовки: {header}")

    # Если это _резолюции.xlsx (есть колонка ОПТС/asud_id)
    asud_keys = ('опт', 'орт', 'асуд', 'asud', 'регистрацион')
    fio_keys = ('фио', 'исполнит')
    okrug_keys = ('округ', 'ао')
    link_keys = ('link', 'ссылк')

    def _col(predicate_keys):
        for i, h in enumerate(header_lower):
            for k in predicate_keys:
                if k in h:
                    return i
        return None

    asud_col = _col(asud_keys)
    if asud_col is not None:
        # Формат _резолюции.xlsx
        fio_col = _col(fio_keys)
        okrug_col = _col(okrug_keys)
        link_col = _col(link_keys)
        log.info(f"Формат _резолюции: ОПТС=col{asud_col}, "
                 f"ФИО=col{fio_col}, Округ=col{okrug_col}, Link=col{link_col}")

        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row:
                continue
            asud_id = str(row[asud_col]).strip() if row[asud_col] else ''
            fio = str(row[fio_col]).strip() if (fio_col is not None and row[fio_col]) else ''
            ao = str(row[okrug_col]).strip() if (okrug_col is not None and row[okrug_col]) else ''
            link = row[link_col] if (link_col is not None) else None

            # Если ФИО пуст, но Округ есть — пытаемся через мапу
            if not fio and ao:
                fio = cfg.DEFAULT_OKRUG_MAP.get(ao, '')

            if not asud_id and not link:
                continue

            rows.append({
                "row_idx": row_idx,
                "asud_id": asud_id,
                "executor": fio,
                "ao": ao,
                "link": link,
                "subject": "",
                "textbody": "",
                "appeal_no": None,
                "ls": "",
            })
        wb.close()
        log.info(f"Загружено: {len(rows)} строк (формат _резолюции)")
        no_asud = sum(1 for r in rows if not r['asud_id'])
        no_fio = sum(1 for r in rows if not r['executor'])
        if no_asud:
            log.warning(f"  без ОПТС/ОРТС: {no_asud} (попробую матч по тексту)")
        if no_fio:
            log.warning(f"  без ФИО: {no_fio} (будут пропущены)")
        return rows

    # FALLBACK: старый формат — Почта_ТЭС.xlsx Лист2
    log.warning("Колонка ОПТС не найдена — пробую старый формат Лист2")
    if 'Лист2' in wb.sheetnames:
        ws = wb['Лист2']

    rows = []
    skipped = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
        if not row or len(row) < 8:
            skipped += 1
            continue
        link = row[0]
        subject = row[1]
        textbody = row[2] or ''
        ls = row[5] if len(row) > 5 else None
        ao = row[6] if len(row) > 6 else None
        fio = row[7] if len(row) > 7 else None
        if not link and not subject:
            skipped += 1
            continue
        executor = _resolve_executor(ao, fio, ls, textbody)
        appeal_no = _extract_appeal_no(textbody)
        rows.append({
            "row_idx": row_idx,
            "asud_id": "",
            "link": link,
            "subject": str(subject or '').strip(),
            "textbody": str(textbody),
            "ls": _norm_ls(ls),
            "ao": str(ao or '').strip(),
            "executor": executor,
            "appeal_no": appeal_no,
        })
    wb.close()
    log.info(f"Загружено: {len(rows)} (старый формат), пропущено: {skipped}")
    return rows


# ============================================================
# UI: переключение учётки
# ============================================================

def switch_account(driver, target_substring):
    """Переключается на учётку, ФИО которой содержит target_substring.
    1. Клик по dropdown-стрелке в шапке профиля
    2. Клик по пункту с нужным ФИО в выпадашке
    """
    log.info(f"Переключение на учётку: {target_substring}")
    time.sleep(2)

    # Шаг 1: клик ▼ рядом с именем профиля
    try:
        # Стрелочка вниз обычно лежит рядом с блоком пользователя в самом верху страницы
        # На скрине она 21x55px, расположена сразу справа от блока с именем
        # Ищем по клик-целям рядом с блоком пользователя
        triggers = driver.find_elements(By.CSS_SELECTOR,
            "img[class*='trigger'], img[class*='Trigger'], div[class*='trigger']")
        # Выбираем те что в верхней части страницы (top < 100px)
        candidate = None
        for t in triggers:
            try:
                if not t.is_displayed():
                    continue
                rect = t.rect
                if rect.get('y', 0) < 120 and rect.get('x', 0) < 400:
                    candidate = t
                    break
            except Exception:
                continue
        if not candidate:
            # Fallback — любая стрелка/треугольник в верхнем левом углу
            log.warning("Стрелка профиля не найдена по trigger-классам, пробую по позиции")
            all_imgs = driver.find_elements(By.TAG_NAME, "img")
            for im in all_imgs:
                try:
                    if not im.is_displayed(): continue
                    rect = im.rect
                    if rect.get('y', 0) < 120 and rect.get('x', 0) < 400 \
                            and 10 <= rect.get('width', 0) <= 30:
                        candidate = im
                        break
                except Exception:
                    continue
        if not candidate:
            log.error("Не нашёл dropdown-стрелку профиля")
            return False
        click(driver, candidate, "▼ профиль")
        time.sleep(1.5)
    except Exception as e:
        log.error(f"Ошибка клика по dropdown профиля: {e}")
        return False

    # Шаг 2: клик по пункту со строкой target_substring
    try:
        WebDriverWait(driver, 10).until(
            lambda d: any(target_substring in (e.text or '')
                          for e in d.find_elements(By.CSS_SELECTOR, "div, span, label")
                          if e.is_displayed()))
    except Exception:
        log.warning(f"Пункт '{target_substring}' в выпадашке не появился")

    items = driver.find_elements(By.XPATH,
        f"//*[contains(normalize-space(text()), '{target_substring}')]")
    target = None
    for it in items:
        try:
            if not it.is_displayed():
                continue
            rect = it.rect
            # должен быть всплывающий пункт под профилем (y > стрелки)
            if rect.get('y', 0) > 100:
                target = it
                break
        except Exception:
            continue
    if not target and items:
        target = next((i for i in items if i.is_displayed()), None)

    if not target:
        log.error(f"Пункт с '{target_substring}' не найден")
        return False

    click(driver, target, f"учётка {target_substring}")
    time.sleep(3)
    log.info("Переключение запущено, жду перезагрузку АСУД")
    _wait_profile_loaded(driver)
    return True


# ============================================================
# UI: сайдбар → "На резолюцию"
# ============================================================

def click_sidebar_section(driver, section_text):
    """Клик по пункту в левом сайдбаре."""
    log.info(f"Сайдбар → '{section_text}'")
    items = driver.find_elements(By.XPATH,
        f"//*[normalize-space(text())='{section_text}']")
    target = None
    for it in items:
        try:
            if it.is_displayed():
                target = it
                break
        except Exception:
            continue
    if not target:
        log.error(f"Пункт сайдбара '{section_text}' не найден")
        return False
    click(driver, target, f"сайдбар: {section_text}")
    time.sleep(2)
    return True


# ============================================================
# UI: поиск и открытие документа в списке "На резолюцию"
# ============================================================

# GXT-сетка АСУД: <tr> с обоими классами obj-list-rec и obj-list-task —
# это строки данных. Заголовки/служебные tr этих классов не имеют.
# Раньше использовался id "CABINET_MENU__RECEIVED__ALL_ACTIVE__TO_RESOLUTION",
# но это id сайдбар-пункта, а не таблицы → находили tr меню вместо данных.
DATA_ROW_XPATH = ("//tr[contains(concat(' ',normalize-space(@class),' '),' obj-list-rec ')"
                  " and contains(concat(' ',normalize-space(@class),' '),' obj-list-task ')]")

# Фильтр-input под колонкой "Номер" — стабильный id из DOM
NUMBER_FILTER_INPUT_ID = "FCPC_Номер-input"
NUMBER_FILTER_CONTAINER_ID = "FCPC_Номер"


def _set_filter_value(driver, container_id, input_id, value):
    """JS-ввод в фильтр колонки — без эмуляции клавиатуры (фон-friendly)."""
    log.debug(f"_set_filter_value: ищу input id={input_id!r}")
    inp = None
    try:
        inp = driver.find_element(By.ID, input_id)
        log.debug(f"  input найден по ID")
    except Exception:
        log.debug(f"  по ID не нашёл, пробую внутри container={container_id!r}")
        try:
            container = driver.find_element(By.ID, container_id)
            inp = container.find_element(By.CSS_SELECTOR, "input[type='text']")
            log.debug(f"  input найден внутри container")
        except Exception as e:
            log.warning(f"Фильтр {container_id}: input не найден ({e})")
            return False
    try:
        driver.execute_script("""
            var el = arguments[0], v = arguments[1];
            el.focus();
            el.value = v;
            el.dispatchEvent(new Event('input', {bubbles:true}));
            el.dispatchEvent(new Event('keyup', {bubbles:true}));
            el.dispatchEvent(new Event('change', {bubbles:true}));
        """, inp, value)
        log.debug(f"  JS dispatch выполнен; value={value!r}")
        return True
    except Exception as e:
        log.warning(f"JS-ввод в фильтр упал: {e}")
        return False


def filter_by_number(driver, asud_id):
    """Вбивает ОПТС/ОРТС-номер в фильтр колонки 'Номер'.
    Возвращает True если удалось ввести."""
    log.info(f"Фильтр 'Номер' = {asud_id}")
    return _set_filter_value(driver, NUMBER_FILTER_CONTAINER_ID,
                              NUMBER_FILTER_INPUT_ID, asud_id)


def clear_filter(driver):
    """Очищает фильтр колонки 'Номер'."""
    _set_filter_value(driver, NUMBER_FILTER_CONTAINER_ID,
                       NUMBER_FILTER_INPUT_ID, "")
    time.sleep(0.5)


def find_doc_row(driver, doc, timeout=8):
    """Находит <tr> в таблице списка после применения фильтра.

    Стратегия:
      1. Если есть doc['asud_id'] — вбить в фильтр 'Номер' → взять первую строку
      2. Иначе — поиск по тексту (appeal_no / TextBody / Subject) как раньше
    """
    asud_id = doc.get('asud_id') or ''
    log.debug(f"find_doc_row: asud_id={asud_id!r}, appeal_no={doc.get('appeal_no')!r}, "
             f"timeout={timeout}s")

    # Главная стратегия: фильтр по точному номеру
    if asud_id:
        log.info(f"[find] стратегия 1 (фильтр 'Номер'): {asud_id}")
        if filter_by_number(driver, asud_id):
            log.debug(f"  фильтр введён, жду дебаунс 1.5s")
            time.sleep(1.5)  # дебаунс GWT-фильтра
            end = time.monotonic() + timeout
            tick = 0
            while time.monotonic() < end:
                tick += 1
                try:
                    rows = driver.find_elements(By.XPATH, DATA_ROW_XPATH)
                    visible_rows = [r for r in rows if r.is_displayed()
                                    and (r.text or '').strip()]
                    log.debug(f"  tick#{tick}: всего obj-list-rec.obj-list-task <tr>="
                             f"{len(rows)}, видимых-с-текстом={len(visible_rows)}")
                    if visible_rows:
                        first_text = (visible_rows[0].text or '').replace('\n', ' ')[:80]
                        log.info(f"[find] МАТЧ по фильтру: {asud_id} → {len(visible_rows)} строк")
                        log.debug(f"  первая строка: {first_text!r}")
                        return visible_rows[0]
                except Exception as e:
                    log.debug(f"  tick#{tick}: исключение {e}")
                time.sleep(0.5)
            log.warning(f"[find] фильтр {asud_id} → 0 строк за {timeout}s")
            return None
        else:
            log.warning(f"[find] не получилось ввести фильтр для {asud_id}")

    # Fallback на текстовый поиск (без фильтра — сканируем что в DOM)
    log.info(f"[find] стратегия 2 (fallback по тексту)")
    end = time.monotonic() + timeout
    tick = 0
    while time.monotonic() < end:
        tick += 1
        if doc.get('appeal_no'):
            try:
                row = driver.find_element(By.XPATH,
                    f"{DATA_ROW_XPATH}[contains(., '{doc['appeal_no']}')]")
                if row.is_displayed():
                    log.info(f"[find] МАТЧ (fallback): appeal № {doc['appeal_no']}")
                    return row
            except Exception:
                pass
        body = (doc.get('textbody') or '').replace('\xa0', ' ').strip()
        snippet = re.sub(r'\s+', ' ', body)[:60].strip().replace("'", "")
        if len(snippet) >= 20:
            try:
                row = driver.find_element(By.XPATH,
                    f"{DATA_ROW_XPATH}[contains(., \"{snippet}\")]")
                if row.is_displayed():
                    log.info(f"[find] МАТЧ (fallback): TextBody {snippet!r}")
                    return row
            except Exception:
                pass
        log.debug(f"  fallback tick#{tick}: не нашёл")
        time.sleep(0.5)
    log.warning(f"[find] не найдено ни одной стратегией за {timeout}s")
    return None


def _card_opened(driver, timeout):
    """Признак открывшейся карточки — кнопка 'Создать резолюцию'."""
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.ID, "header-action-btn-add_resolution")))
        return True
    except Exception:
        return False


def _refresh_first_row(driver):
    """Возвращает свежий <tr> первой видимой строки таблицы (после фильтра).

    Хитрость GXT: после фильтра старая ссылка на <tr> часто становится
    stale — пересортировка перерисовывает грид.
    """
    try:
        rows = driver.find_elements(By.XPATH, DATA_ROW_XPATH)
        log.debug(f"_refresh_first_row: data <tr>={len(rows)}")
        for idx, r in enumerate(rows):
            try:
                if r.is_displayed() and (r.text or '').strip():
                    cls_preview = (r.get_attribute('class') or '')[:80]
                    log.debug(f"  → свежий ref на data-<tr>[{idx}], class={cls_preview!r}")
                    return r
            except Exception:
                continue
        log.debug(f"  → ни одной видимой data-<tr> с текстом")
    except Exception as e:
        log.debug(f"  ошибка поиска data-<tr>: {e}")
    return None


def _meaningful_cell(row):
    """Берёт ячейку с текстом (subject/тип) — не чекбокс, не иконку."""
    try:
        tds = row.find_elements(By.XPATH, ".//td")
    except Exception as e:
        log.debug(f"_meaningful_cell: не нашёл <td>: {e}")
        return None
    log.debug(f"_meaningful_cell: всего <td>={len(tds)}")
    best = None
    for idx, td in enumerate(tds):
        try:
            if not td.is_displayed():
                continue
            txt = (td.text or '').strip()
            w = td.size.get('width', 0)
            if len(txt) > 10:  # пропускаем чекбоксы/иконки
                log.debug(f"  → выбрана <td>[{idx}] w={w} text={txt[:40]!r}")
                return td
            if best is None and w > 50:
                best = td
                log.debug(f"  кандидат <td>[{idx}] w={w} (без текста)")
        except Exception:
            continue
    log.debug(f"  → fallback на best={'найден' if best else 'нет'}")
    return best


def open_doc_card(driver, row):
    """Открывает карточку документа.

    GXT-сетка обычно реагирует так: одиночный клик выделяет строку
    (добавляет класс rowSelected), а двойной открывает карточку.
    ActionChains.double_click по разным причинам часто не срабатывает,
    поэтому пробуем несколько стратегий по очереди.
    """
    log.info("[open] начинаю открывать карточку документа")
    # Свежая ссылка — пред. могла стать stale
    fresh = _refresh_first_row(driver) or row
    log.debug(f"[open] fresh={'свежая ссылка' if fresh is not row else 'та же что передали'}")

    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", fresh)
        time.sleep(0.3)
        log.debug(f"[open] scrollIntoView OK")
    except Exception as e:
        log.debug(f"[open] scrollIntoView err: {e}")

    # Берём содержательную ячейку (Subject/Тип) — не чекбокс
    target_cell = _meaningful_cell(fresh) or fresh
    log.debug(f"[open] target_cell={'<td>' if target_cell is not fresh else '<tr>'}")

    # ── Стратегия 1: single click + Enter (самый надёжный для GXT)
    log.info("[open] strat1: single click + Enter")
    try:
        ActionChains(driver).move_to_element(target_cell).pause(0.15).click().perform()
        log.debug(f"  click выполнен")
        try:
            cls = (fresh.get_attribute('class') or '')
            log.debug(f"  class у <tr> после клика: {cls!r}")
        except Exception:
            pass
        time.sleep(0.4)
        ActionChains(driver).send_keys(Keys.ENTER).perform()
        log.debug(f"  Enter отправлен")
    except Exception as e:
        log.debug(f"  strat1 err: {e}")
    if _card_opened(driver, timeout=5):
        log.info("[open] УСПЕХ: strat1 click + Enter")
        return True
    log.info("[open] strat1 не сработал, пробую strat2")

    # ── Стратегия 2: ActionChains double-click по содержательной ячейке
    log.info("[open] strat2: ActionChains dblclick по <td>")
    try:
        fresh2 = _refresh_first_row(driver) or fresh
        cell2 = _meaningful_cell(fresh2) or fresh2
        ActionChains(driver).move_to_element(cell2).pause(0.2).double_click().perform()
        log.debug(f"  double_click выполнен")
    except Exception as e:
        log.debug(f"  strat2 err: {e}")
    if _card_opened(driver, timeout=4):
        log.info("[open] УСПЕХ: strat2 ActionChains dblclick")
        return True
    log.info("[open] strat2 не сработал, пробую strat3")

    # ── Стратегия 3: JS — полная mouse-event цепочка (mousedown/up/click x2 + dblclick)
    log.info("[open] strat3: JS mouse-event chain")
    try:
        fresh3 = _refresh_first_row(driver) or fresh
        cell3 = _meaningful_cell(fresh3) or fresh3
        driver.execute_script("""
            const el = arguments[0];
            const r = el.getBoundingClientRect();
            const x = r.left + r.width/2, y = r.top + r.height/2;
            const opts = {bubbles:true, cancelable:true, view:window,
                          button:0, buttons:1, clientX:x, clientY:y};
            for (const t of ['mousedown','mouseup','click']) {
                el.dispatchEvent(new MouseEvent(t, {...opts, detail:1}));
            }
            for (const t of ['mousedown','mouseup','click']) {
                el.dispatchEvent(new MouseEvent(t, {...opts, detail:2}));
            }
            el.dispatchEvent(new MouseEvent('dblclick', {...opts, detail:2}));
        """, cell3)
        log.debug(f"  JS event chain dispatched")
    except Exception as e:
        log.debug(f"  strat3 err: {e}")
    if _card_opened(driver, timeout=4):
        log.info("[open] УСПЕХ: strat3 JS event chain")
        return True
    log.info("[open] strat3 не сработал, пробую strat4")

    # ── Стратегия 4: клик по ссылке/anchor если есть
    log.info("[open] strat4: click по <a>/anchor внутри строки")
    try:
        fresh4 = _refresh_first_row(driver) or fresh
        links = fresh4.find_elements(By.XPATH,
            ".//a | .//*[contains(@class,'gwt-Anchor')] | .//*[contains(@class,'cellClickable')]")
        log.debug(f"  найдено anchor/link={len(links)}")
        for idx, lnk in enumerate(links):
            if lnk.is_displayed():
                log.debug(f"  кликаю anchor[{idx}] tag={lnk.tag_name}")
                try:
                    driver.execute_script("arguments[0].click();", lnk)
                except Exception:
                    try:
                        lnk.click()
                    except Exception:
                        continue
                if _card_opened(driver, timeout=4):
                    log.info(f"[open] УСПЕХ: strat4 link-click anchor[{idx}]")
                    return True
    except Exception as e:
        log.debug(f"  strat4 err: {e}")

    log.warning("[open] ПРОВАЛ: все 4 стратегии не сработали")
    return False


# ============================================================
# UI: диалог "Корневая резолюция"
# ============================================================

def _wait_data_value(driver, container, target_value="true", timeout=5):
    """Ждёт пока у контейнера-тоггла data-value станет target_value."""
    end = time.monotonic() + timeout
    while time.monotonic() < end:
        try:
            if container.get_attribute('data-value') == target_value:
                return True
        except Exception:
            pass
        time.sleep(0.2)
    return False


def toggle_switch(driver, label_text, target_value="true"):
    """Переключает тоггл рядом с label_text в нужное состояние."""
    try:
        label = driver.find_element(By.XPATH,
            f"//*[normalize-space(text())='{label_text}']")
        container = label.find_element(By.XPATH,
            "./following::*[contains(@class,'switcherContainer')][1]")
    except Exception as e:
        log.warning(f"Тоггл '{label_text}' не найден: {e}")
        return False

    cur = container.get_attribute('data-value')
    if cur == target_value:
        log.info(f"Тоггл '{label_text}' уже = {target_value}")
        return True

    click(driver, container, f"тоггл {label_text} → {target_value}")
    if _wait_data_value(driver, container, target_value, timeout=3):
        log.info(f"Тоггл '{label_text}' = {target_value}")
        return True
    log.warning(f"Тоггл '{label_text}' не переключился")
    return False


def select_content_template(driver, template_text):
    """Выбирает в поле "Содержание" пункт из выпадашки."""
    try:
        # input с placeholder "Общие формулировки"
        inp = None
        candidates = driver.find_elements(By.CSS_SELECTOR,
            "input[placeholder='Общие формулировки']")
        for c in candidates:
            if c.is_displayed():
                inp = c
                break
        if not inp:
            # Fallback — поиск по label "Содержание"
            inp = find_input_near_label(driver, "Содержание")
        if not inp:
            log.error("Поле 'Содержание' не найдено")
            return False

        click(driver, inp, "Содержание")
        time.sleep(1)

        # Дропдаун с пунктами
        items = driver.find_elements(By.XPATH,
            f"//*[normalize-space(text())='{template_text}']")
        target = None
        for it in items:
            try:
                if it.is_displayed():
                    target = it
                    break
            except Exception:
                continue
        if not target:
            log.error(f"Пункт '{template_text}' в выпадашке не найден")
            return False
        click(driver, target, f"Содержание: {template_text}")
        time.sleep(0.5)
        return True
    except Exception as e:
        log.error(f"Ошибка выбора шаблона содержания: {e}")
        return False


def add_business_days(start, days):
    cur = start
    added = 0
    while added < days:
        cur += timedelta(days=1)
        if cur.weekday() < 5:  # 0-4 = пн-пт
            added += 1
    return cur


def set_stage_date(driver, n_workdays):
    """Заполняет дату в поле 'Контрольный этап' = today + n рабочих дней."""
    deadline = add_business_days(date.today(), n_workdays).strftime("%d.%m.%Y")
    try:
        inp = driver.find_element(By.CSS_SELECTOR,
            "input[id*='stage_control_date']")
        js_set_value(driver, inp, deadline)
        log.info(f"Контрольный этап: {deadline}")
        return True
    except Exception as e:
        log.warning(f"Поле даты этапа не найдено: {e}")
        return False


def fill_executor(driver, fio):
    """Вбивает ФИО в поле 'Исполнитель' (combobox), выбирает из выпадашки."""
    try:
        # Ищем поле по label "Исполнитель"
        inp = find_input_near_label(driver, "Исполнитель")
        if not inp:
            # Fallback — input id select_combobox-input
            inp = driver.find_element(By.ID, "select_combobox-input")
        if not inp:
            log.error("Поле 'Исполнитель' не найдено")
            return False

        surname = fio.split()[0]
        inp.click()
        time.sleep(0.3)
        inp.clear()
        time.sleep(0.3)
        for ch in surname:
            inp.send_keys(ch)
            time.sleep(0.08)
        log.info(f"Введена фамилия: {surname}")
        time.sleep(2)

        # Кандидаты в выпадашке
        results = driver.find_elements(By.XPATH,
            f"//*[contains(text(),'{surname}')]")
        candidates = [r for r in results
                      if r.is_displayed() and r != inp
                      and r.tag_name.lower() != 'input']
        if not candidates:
            inp.send_keys(Keys.ENTER)
            time.sleep(2)
            results = driver.find_elements(By.XPATH,
                f"//*[contains(text(),'{surname}')]")
            candidates = [r for r in results
                          if r.is_displayed() and r != inp
                          and r.tag_name.lower() != 'input']

        log.info(f"Кандидатов: {len(candidates)}")
        target = None
        for idx, r in enumerate(candidates, 1):
            try:
                txt = (r.text or '').strip()
                if len(txt) > 200:
                    continue
                ok = match_correspondent(txt, fio)
                preview = txt.replace('\n', ' ')[:80]
                log.info(f"  [{idx}] {'OK' if ok else '--'} | {preview!r}")
                if ok and target is None:
                    target = r
            except Exception:
                continue

        if not target:
            log.error(f"Исполнитель '{fio}' не найден в выпадашке")
            return False

        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", target)
        time.sleep(0.3)
        ActionChains(driver).move_to_element(target).pause(0.3).click().perform()
        time.sleep(1)
        log.info(f"Исполнитель выбран: {fio}")
        return True
    except Exception as e:
        log.error(f"Ошибка ввода исполнителя: {e}")
        return False


def _wait_button_enabled(driver, btn_id, timeout=15):
    """Ждёт пока кнопка с data-disabled='1' не станет активной."""
    end = time.monotonic() + timeout
    while time.monotonic() < end:
        try:
            btn = driver.find_element(By.ID, btn_id)
            if btn.get_attribute('data-disabled') != '1' and btn.is_displayed():
                return btn
        except Exception:
            pass
        time.sleep(0.3)
    return None


def _click_confirm_yes(driver, timeout=10):
    """Ждёт и кликает 'Да' в confirm-диалоге АСУД (с fallback'ами).

    Используется после 'Сохранить и отправить' (подтверждение
    отправки адресатам) — аналогично confirm после 'На резолюцию'
    в основных скриптах.
    """
    yes_btn = None
    end = time.monotonic() + timeout
    while time.monotonic() < end:
        # 1) Точный id
        try:
            btn = driver.find_element(By.ID, "confirm_dialog_btn_yes")
            if btn.is_displayed():
                yes_btn = btn
                break
        except Exception:
            pass
        # 2) Substring id (GWT может префиксовать)
        try:
            btn = driver.find_element(By.CSS_SELECTOR,
                "[id*='confirm_dialog_btn_yes'], [id*='confirm'][id*='yes']")
            if btn.is_displayed():
                yes_btn = btn
                break
        except Exception:
            pass
        # 3) По тексту "Да"
        try:
            for b in driver.find_elements(By.XPATH,
                    "//*[normalize-space(text())='Да']"):
                if b.is_displayed():
                    yes_btn = b
                    break
        except Exception:
            pass
        if yes_btn:
            break
        time.sleep(0.5)

    if not yes_btn:
        log.warning("Confirm-диалог 'Да' не появился")
        return False

    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", yes_btn)
        time.sleep(0.3)
    except Exception:
        pass
    clicked = False
    try:
        ActionChains(driver).move_to_element(yes_btn).pause(0.3).click().perform()
        log.info(f"Клик 'Да' (ActionChains): id={yes_btn.get_attribute('id')}")
        clicked = True
    except Exception:
        pass
    if not clicked:
        try:
            driver.execute_script("arguments[0].click();", yes_btn)
            log.info("Клик 'Да' (JS)")
            clicked = True
        except Exception:
            pass
    if not clicked:
        try:
            yes_btn.click()
            log.info("Клик 'Да' (native)")
            clicked = True
        except Exception:
            pass
    if clicked:
        try:
            ActionChains(driver).send_keys(Keys.ENTER).perform()
        except Exception:
            pass
    time.sleep(2)
    return clicked


def submit_resolution(driver):
    """Финальный шаг: 'Сохранить и отправить' → confirm 'Да' → закрыть карточку."""
    log.debug("[submit] жду активации 'Сохранить и отправить' (id=save_and_send_btn)")
    btn = _wait_button_enabled(driver, "save_and_send_btn", timeout=15)
    if not btn:
        log.error("Кнопка 'Сохранить и отправить' не активировалась")
        return False
    log.info("[submit] клик 'Сохранить и отправить'")
    click(driver, btn, "Сохранить и отправить")
    time.sleep(2)

    # Подтверждение отправки адресатам
    log.info("[submit] жду confirm-диалог 'Да'")
    confirmed = _click_confirm_yes(driver, timeout=10)
    log.info(f"[submit] confirm 'Да': {'OK' if confirmed else 'не появился'}")
    time.sleep(2)

    # Может остаться открытой карточка документа — её закрываем отдельно
    # в close_card_after_resolution. На случай если "Корневая резолюция"
    # ещё открыта — закроем её через крестик в модалке.
    try:
        title = driver.find_element(By.XPATH,
            "//*[contains(text(),'Корневая резолюция')]")
        if title.is_displayed():
            log.warning("[submit] модалка 'Корневая резолюция' ещё открыта — крестик")
            close_open_modals(driver)
            time.sleep(1)
        else:
            log.debug("[submit] модалка 'Корневая резолюция' закрылась штатно")
    except Exception:
        log.debug("[submit] модалка 'Корневая резолюция' уже не в DOM")
    return True


def close_card_after_resolution(driver):
    """После выдачи резолюции возвращаемся в список через #header-close-btn."""
    time.sleep(2)
    closed = False
    try:
        close_btn = driver.find_element(By.ID, "header-close-btn")
        if close_btn.is_displayed():
            log.debug("[close] header-close-btn найден и видим")
            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", close_btn)
                time.sleep(0.3)
            except Exception:
                pass
            try:
                ActionChains(driver).move_to_element(close_btn).pause(0.3).click().perform()
                closed = True
                log.info("[close] карточка закрыта (ActionChains)")
            except Exception as e:
                log.debug(f"[close] ActionChains err: {e}")
            if not closed:
                try:
                    driver.execute_script("arguments[0].click();", close_btn)
                    closed = True
                    log.info("[close] карточка закрыта (JS click)")
                except Exception as e:
                    log.debug(f"[close] JS click err: {e}")
            if closed:
                time.sleep(2)
                return
        else:
            log.debug("[close] header-close-btn не видим")
    except Exception as e:
        log.debug(f"[close] header-close-btn не найден: {e}")
    log.info("[close] карточка уже закрыта или header-close-btn не найден")


# ============================================================
# DOCUMENT FLOW (один документ)
# ============================================================

def process_one(driver, doc, index, total):
    """Обработка одной строки реестра: найти, открыть, выдать резолюцию, закрыть."""
    t_start = time.monotonic()
    log.info("=" * 50)
    log.info(f"ДОКУМЕНТ {index}/{total}: link={doc.get('link')!r}, "
             f"appeal_no={doc.get('appeal_no')}, "
             f"исполнитель={doc.get('executor')}")
    log.debug(f"  доп: asud_id={doc.get('asud_id')!r}, ao={doc.get('ao')!r}, "
             f"row_idx={doc.get('row_idx')}")

    if not doc.get('executor'):
        log.warning(f"Row {doc['row_idx']}: исполнитель не определён → пропускаю")
        return False

    # 1. Найти строку в списке
    log.info(f"--- ШАГ 1/12: поиск документа в списке ---")
    row = find_doc_row(driver, doc, timeout=10)
    if not row:
        log.warning(f"Row {doc['row_idx']}: документ в списке не найден → пропускаю")
        return False
    log.debug(f"  Шаг 1 OK ({time.monotonic()-t_start:.1f}s)")

    # 2. Открыть карточку
    log.info(f"--- ШАГ 2/12: открыть карточку ---")
    if not open_doc_card(driver, row):
        return False
    log.debug(f"  Шаг 2 OK ({time.monotonic()-t_start:.1f}s)")

    # 3. Создать резолюцию
    log.info(f"--- ШАГ 3/12: кнопка 'Создать резолюцию' ---")
    try:
        btn = driver.find_element(By.ID, "header-action-btn-add_resolution")
        click(driver, btn, "Создать резолюцию")
        time.sleep(2)
        log.debug(f"  Шаг 3 OK ({time.monotonic()-t_start:.1f}s)")
    except Exception as e:
        log.error(f"Кнопка 'Создать резолюцию' не найдена: {e}")
        return False

    # 4. Содержание
    log.info(f"--- ШАГ 4/12: выбор шаблона 'Содержание' ---")
    select_content_template(driver,
        settings.get("resolution_content", cfg.DEFAULTS["resolution_content"]))
    log.debug(f"  Шаг 4 OK ({time.monotonic()-t_start:.1f}s)")

    # 5. Тоггл "Требуется отчёт"
    if settings.get("require_report", True):
        log.info(f"--- ШАГ 5/12: тоггл 'Требуется отчёт' ---")
        toggle_switch(driver, "Требуется отчёт", "true")
        log.debug(f"  Шаг 5 OK ({time.monotonic()-t_start:.1f}s)")

    # 6. Тоггл "Контрольная резолюция"
    if settings.get("control_resolution", True):
        log.info(f"--- ШАГ 6/12: тоггл 'Контрольная резолюция' ---")
        toggle_switch(driver, "Контрольная резолюция", "true")
        time.sleep(0.5)
        # 7. Дата контрольного этапа
        log.info(f"--- ШАГ 7/12: дата контрольного этапа ---")
        set_stage_date(driver,
            settings.get("workdays", cfg.DEFAULTS["workdays"]))
        log.debug(f"  Шаги 6-7 OK ({time.monotonic()-t_start:.1f}s)")

    # 8. Исполнитель
    log.info(f"--- ШАГ 8/12: исполнитель = {doc['executor']} ---")
    if not fill_executor(driver, doc['executor']):
        log.error(f"Row {doc['row_idx']}: не получилось ввести исполнителя")
        # Закроем диалог чтобы продолжить — но НЕ сохраним
        close_open_modals(driver)
        return False
    log.debug(f"  Шаг 8 OK ({time.monotonic()-t_start:.1f}s)")

    # 9. Клик "Добавить"
    log.info(f"--- ШАГ 9/12: кнопка 'Добавить' ---")
    add_btn = _wait_button_enabled(driver, "add_btn", timeout=10)
    if not add_btn:
        log.error("Кнопка 'Добавить' не активировалась")
        close_open_modals(driver)
        return False
    click(driver, add_btn, "Добавить")
    time.sleep(1.5)
    log.debug(f"  Шаг 9 OK ({time.monotonic()-t_start:.1f}s)")

    # 10. Клик "Сохранить и отправить"
    log.info(f"--- ШАГ 10/12: 'Сохранить и отправить' + confirm ---")
    if not submit_resolution(driver):
        log.error("Не удалось сохранить и отправить")
        return False
    log.debug(f"  Шаг 10 OK ({time.monotonic()-t_start:.1f}s)")

    # 11. Закрыть карточку (если осталась открыта)
    log.info(f"--- ШАГ 11/12: закрытие карточки ---")
    close_card_after_resolution(driver)
    log.debug(f"  Шаг 11 OK ({time.monotonic()-t_start:.1f}s)")

    # 12. Очистить фильтр для следующей итерации
    log.info(f"--- ШАГ 12/12: очистка фильтра ---")
    clear_filter(driver)
    log.debug(f"  Шаг 12 OK")
    log.info(f"ДОКУМЕНТ {index}/{total} ОБРАБОТАН за {time.monotonic()-t_start:.1f}s")
    return True


# ============================================================
# MAIN
# ============================================================

def _choose_xlsx(base_dir):
    files = [f for f in os.listdir(base_dir) if f.lower().endswith('.xlsx')]
    if not files:
        log.error(f"Нет .xlsx в {base_dir}")
        input("Enter...")
        sys.exit(1)
    # Сортируем — файлы с '_резолюции' в имени идут первыми (наш формат)
    files.sort(key=lambda f: (0 if 'резолюции' in f.lower() else 1, f))
    if len(files) == 1:
        log.info(f"Файл: {files[0]}")
        return os.path.join(base_dir, files[0])
    print(f"\nНайдено {len(files)} xlsx-файлов:")
    for i, f in enumerate(files, 1):
        marker = ' ← рекомендую' if 'резолюции' in f.lower() and i == 1 else ''
        print(f"  {i}. {f}{marker}")
    choice = input("Выбери номер [1]: ").strip() or "1"
    try:
        return os.path.join(base_dir, files[int(choice) - 1])
    except (ValueError, IndexError):
        log.error("Неверный выбор")
        sys.exit(1)


def _is_port_open(host, port, timeout=0.5):
    import socket
    try:
        with socket.create_connection((host, port), timeout=timeout):
            return True
    except (socket.error, OSError):
        return False


def _start_browser(base_dir):
    driver_path = os.path.join(base_dir, "msedgedriver.exe")
    if not os.path.exists(driver_path):
        log.error(f"msedgedriver.exe не найден в {base_dir}")
        input("Enter...")
        sys.exit(1)

    service = EdgeService(executable_path=driver_path)
    debugger_port = settings.get("debugger_port")
    if debugger_port and _is_port_open("127.0.0.1", int(debugger_port)):
        try:
            options = EdgeOptions()
            options.add_experimental_option("debuggerAddress",
                f"127.0.0.1:{debugger_port}")
            driver = webdriver.Edge(service=service, options=options)
            log.info(f"Подключился к открытому Edge на :{debugger_port} "
                     f"(URL: {driver.current_url or '?'})")
            return driver, True
        except Exception as e:
            log.warning(f"Не удалось подключиться: {e}")
    elif debugger_port:
        log.info(f"Edge не запущен с debug-портом {debugger_port}. Стартую новый.")

    options = EdgeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--auth-server-whitelist=*.interrao.ru")
    options.add_argument("--auth-negotiate-delegate-whitelist=*.interrao.ru")
    options.add_argument("--log-level=3")
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    return webdriver.Edge(service=service, options=options), False


def _wait_profile_loaded(driver, max_wait=120):
    """Ждёт готовности АСУД для resolutions: только readyState + кнопка
    'Создать документ' (как индикатор что главная отрисована).
    Таблицу с задачами НЕ ждём — у Есиной inbox может быть пустым."""
    log.info("Жду готовности АСУД...")
    try:
        WebDriverWait(driver, max_wait).until(
            lambda d: d.execute_script("return document.readyState === 'complete'"))
    except Exception:
        log.warning("readyState не complete")

    try:
        WebDriverWait(driver, max_wait).until(
            EC.element_to_be_clickable((By.ID, "mainscreen-create-button")))
        log.info("АСУД готов")
    except Exception:
        log.warning("Кнопка 'Создать документ' не появилась — продолжаю")
    time.sleep(1)


def _go_to_asud(driver, url, attached):
    try:
        cur = (driver.current_url or "").lower()
    except Exception:
        cur = ""
    if attached and "asudik" in cur:
        log.info(f"Использую вкладку АСУД: {driver.current_url}")
    else:
        log.info(f"Открываю {url}")
        driver.get(url)
    _wait_profile_loaded(driver)


def _maybe_quit(driver, attached):
    if attached:
        log.info("Оставляю Edge открытым")
    else:
        try:
            driver.quit()
            log.info("Браузер закрыт")
        except Exception as e:
            log.warning(f"Ошибка закрытия: {e}")


def main():
    global settings
    settings = cfg.load()

    log.info("=" * 50)
    log.info("АСУД ИК — выдача резолюций (под Халецкой)")
    log.info("=" * 50)

    base_dir = cfg.get_base_dir()
    _attach_file_logger(base_dir)
    excel_path = _choose_xlsx(base_dir)
    log.info(f"Реестр: {excel_path}")

    docs = load_excel(excel_path)
    if not docs:
        log.error("Реестр пуст")
        input("Enter...")
        sys.exit(1)

    # Превью
    print(f"\nПервые 5:")
    for i, d in enumerate(docs[:5], 1):
        flag = '✓' if d['executor'] else '!'
        print(f"  {i}. [{flag}] {d.get('appeal_no') or '???':>10} | "
              f"{(d.get('executor') or 'ИСПОЛНИТЕЛЬ?')[:30]:30} | "
              f"{d.get('subject', '')[:40]}")
    no_executor = sum(1 for d in docs if not d['executor'])
    print(f"\nВсего: {len(docs)} (без исполнителя: {no_executor} — будут пропущены)")

    if input("Начать? (да/нет): ").strip().lower() not in ("да", "д", "y", "yes", ""):
        print("Отменено.")
        sys.exit(0)

    driver, attached = _start_browser(base_dir)
    try:
        url = settings.get("asud_url", cfg.DEFAULTS["asud_url"])
        _go_to_asud(driver, url, attached)

        # Переключение учётки
        if not switch_account(driver,
                settings.get("target_account", cfg.DEFAULTS["target_account"])):
            log.error("Не удалось переключиться на учётку. Прерываю.")
            input("Enter...")
            return

        # Сайдбар → "На резолюцию"
        click_sidebar_section(driver,
            settings.get("sidebar_section", cfg.DEFAULTS["sidebar_section"]))
        time.sleep(2)

        done, err, skip = 0, 0, 0
        for i, doc in enumerate(docs, 1):
            try:
                ok = process_one(driver, doc, i, len(docs))
                if ok:
                    done += 1
                else:
                    skip += 1
            except Exception as e:
                log.error(f"ОШИБКА документ {i}: {e}")
                err += 1
                # Возврат в список — обновим страницу
                try:
                    driver.get(url)
                    _wait_profile_loaded(driver)
                    switch_account(driver,
                        settings.get("target_account", cfg.DEFAULTS["target_account"]))
                    click_sidebar_section(driver,
                        settings.get("sidebar_section", cfg.DEFAULTS["sidebar_section"]))
                    time.sleep(2)
                except Exception:
                    pass

        elapsed_seconds = time.monotonic() - start_time
        elapsed = timedelta(seconds=int(elapsed_seconds))
        avg = (timedelta(seconds=int(elapsed_seconds / done))
               if done else None)
        summary = [
            "",
            "=" * 60,
            "ГОТОВО!",
            f"  Обработано:  {done} / {len(docs)}",
            f"  Пропущено:   {skip}",
            f"  Ошибок:      {err}",
            f"  Затрачено:   {elapsed}" + (f"  (в среднем {avg}/док)" if avg else ""),
            "=" * 60,
        ]
        for line in summary:
            log.info(line)
            print(line)
        input("\nEnter для закрытия...")
    except Exception as e:
        log.error(f"Ошибка: {e}")
        input("Enter...")
    finally:
        _maybe_quit(driver, attached)


if __name__ == "__main__":
    main()
