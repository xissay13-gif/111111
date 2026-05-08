"""
app.py — Единая точка входа для АСУД-автоматизации создания документов.

Поддерживаемые режимы:
  • mix         — создание + регистрация + На резолюцию + .msg по Link
  • auto-create — создание + регистрация + На резолюцию (без .msg по Link)
  • smart       — создание ТОЛЬКО как черновик + .msg по Link (без регистрации,
                  корреспондент = «Неизвестный Неизвестный Неизвестный»)
  • email       — создание прямо из .msg-писем (без xlsx-реестра): рекурсивно
                  обходит папку, парсит письма, ФИО абонента берёт из тела.

Выдача резолюций — отдельный exe, ветка clean-resolutions.

Запуск:
  python app.py                       # auto-detect режима по xlsx
  python app.py --mode=mix
  python app.py --mode=auto-create
  python app.py --mode=smart
  python app.py --mode=email
  python app.py --xlsx=path.xlsx --mode=...
  python app.py --folder=path         # → авто-режим email
  python app.py --headless            # фоновый режим (Edge без GUI)

Auto-detect:
  • Передан --folder или --mode=email                         → email
  • Лист содержит колонку 'Link' (старый mix-формат)          → mix
  • Лист 'результат' (новый формат) или Subject/корреспондент → auto-create
  Smart НЕ определяется автоматически (тот же xlsx что у mix), нужен --mode=smart
"""

import argparse
import logging
import os
import sys

import openpyxl

from shared import config as cfg

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
    handlers=[logging.StreamHandler()],
)
log = logging.getLogger("asud.app")


_MODE_DESCRIPTIONS = {
    'mix':         'Создание + регистрация + На резолюцию + .msg по Link',
    'auto-create': 'Создание + регистрация + На резолюцию (без .msg)',
    'smart':       'Создание как черновик + .msg (без регистрации, фикс. корреспондент)',
    'email':       'Создание прямо из .msg-писем (без xlsx-реестра)',
}
_MODES = ['mix', 'auto-create', 'smart', 'email']


def detect_mode(xlsx_path):
    """Определяет режим по структуре xlsx (mix или auto-create)."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        # Лист 'результат' — новый auto-create формат
        if 'результат' in wb.sheetnames:
            wb.close()
            return 'auto-create'
        ws = wb.active
        headers = [str(c.value or '').lower()
                   for c in next(ws.iter_rows(max_row=1))]
        wb.close()
        # Колонка 'link' → mix-flow (есть .msg по ссылке)
        if any('link' in h for h in headers):
            return 'mix'
    except Exception as e:
        log.warning(f"Не удалось прочитать {xlsx_path} для auto-detect: {e}")

    return 'auto-create'


def pick_mode(xlsx_path):
    """Интерактивный выбор режима с подсказкой auto-detect.
    Email-режим в этом меню не показываем — он работает с папкой, не с xlsx."""
    xlsx_modes = [m for m in _MODES if m != 'email']
    auto = detect_mode(xlsx_path)
    print(f"\nРеестр: {os.path.basename(xlsx_path)}")
    print("Какой процесс запустить?")
    for i, m in enumerate(xlsx_modes, 1):
        marker = '  ← рекомендую (auto-detect)' if m == auto else ''
        print(f"  {i}. {m:11} — {_MODE_DESCRIPTIONS[m]}{marker}")
    print(f"\n[Enter] = {auto}  (рекомендуется по реестру)")
    choice = input(f"Номер режима (1-{len(xlsx_modes)}) или Enter: ").strip()
    if not choice:
        return auto
    try:
        return xlsx_modes[int(choice) - 1]
    except (ValueError, IndexError):
        log.warning(f"Неверный выбор '{choice}' — использую {auto}")
        return auto


def pick_source():
    """Интерактивный верхне-уровневый выбор: реестр (xlsx) или папка с .msg.
    Вызывается когда юзер не передал ни --mode, ни --xlsx, ни --folder."""
    print("\nЧто обработать?")
    print("  1. Реестр (.xlsx)        — режимы mix / auto-create / smart")
    print("  2. Папку с .msg-письмами — режим email")
    choice = input("Номер (1-2) [1]: ").strip()
    return 'email' if choice == '2' else 'xlsx'


def main():
    parser = argparse.ArgumentParser(
        description="АСУД ИК — автоматизация документооборота")
    parser.add_argument('--mode', choices=_MODES,
                        help="Режим работы (если не задан — auto-detect по xlsx)")
    parser.add_argument('--xlsx', help="Путь к реестру (если не задан — спрашиваем)")
    parser.add_argument('--folder', help="Папка с .msg-письмами (включает режим email)")
    parser.add_argument('--watch', action='store_true',
                        help="Непрерывный мониторинг папки (только email-режим). "
                             "Ctrl+C — остановка после текущего документа")
    parser.add_argument('--headless', action='store_true',
                        help="Запустить Edge без GUI (фоновый режим, требует Стадии 1б)")
    args = parser.parse_args()

    base_dir = cfg.get_base_dir()
    log.info(f"Базовая папка: {base_dir}")
    if args.headless:
        os.environ['ASUD_HEADLESS'] = '1'
        log.info("Режим: HEADLESS (Edge без GUI)")

    # === Определяем источник: email vs xlsx =================================
    if args.folder or args.mode == 'email':
        source = 'email'
    elif args.xlsx or args.mode in ('mix', 'auto-create', 'smart'):
        source = 'xlsx'
    else:
        # Ничего не указано — спрашиваем
        source = pick_source()

    # === EMAIL-источник =====================================================
    if source == 'email':
        if args.folder:
            os.environ['ASUD_EMAIL_FOLDER'] = args.folder
            log.info(f"Режим: email (папка: {args.folder})")
        else:
            log.info("Режим: email")
        os.environ['ASUD_MODE'] = 'email'
        if args.watch:
            os.environ['ASUD_WATCH'] = '1'
            log.info("Режим: WATCH (непрерывный мониторинг, Ctrl+C для остановки)")
            from flows.email import daemon_main as flow_main
        else:
            from flows.email import main as flow_main
        flow_main()
        return

    # === XLSX-источник: mix / auto-create / smart ===========================
    # Если xlsx не указан — выбираем интерактивно
    xlsx_path = args.xlsx
    if not xlsx_path:
        # Сначала пробуем выбрать из base_dir
        candidates = [f for f in os.listdir(base_dir)
                      if f.lower().endswith('.xlsx') and not f.startswith('~')]
        if len(candidates) == 1:
            xlsx_path = os.path.join(base_dir, candidates[0])
            log.info(f"Найден единственный xlsx: {candidates[0]}")
        elif len(candidates) > 1:
            print("\nДоступные реестры:")
            for i, name in enumerate(candidates, 1):
                print(f"  {i}. {name}")
            choice = input(f"Выбери (1-{len(candidates)}): ").strip()
            try:
                xlsx_path = os.path.join(base_dir, candidates[int(choice) - 1])
            except (ValueError, IndexError):
                log.error("Неверный выбор")
                sys.exit(1)
        else:
            log.error(f"Не нашёл .xlsx в {base_dir}")
            sys.exit(1)

    # Определяем режим: явный флаг → интерактивный выбор → auto-detect
    if args.mode:
        mode = args.mode
        log.info(f"Режим: {mode} (через --mode)")
    else:
        mode = pick_mode(xlsx_path)
        log.info(f"Режим выбран: {mode}")

    # Запуск соответствующего flow
    if mode == 'mix':
        from flows.mix import main as flow_main
    elif mode == 'auto-create':
        from flows.auto_create import main as flow_main
    elif mode == 'smart':
        from flows.smart import main as flow_main
    else:
        log.error(f"Неизвестный режим: {mode}")
        sys.exit(1)

    # Каждый flow.main() сам читает xlsx — передаём через env-переменную
    # (минимальный contract без переписывания main каждого flow).
    # ASUD_MODE передаём для отображения в превью flow'а.
    os.environ['ASUD_XLSX'] = xlsx_path
    os.environ['ASUD_MODE'] = mode
    flow_main()


if __name__ == "__main__":
    main()
