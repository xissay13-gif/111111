"""
auto_create_correspondent.py — Пакетное создание Входящих документов в АСУД ИК.

Читает Excel (B=Содержание, C=Корреспондент), автоматически создаёт
корреспондента если не найден. Регистрирует + На резолюцию.

Модули:
  config.py       — настройки (+ config.json)
  ui.py           — Selenium UI-хелперы
  correspondent.py — создание корреспондентов
  attachments.py  — прикрепление файлов (пустышка)
"""

import os
import re
import sys
import time
import logging
from datetime import date, datetime, timedelta

import openpyxl
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import config as cfg
from ui import (click, wait_and_click, find_input_near_label,
                wait_asud_loaded, wait_modal_closed, close_open_modals, js_set_value)
from correspondent import (fill_correspondent_field, match_correspondent)
from attachments import get_dummy_msg, attach_content


# ================= LOGGING =================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
)
log = logging.getLogger("asud")
start_time = time.monotonic()


# ================= EXCEL =================

def load_excel(file_path):
    """Читает Excel. Колонки: B=Содержание, C=Корреспондент."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        content = row[1]   # B
        corr = row[2]      # C
        if content and corr:
            rows.append({
                "содержание": str(content).strip(),
                "корреспондент": str(corr).strip(),
            })
    wb.close()
    log.info(f"Загружено: {len(rows)} документов")
    return rows


# ================= FORM FILLING =================

def fill_text(driver, text):
    """Заполняет краткое содержание (textarea)."""
    try:
        areas = driver.find_elements(By.TAG_NAME, "textarea")
        visible = [a for a in areas if a.is_displayed()]
        if visible:
            visible[0].click()
            time.sleep(0.3)
            visible[0].clear()
            visible[0].send_keys(text)
            log.info("Краткое содержание заполнено")
        else:
            log.warning("Textarea не найдена")
    except Exception as e:
        log.error(f"Ошибка содержания: {e}")
    time.sleep(0.5)


def fill_corr_number(driver, index):
    """Заполняет 'Номер у корреспондента' = 'б/н (N)'."""
    value = f"б/н ({index})"
    inp = find_input_near_label(driver, "Номер у корреспондента")
    if not inp:
        log.warning("Поле 'Номер у корреспондента' не найдено")
        return
    try:
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", inp)
        time.sleep(0.3)
        inp.click()
        time.sleep(0.3)
        inp.clear()
        inp.send_keys(value)
        inp.send_keys(Keys.TAB)
        log.info(f"Номер: {value}")
    except Exception:
        js_set_value(driver, inp, value)
        log.info(f"Номер (JS): {value}")


def fill_corr_date(driver):
    """Заполняет 'Дата у корреспондента' = сегодня."""
    from datetime import date
    today = date.today().strftime("%d.%m.%Y")
    labels = driver.find_elements(By.XPATH,
        "//*[normalize-space(text())='Дата у корреспондента']")
    inp = None
    for label in labels:
        try:
            if not label.is_displayed():
                continue
            for level in range(1, 6):
                parent = label
                for _ in range(level):
                    parent = parent.find_element(By.XPATH, "..")
                inputs = parent.find_elements(By.CSS_SELECTOR, "input[type='text']")
                visible = [i for i in inputs
                           if i.is_displayed() and i.get_attribute("readonly") is None]
                if visible:
                    inp = visible[0]
                    break
            if inp:
                break
        except Exception:
            continue
    if not inp:
        log.warning("Поле 'Дата у корреспондента' не найдено")
        return
    try:
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", inp)
        driver.execute_script("arguments[0].focus(); arguments[0].click();", inp)
        time.sleep(0.3)
        inp.send_keys(Keys.CONTROL + "a")
        time.sleep(0.2)
        inp.send_keys(Keys.DELETE)
        time.sleep(0.2)
        inp.send_keys(today)
        inp.send_keys(Keys.TAB)
        log.info(f"Дата: {today}")
    except Exception:
        js_set_value(driver, inp, today)
        log.info(f"Дата (JS): {today}")


def fill_delivery_method(driver):
    """Выбирает 'Электронная почта'."""
    target = settings.get("delivery_method", "Электронная почта")
    labels = driver.find_elements(By.XPATH,
        "//*[normalize-space(text())='Способ получения']")
    trigger = None
    for label in labels:
        try:
            if not label.is_displayed():
                continue
            for level in range(1, 8):
                parent = label
                for _ in range(level):
                    parent = parent.find_element(By.XPATH, "..")
                for sel in ["input[type='text']", "div[class*='trigger']", "img[class*='trigger']"]:
                    try:
                        el = parent.find_element(By.CSS_SELECTOR, sel)
                        if el.is_displayed():
                            trigger = el
                            break
                    except Exception:
                        continue
                if trigger:
                    break
            if trigger:
                break
        except Exception:
            continue
    if not trigger:
        log.warning("Поле 'Способ получения' не найдено")
        return
    click(driver, trigger, "Способ получения")
    time.sleep(1.5)
    for attempt in range(3):
        candidates = driver.find_elements(By.XPATH, f"//*[contains(text(),'{target}')]")
        for c in candidates:
            try:
                if c.is_displayed() and c.tag_name.lower() != 'input':
                    click(driver, c, target)
                    log.info(f"Способ получения: {target}")
                    return
            except Exception:
                continue
        time.sleep(1)
    log.warning(f"'{target}' не найдена")


def add_addressee(driver, person_name):
    """Добавляет адресата через combobox."""
    inp = find_input_near_label(driver, "Адресаты")
    if not inp:
        log.warning("Поле адресата не найдено")
        return
    surname = person_name.split()[0]
    inp.click()
    time.sleep(0.5)
    inp.clear()
    time.sleep(0.3)
    for char in surname:
        inp.send_keys(char)
        time.sleep(0.1)
    time.sleep(2)
    results = driver.find_elements(By.XPATH, f"//*[contains(text(),'{surname}')]")
    all_r = [r for r in results if r.is_displayed() and r != inp and r.tag_name.lower() != 'input']
    if not all_r:
        inp.send_keys(Keys.ENTER)
        time.sleep(2)
        results = driver.find_elements(By.XPATH, f"//*[contains(text(),'{surname}')]")
        all_r = [r for r in results if r.is_displayed() and r != inp and r.tag_name.lower() != 'input']
    target = None
    for r in all_r:
        try:
            if match_correspondent(r.text, person_name):
                target = r
                break
        except Exception:
            continue
    if not target and all_r:
        target = all_r[0]
    if target:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", target)
        time.sleep(0.3)
        ActionChains(driver).move_to_element(target).pause(0.3).click().perform()
        time.sleep(1)
        log.info(f"Адресат: {person_name}")
    else:
        log.warning(f"Адресат не найден: {person_name}")


# ================= REGISTRATION =================

_ASUD_ID_RE = re.compile(
    r'\b([А-Я]{2,5}(?:/[А-Я0-9.-]+){2,})\b',
    re.UNICODE)


def _looks_like_asud_id(txt):
    if not txt:
        return False
    txt = txt.strip()
    if '/' not in txt or len(txt) < 6:
        return False
    if re.match(r'^\d{2}\.\d{2}\.\d{4}', txt):
        return False
    return True


def capture_asud_id(driver, timeout=15):
    """Читает регистрационный номер документа после регистрации.
    3 стратегии (ScreenHeader1/<b>; XPath с фильтром; regex по странице)."""
    end = time.monotonic() + timeout
    last_dump = 0
    while time.monotonic() < end:
        try:
            header = driver.find_element(By.CSS_SELECTOR,
                "[data-marker='ScreenHeader1']")
            for b in header.find_elements(By.CSS_SELECTOR, "b"):
                try:
                    txt = (b.text or "").strip()
                    if _looks_like_asud_id(txt):
                        log.info(f"  asud_id [s1]: {txt!r}")
                        return txt
                except Exception:
                    continue
        except Exception:
            pass
        try:
            for b in driver.find_elements(By.XPATH,
                    "//*[@data-marker='ScreenHeader1']//b"):
                txt = (b.text or "").strip()
                if _looks_like_asud_id(txt):
                    log.info(f"  asud_id [s2]: {txt!r}")
                    return txt
        except Exception:
            pass
        if time.monotonic() - last_dump > 3:
            last_dump = time.monotonic()
            try:
                headers = driver.find_elements(By.CSS_SELECTOR,
                    "[data-marker='ScreenHeader1']")
                log.info(f"  ждём asud_id... ScreenHeader1: {len(headers)}")
                if headers:
                    inner = (headers[0].text or "")[:200].replace('\n', ' | ')
                    log.info(f"    содержимое: {inner!r}")
            except Exception:
                pass
        time.sleep(0.3)
    try:
        body_text = driver.find_element(By.TAG_NAME, "body").text
        m = _ASUD_ID_RE.search(body_text)
        if m:
            log.info(f"  asud_id [s3 regex]: {m.group(1)!r}")
            return m.group(1)
    except Exception:
        pass
    log.warning("Регистрационный номер не захватили — пуст в output")
    return None


def register_and_resolve(driver, index, total):
    """Регистрирует + На резолюцию + Да. Возвращает asud_id (или None)."""
    log.info("Регистрирую...")
    registered = False
    asud_id = None
    try:
        btn = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,
                "#header-action-btn-register, [id*='header-action-btn-register']")))
        click(driver, btn, "Зарегистрировать")
        time.sleep(3)
        asud_id = capture_asud_id(driver, timeout=10)
        if asud_id:
            log.info(f"Документ {index}/{total} ЗАРЕГИСТРИРОВАН: {asud_id}")
        else:
            log.warning(f"Документ {index}/{total} ЗАРЕГИСТРИРОВАН (номер не захватили)")
        registered = True
    except Exception:
        try:
            btn = driver.find_element(By.XPATH, "//div[contains(text(),'Зарегистрировать')]")
            click(driver, btn, "Зарегистрировать (fallback)")
            time.sleep(3)
            registered = True
        except Exception as e:
            log.error(f"'Зарегистрировать' не найдена: {e}")

    if not registered:
        return None

    # На резолюцию
    res_btn = None
    for _ in range(10):
        try:
            btn = driver.find_element(By.ID, "header-action-btn-send_on_resolution")
            if btn.is_displayed():
                res_btn = btn
                break
        except Exception:
            pass
        try:
            for b in driver.find_elements(By.XPATH, "//*[contains(text(),'На резолюцию')]"):
                if b.is_displayed():
                    res_btn = b
                    break
        except Exception:
            pass
        if res_btn:
            break
        time.sleep(1)

    if res_btn:
        click(driver, res_btn, "На резолюцию")
        time.sleep(2)

        # Да
        yes_btn = None
        for _ in range(10):
            # 1) Точный id
            try:
                btn = driver.find_element(By.ID, "confirm_dialog_btn_yes")
                if btn.is_displayed():
                    yes_btn = btn
                    break
            except Exception:
                pass
            # 2) Substring id (GWT иногда добавляет префиксы/суффиксы)
            try:
                btn = driver.find_element(By.CSS_SELECTOR,
                    "[id*='confirm_dialog_btn_yes'], [id*='confirm'][id*='yes']")
                if btn.is_displayed():
                    yes_btn = btn
                    break
            except Exception:
                pass
            # 3) По тексту "Да"
            try:
                for b in driver.find_elements(By.XPATH, "//*[normalize-space(text())='Да']"):
                    if b.is_displayed():
                        yes_btn = b
                        break
            except Exception:
                pass
            if yes_btn:
                break
            time.sleep(1)

        if yes_btn:
            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", yes_btn)
                time.sleep(0.3)
            except Exception:
                pass
            clicked = False
            # 1) ActionChains
            try:
                ActionChains(driver).move_to_element(yes_btn).pause(0.3).click().perform()
                log.info(f"Клик 'Да' (ActionChains): id={yes_btn.get_attribute('id')}")
                clicked = True
            except Exception as e:
                log.info(f"ActionChains 'Да' не сработал: {e}")
            # 2) JS
            if not clicked:
                try:
                    driver.execute_script("arguments[0].click();", yes_btn)
                    log.info("Клик 'Да' (JS)")
                    clicked = True
                except Exception as e:
                    log.info(f"JS 'Да' не сработал: {e}")
            # 3) Нативный
            if not clicked:
                try:
                    yes_btn.click()
                    log.info("Клик 'Да' (native)")
                    clicked = True
                except Exception as e:
                    log.info(f"Native 'Да' не сработал: {e}")
            # 4) Enter — GWT-диалоги часто принимают его
            if clicked:
                try:
                    ActionChains(driver).send_keys(Keys.ENTER).perform()
                except Exception:
                    pass
            time.sleep(3)
            log.info(f"Документ {index}/{total} НА РЕЗОЛЮЦИИ")
        else:
            log.warning("Диалог 'Да' не появился за 10 сек")
    else:
        log.warning("'На резолюцию' не появилась")
    return asud_id


# ================= DOCUMENT FLOW =================

def create_one_document(driver, doc_data, index, total):
    """Создаёт один входящий документ."""
    log.info(f"{'='*50}")
    log.info(f"ДОКУМЕНТ {index}/{total}")
    log.info(f"Содержание: {doc_data['содержание'][:60]}...")
    log.info(f"Корреспондент: {doc_data['корреспондент']}")

    el = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
        EC.presence_of_element_located((By.ID, "mainscreen-create-button")))
    time.sleep(1)
    click(driver, el, "Создать документ")
    time.sleep(3)

    wait_and_click(driver, By.XPATH,
        "//div[contains(text(),'Входящий документ')]", "Входящий документ")
    time.sleep(1)

    subtype = settings.get("doc_subtype", "Письма, заявления и жалобы граждан, акционеров")
    short = subtype[:30]
    wait_and_click(driver, By.XPATH,
        f"//div[contains(text(),'{short}')] | //td[contains(text(),'{short}')]", subtype)
    time.sleep(0.5)

    wait_and_click(driver, By.XPATH,
        "//button[contains(text(),'Создать документ')] | //div[contains(text(),'Создать документ')]",
        "Создать документ")
    time.sleep(5)

    fill_text(driver, doc_data["содержание"])
    fill_correspondent_field(driver, doc_data["корреспондент"])
    fill_corr_number(driver, index)
    fill_corr_date(driver)

    for person in settings.get("addressees", ["Басманов Александр Владимирович"]):
        add_addressee(driver, person)
        time.sleep(0.5)

    fill_delivery_method(driver)
    time.sleep(0.5)

    try:
        save_btn = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            EC.element_to_be_clickable((By.ID, "header-save-btn")))
        click(driver, save_btn, "Сохранить")
        time.sleep(3)
        log.info(f"Документ {index}/{total} сохранён")
    except Exception as e:
        log.error(f"Ошибка сохранения: {e}")

    if doc_data.get("файл"):
        attach_content(driver, doc_data["файл"])
        wait_modal_closed(driver)

    asud_id = register_and_resolve(driver, index, total)

    time.sleep(2)
    try:
        close_btn = driver.find_element(By.ID, "header-close-btn")
        if close_btn.is_displayed():
            ActionChains(driver).move_to_element(close_btn).pause(0.3).click().perform()
            time.sleep(2)
    except Exception:
        pass

    try:
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "mainscreen-create-button")))
    except Exception:
        driver.get(settings.get("asud_url", cfg.DEFAULTS["asud_url"]))
        wait_asud_loaded(driver)
    return asud_id


# ================= OUTPUT XLSX (для clean-resolutions) =================

OKRUG_TO_FIO = {
    "САО": "Гренц Екатерина Александровна",
    "ЦАО": "Емельянова Татьяна Николаевна",
    "ОАО": "Рендюк Юлия Павловна",
    "ЛАО": "Вырва Елена Анатольевна",
    "КАО": "Кравец Татьяна Александровна",
}


def _output_xlsx_path(excel_path):
    base, _ext = os.path.splitext(excel_path)
    return base + "_резолюции.xlsx"


def _ensure_output_xlsx(path):
    if os.path.isfile(path):
        return
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Резолюции"
        ws.append(["ОПТС", "Округ", "ФИО", "Link"])
        for c in range(1, 5):
            ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
        widths = {1: 30, 2: 8, 3: 35, 4: 22}
        for col, w in widths.items():
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)].width = w
        ws.freeze_panes = "A2"
        wb.save(path)
    except Exception as e:
        log.warning(f"Не удалось создать {path}: {e}")


def _append_output_row(path, doc_data, asud_id):
    """Дописывает в output xlsx: ОПТС | Округ | ФИО | Link.
    В auto-create нет колонки Link — оставляем пусто; адрес парсится
    из 'Содержание' (что в auto-create заменяет TextBody)."""
    okrug = None
    try:
        from okrug_parser import okrug_from_textbody
        okrug = okrug_from_textbody(doc_data.get("содержание"),
                                     base_dir_fn=cfg.get_base_dir)
    except Exception as e:
        log.warning(f"okrug_parser упал: {e}")
    fio = OKRUG_TO_FIO.get(okrug) if okrug else None
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        ws.append([asud_id or "", okrug or "", fio or "", ""])
        wb.save(path)
        wb.close()
        log.info(f"  → {os.path.basename(path)}: "
                 f"{asud_id or '—'} | {okrug or '—'} | {fio or '—'}")
    except Exception as e:
        log.warning(f"Не удалось записать в {path}: {e}")


# ================= MAIN =================

settings = {}


def main():
    global settings
    settings = cfg.load()

    log.info("=" * 50)
    log.info("АСУД ИК — Входящие + автосоздание корреспондентов")
    log.info("=" * 50)

    base_dir = cfg.get_base_dir()

    xlsx_files = [f for f in os.listdir(base_dir) if f.lower().endswith('.xlsx')]
    if not xlsx_files:
        log.error(f"Нет .xlsx в {base_dir}")
        input("Enter...")
        sys.exit(1)
    elif len(xlsx_files) == 1:
        excel_path = os.path.join(base_dir, xlsx_files[0])
    else:
        print(f"\nНайдено {len(xlsx_files)} xlsx:")
        for i, f in enumerate(xlsx_files, 1):
            print(f"  {i}. {f}")
        choice = input("Номер: ").strip()
        try:
            excel_path = os.path.join(base_dir, xlsx_files[int(choice) - 1])
        except (ValueError, IndexError):
            log.error("Неверный выбор")
            sys.exit(1)

    msg_path = get_dummy_msg(base_dir)
    docs = load_excel(excel_path)
    for doc in docs:
        doc["файл"] = msg_path

    if not docs:
        log.error("Нет данных!")
        input("Enter...")
        sys.exit(1)

    print(f"\nПервые 5:")
    for i, d in enumerate(docs[:5], 1):
        print(f"  {i}. {d['корреспондент']} | {d['содержание'][:50]}...")
    print(f"\nВсего: {len(docs)}")

    if input("Начать? (да/нет): ").strip().lower() not in ("да", "д", "y", "yes", ""):
        sys.exit(0)

    driver_path = os.path.join(base_dir, "msedgedriver.exe")
    if not os.path.exists(driver_path):
        log.error(f"msedgedriver.exe не найден")
        input("Enter...")
        sys.exit(1)

    options = EdgeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--auth-server-whitelist=*.interrao.ru")
    options.add_argument("--auth-negotiate-delegate-whitelist=*.interrao.ru")
    options.add_argument("--log-level=3")
    options.add_experimental_option('excludeSwitches', ['enable-logging'])

    driver = webdriver.Edge(service=EdgeService(executable_path=driver_path), options=options)

    try:
        url = settings.get("asud_url", cfg.DEFAULTS["asud_url"])
        driver.get(url)
        wait_asud_loaded(driver)

        # Output xlsx: ОПТС|Округ|ФИО|Link рядом с реестром
        output_path = _output_xlsx_path(excel_path)
        _ensure_output_xlsx(output_path)
        log.info(f"Output xlsx: {output_path}")

        done_count, err_count = 0, 0
        for i, doc in enumerate(docs, 1):
            try:
                asud_id = create_one_document(driver, doc, i, len(docs))
                _append_output_row(output_path, doc, asud_id)
                done_count += 1
            except Exception as e:
                log.error(f"ОШИБКА документ {i}: {e}")
                err_count += 1
                driver.get(url)
                wait_asud_loaded(driver)

        elapsed_seconds = time.monotonic() - start_time
        elapsed = timedelta(seconds=int(elapsed_seconds))
        avg = timedelta(seconds=int(elapsed_seconds / done_count)) if done_count else None

        summary = [
            "",
            "=" * 60,
            f"ГОТОВО!",
            f"  Обработано: {done_count} / {len(docs)}",
            f"  Ошибок:     {err_count}",
            f"  Затрачено:  {elapsed}" + (f"  (в среднем {avg}/док)" if avg else ""),
            "=" * 60,
        ]
        for line in summary:
            log.info(line)
            print(line)

        input("\nEnter для закрытия...")
    except Exception as e:
        log.error(f"Ошибка: {e}")
        input("Enter...")
    finally:
        driver.quit()


if __name__ == "__main__":
    main()
