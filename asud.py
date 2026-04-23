"""
asud.py — Пакетное создание Входящих документов в АСУД ИК.

При запуске выбирается сценарий:

  1. auto-create — Excel: B=Содержание, C=Корреспондент.
     Корреспондент берётся из Excel; если его нет в АСУД — создаётся
     карточка физ. лица. Вид документа — фиксированный (config.doc_subtype).
     Вложение — dummy .msg рядом с exe. Регистрирует + На резолюцию.

  2. mix — Excel: Лист2 (A=Link, B=Subject, C=TextBody, D=Тип, E=To).
     ФИО корреспондента извлекается из TextBody (regex по маркеру "Ф.И.О.:"
     или 3-словному паттерну с отчеством). Вид документа — из колонки D
     через DOC_TYPE_MAP. Вложение — .msg по Link из outlook_dir
     (рекурсивно по подпапкам). Если ФИО найдено → регистрация;
     если не найдено → черновик + WARNING. Resume после крэша
     через per-xlsx state-файл.

Модули:
  config.py        — настройки (+ config.json)
  ui.py            — Selenium UI-хелперы
  correspondent.py — выбор/создание корреспондента + extract_fio_from_text
  attachments.py   — поиск и прикрепление файлов
"""

import os
import re
import sys
import json
import time
import logging
from datetime import date, datetime, timedelta

import openpyxl
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import config as cfg
from ui import (click, wait_and_click, find_input_near_label,
                wait_asud_loaded, wait_modal_closed, close_open_modals, js_set_value)
from correspondent import (fill_correspondent_field, match_strict, fio_to_initials,
                           extract_fio_from_text)
from attachments import find_msg_by_link, get_dummy_msg, attach_content, move_to_done


# ================= LOGGING =================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
    handlers=[logging.StreamHandler()],
)
log = logging.getLogger("asud")
start_time = time.monotonic()


# ================= EXCEL =================

def _clean_body(text):
    """Очищает TextBody от служебных строк (Original Message, ВНИМАНИЕ…)."""
    if not text:
        return ""
    t = str(text).replace('_x000D_', '\n')
    lines = t.split('\n')
    cleaned = []
    for line in lines:
        s = line.strip()
        if re.search(r'внимание!?\s*письмо\s+было\s+отправлено\s+внешним', s, re.IGNORECASE):
            continue
        if re.match(r'^-{3,}\s*Original\s*Message\s*-{3,}$', s, re.IGNORECASE):
            continue
        cleaned.append(line)
    t = '\n'.join(cleaned)
    t = re.sub(r'\n\s*\n\s*\n+', '\n\n', t)
    return t.strip()


# ================= STATE (resume после крэша) =================

def _link_key(link):
    """Стабильный строковый ключ из Link (для state-файла)."""
    if link is None:
        return ""
    if isinstance(link, (datetime, date)):
        try:
            return link.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(link)
    return str(link).strip()


def _state_path(xlsx_path):
    """Путь к state-файлу рядом с exe, привязан к имени Excel."""
    name = os.path.splitext(os.path.basename(xlsx_path))[0]
    # Безопасное имя файла
    safe = re.sub(r'[^\w.\-]+', '_', name)
    return os.path.join(cfg.get_base_dir(), f"mix_state_{safe}.json")


def load_state(xlsx_path):
    """Загружает set обработанных Link-ключей."""
    path = _state_path(xlsx_path)
    if not os.path.isfile(path):
        return set()
    try:
        with open(path, encoding='utf-8') as f:
            data = json.load(f)
        return set(data.get('processed', []))
    except Exception as e:
        log.warning(f"Не удалось прочитать state {path}: {e}")
        return set()


def save_state(xlsx_path, processed_set):
    """Атомарно перезаписывает state-файл."""
    path = _state_path(xlsx_path)
    try:
        tmp = path + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump({'processed': sorted(processed_set)}, f,
                      ensure_ascii=False, indent=2)
        os.replace(tmp, path)
    except Exception as e:
        log.warning(f"Не удалось сохранить state {path}: {e}")


def load_excel_mix(file_path):
    """[MIX] Читает Лист2. Колонки: A=Link, B=Subject, C=TextBody, D=Тип.

    ФИО корреспондента извлекается из TextBody.
    Если не найдено — ставится заглушка (unknown_correspondent) и флаг corr_found=False
    → такой документ потом останется в черновиках.
    """
    sheet_name = settings.get("sheet_name", cfg.DEFAULTS["sheet_name"])
    unknown = settings.get("unknown_correspondent", cfg.DEFAULTS["unknown_correspondent"])

    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        log.warning(f"Лист '{sheet_name}' не найден, использую активный: {wb.active.title}")
        ws = wb.active

    rows = []
    skipped = 0
    unknown_rows = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
        if not row or len(row) < 4:
            skipped += 1
            continue
        link = row[0]
        subject = row[1]
        body = row[2]
        type_idx = row[3]

        if not subject:
            skipped += 1
            continue
        try:
            type_idx = int(type_idx) if type_idx is not None else 0
        except (ValueError, TypeError):
            type_idx = 0
        if type_idx == 0 or type_idx not in cfg.DOC_TYPE_MAP:
            skipped += 1
            continue

        clean_subject = re.sub(r'^(FW:|RE:|Fwd:)\s*', '',
                               str(subject).strip(), flags=re.IGNORECASE)
        body_clean = _clean_body(body) if body else clean_subject

        fio, fio_src = extract_fio_from_text(body)
        if fio:
            correspondent = fio
            corr_found = True
        else:
            correspondent = unknown
            corr_found = False
            unknown_rows.append((row_idx, clean_subject))

        rows.append({
            "row_idx": row_idx,
            "содержание": body_clean,
            "корреспондент": correspondent,
            "корр_найден": corr_found,
            "корр_источник": fio_src,
            "тема": clean_subject,
            "тип_индекс": type_idx,
            "тип_название": cfg.DOC_TYPE_MAP[type_idx],
            "link": link,
        })
    wb.close()

    log.info(f"Загружено: {len(rows)}, пропущено: {skipped}")
    log.info(f"  ФИО найдено: {sum(1 for r in rows if r['корр_найден'])}")
    log.info(f"  ФИО НЕ найдено: {len(unknown_rows)} (уйдут в черновики)")
    for ri, subj in unknown_rows:
        log.info(f"    → Row {ri}: {subj[:60]}")

    return rows


def load_excel_auto_create(file_path):
    """[AUTO-CREATE] Читает активный лист. Колонки: B=Содержание, C=Корреспондент."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row,
                                               values_only=True), 2):
        if not row or len(row) < 3:
            continue
        content = row[1]  # B
        corr = row[2]     # C
        if content and corr:
            rows.append({
                "row_idx": row_idx,
                "содержание": str(content).strip(),
                "корреспондент": str(corr).strip(),
            })
    wb.close()
    log.info(f"Загружено: {len(rows)} документов")
    return rows


# ================= FORM FILLING =================

def fill_text(driver, text):
    """Заполняет краткое содержание (textarea)."""
    try:
        areas = driver.find_elements(By.TAG_NAME, "textarea")
        visible = [a for a in areas if a.is_displayed()]
        if visible:
            visible[0].click()
            time.sleep(0.3)
            visible[0].clear()
            visible[0].send_keys(text)
            log.info("Краткое содержание заполнено")
        else:
            log.warning("Textarea не найдена")
    except Exception as e:
        log.error(f"Ошибка заполнения содержания: {e}")
    time.sleep(0.5)


def fill_corr_number(driver, value):
    """Заполняет 'Номер у корреспондента' готовым значением (строкой)."""
    inp = find_input_near_label(driver, "Номер у корреспондента")
    if not inp:
        log.warning("Поле 'Номер у корреспондента' не найдено")
        return
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
        time.sleep(0.3)
        inp.click()
        time.sleep(0.3)
        inp.clear()
        time.sleep(0.2)
        inp.send_keys(value)
        inp.send_keys(Keys.TAB)
        log.info(f"Номер: {value}")
    except Exception:
        js_set_value(driver, inp, value)
        log.info(f"Номер (JS): {value}")


def fill_corr_date(driver):
    """Заполняет 'Дата у корреспондента' = сегодня."""
    today = date.today().strftime("%d.%m.%Y")
    labels = driver.find_elements(By.XPATH,
        "//*[normalize-space(text())='Дата у корреспондента']")
    inp = None
    for label in labels:
        try:
            if not label.is_displayed():
                continue
            for level in range(1, 6):
                parent = label
                for _ in range(level):
                    parent = parent.find_element(By.XPATH, "..")
                inputs = parent.find_elements(By.CSS_SELECTOR, "input[type='text']")
                visible = [i for i in inputs
                           if i.is_displayed() and i.get_attribute("readonly") is None]
                if visible:
                    inp = visible[0]
                    break
            if inp:
                break
        except Exception:
            continue

    if not inp:
        log.warning("Поле 'Дата у корреспондента' не найдено")
        return
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
        driver.execute_script("arguments[0].focus(); arguments[0].click();", inp)
        time.sleep(0.3)
        inp.send_keys(Keys.CONTROL + "a")
        time.sleep(0.2)
        inp.send_keys(Keys.DELETE)
        time.sleep(0.2)
        inp.send_keys(today)
        inp.send_keys(Keys.TAB)
        log.info(f"Дата: {today}")
    except Exception:
        js_set_value(driver, inp, today)
        log.info(f"Дата (JS): {today}")


def fill_delivery_method(driver):
    """Выбирает 'Электронная почта' в 'Способ получения'."""
    target_text = settings.get("delivery_method", "Электронная почта")
    labels = driver.find_elements(By.XPATH,
        "//*[normalize-space(text())='Способ получения']")
    trigger = None
    for label in labels:
        try:
            if not label.is_displayed():
                continue
            for level in range(1, 8):
                parent = label
                for _ in range(level):
                    parent = parent.find_element(By.XPATH, "..")
                inputs = parent.find_elements(By.CSS_SELECTOR, "input[type='text']")
                for i in inputs:
                    if i.is_displayed():
                        trigger = i
                        break
                if trigger:
                    break
                for sel in ["div[class*='trigger']", "img[class*='trigger']"]:
                    try:
                        el = parent.find_element(By.CSS_SELECTOR, sel)
                        if el.is_displayed():
                            trigger = el
                            break
                    except Exception:
                        continue
                if trigger:
                    break
            if trigger:
                break
        except Exception:
            continue

    if not trigger:
        log.warning("Поле 'Способ получения' не найдено")
        return

    click(driver, trigger, "Способ получения")
    time.sleep(1.5)

    option = None
    for _ in range(3):
        candidates = driver.find_elements(By.XPATH,
            f"//*[contains(text(),'{target_text}')]")
        for c in candidates:
            try:
                if c.is_displayed() and c.tag_name.lower() != 'input':
                    option = c
                    break
            except Exception:
                continue
        if option:
            break
        time.sleep(1)

    if option:
        click(driver, option, target_text)
        log.info(f"Способ получения: {target_text}")
    else:
        log.warning(f"'{target_text}' не найдена в списке")


def add_addressee(driver, person_name):
    """Добавляет одного адресата через combobox."""
    inp = find_input_near_label(driver, "Адресаты")
    if not inp:
        log.warning("Поле адресата не найдено")
        return

    surname = person_name.split()[0]
    inp.click()
    time.sleep(0.5)
    inp.clear()
    time.sleep(0.3)
    for char in surname:
        inp.send_keys(char)
        time.sleep(0.1)
    time.sleep(2)

    from correspondent import match_correspondent
    results = driver.find_elements(By.XPATH, f"//*[contains(text(),'{surname}')]")
    all_results = [r for r in results
                   if r.is_displayed() and r != inp and r.tag_name.lower() != 'input']

    if not all_results:
        inp.send_keys(Keys.ENTER)
        time.sleep(2)
        results = driver.find_elements(By.XPATH, f"//*[contains(text(),'{surname}')]")
        all_results = [r for r in results
                       if r.is_displayed() and r != inp and r.tag_name.lower() != 'input']

    target = None
    for r in all_results:
        try:
            if match_correspondent(r.text, person_name):
                target = r
                break
        except Exception:
            continue
    if not target and all_results:
        target = all_results[0]

    if target:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", target)
        time.sleep(0.3)
        ActionChains(driver).move_to_element(target).pause(0.3).click().perform()
        time.sleep(1)
        log.info(f"Адресат добавлен: {person_name}")
    else:
        log.warning(f"Адресат не найден: {person_name}")


# ================= REGISTRATION =================

def register_and_resolve(driver, index, total):
    """Регистрирует + На резолюцию + Да. Возвращает True при успехе."""
    log.info("Регистрирую...")
    registered = False
    try:
        btn = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,
                "#header-action-btn-register, [id*='header-action-btn-register']")))
        click(driver, btn, "Зарегистрировать")
        time.sleep(3)
        log.info(f"Документ {index}/{total} ЗАРЕГИСТРИРОВАН")
        registered = True
    except Exception:
        try:
            btn = driver.find_element(By.XPATH, "//div[contains(text(),'Зарегистрировать')]")
            click(driver, btn, "Зарегистрировать (fallback)")
            time.sleep(3)
            registered = True
        except Exception as e:
            log.error(f"'Зарегистрировать' не найдена: {e}")

    if not registered:
        return False

    # На резолюцию
    res_btn = None
    for _ in range(10):
        try:
            btn = driver.find_element(By.ID, "header-action-btn-send_on_resolution")
            if btn.is_displayed():
                res_btn = btn
                break
        except Exception:
            pass
        try:
            for b in driver.find_elements(By.XPATH, "//*[contains(text(),'На резолюцию')]"):
                if b.is_displayed():
                    res_btn = b
                    break
        except Exception:
            pass
        if res_btn:
            break
        time.sleep(1)

    if not res_btn:
        log.warning("'На резолюцию' не появилась")
        return True

    click(driver, res_btn, "На резолюцию")
    time.sleep(2)

    # Да
    yes_btn = None
    for _ in range(10):
        # 1) Точный id
        try:
            btn = driver.find_element(By.ID, "confirm_dialog_btn_yes")
            if btn.is_displayed():
                yes_btn = btn
                break
        except Exception:
            pass
        # 2) Substring id (GWT может добавлять префиксы/суффиксы)
        try:
            btn = driver.find_element(By.CSS_SELECTOR,
                "[id*='confirm_dialog_btn_yes'], [id*='confirm'][id*='yes']")
            if btn.is_displayed():
                yes_btn = btn
                break
        except Exception:
            pass
        # 3) По тексту "Да" в любом видимом элементе
        try:
            for b in driver.find_elements(By.XPATH, "//*[normalize-space(text())='Да']"):
                if b.is_displayed():
                    yes_btn = b
                    break
        except Exception:
            pass
        if yes_btn:
            break
        time.sleep(1)

    if yes_btn:
        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", yes_btn)
            time.sleep(0.3)
        except Exception:
            pass
        clicked = False
        # 1) ActionChains
        try:
            ActionChains(driver).move_to_element(yes_btn).pause(0.3).click().perform()
            log.info(f"Клик 'Да' (ActionChains): id={yes_btn.get_attribute('id')}")
            clicked = True
        except Exception as e:
            log.info(f"ActionChains 'Да' не сработал: {e}")
        # 2) JS
        if not clicked:
            try:
                driver.execute_script("arguments[0].click();", yes_btn)
                log.info("Клик 'Да' (JS)")
                clicked = True
            except Exception as e:
                log.info(f"JS 'Да' не сработал: {e}")
        # 3) Нативный
        if not clicked:
            try:
                yes_btn.click()
                log.info("Клик 'Да' (native)")
                clicked = True
            except Exception as e:
                log.info(f"Native 'Да' не сработал: {e}")
        # 4) Enter — GWT-диалоги часто принимают его
        if clicked:
            try:
                ActionChains(driver).send_keys(Keys.ENTER).perform()
            except Exception:
                pass
        time.sleep(3)
        log.info(f"Документ {index}/{total} НА РЕЗОЛЮЦИИ")
    else:
        log.warning("Диалог 'Да' не появился за 10 сек")
    return True


def close_card_and_wait_main(driver):
    """Закрывает карточку через header-close-btn и ждёт главную страницу."""
    time.sleep(2)
    try:
        close_btn = driver.find_element(By.ID, "header-close-btn")
        if close_btn.is_displayed():
            ActionChains(driver).move_to_element(close_btn).pause(0.3).click().perform()
            time.sleep(2)
            log.info("Карточка закрыта")
        else:
            log.info("Карточка уже закрыта")
    except Exception:
        log.info("Карточка уже закрыта")

    try:
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "mainscreen-create-button")))
    except Exception:
        log.warning("Главная не загрузилась — перезагружаю")
        driver.get(settings.get("asud_url", cfg.DEFAULTS["asud_url"]))
        wait_asud_loaded(driver)


# ================= DOCUMENT FLOW =================

def _link_to_corr_number(link):
    """Строит 'Номер у корреспондента' из Link: 'б/н <ISO-дата>' или 'б/н'."""
    if isinstance(link, (datetime, date)):
        return f"б/н {link.strftime('%d.%m.%Y %H-%M-%S')}"
    if link:
        s = str(link).strip()
        return f"б/н {s}" if s else "б/н"
    return "б/н"


def process_mix(driver, doc_data, index, total):
    """[MIX] Создаёт один входящий документ по схеме mix-сценария."""
    log.info(f"{'='*50}")
    log.info(f"ДОКУМЕНТ {index}/{total}: {doc_data['тема'][:60]}")
    log.info(f"Корреспондент: {doc_data['корреспондент']} "
             f"({'найден ' + (doc_data['корр_источник'] or '') if doc_data['корр_найден'] else 'ЗАГЛУШКА'})")
    log.info(f"Тип: {doc_data['тип_название']}")

    # [1] Кнопка создания
    el = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
        EC.presence_of_element_located((By.ID, "mainscreen-create-button")))
    time.sleep(1)
    click(driver, el, "Создать документ")
    time.sleep(3)

    # [2] Входящий документ
    wait_and_click(driver, By.XPATH,
        "//div[contains(text(),'Входящий документ')]", "Входящий документ")
    time.sleep(1)

    # [3] Вид (по DOC_TYPE_MAP)
    subtype = doc_data.get("тип_название", "Письма, заявления и жалобы граждан, акционеров")
    short = subtype[:30]
    wait_and_click(driver, By.XPATH,
        f"//div[contains(text(),'{short}')] | //td[contains(text(),'{short}')]", subtype)
    time.sleep(0.5)

    wait_and_click(driver, By.XPATH,
        "//button[contains(text(),'Создать документ')] | //div[contains(text(),'Создать документ')]",
        "Создать документ")
    time.sleep(5)

    # [4] Заполнение формы
    fill_text(driver, doc_data["содержание"])
    fill_correspondent_field(driver, doc_data["корреспондент"])
    fill_corr_number(driver, _link_to_corr_number(doc_data.get("link")))
    fill_corr_date(driver)

    for addr in settings.get("addressees", cfg.DEFAULTS["addressees"]):
        add_addressee(driver, addr)
        time.sleep(0.5)

    fill_delivery_method(driver)
    time.sleep(0.5)

    # [5] Сохранение
    try:
        save_btn = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            EC.element_to_be_clickable((By.ID, "header-save-btn")))
        click(driver, save_btn, "Сохранить")
        time.sleep(3)
        log.info(f"Документ {index}/{total} сохранён")
    except Exception as e:
        log.error(f"Ошибка сохранения: {e}")

    # [6] Прикрепление по Link
    outlook_dir = settings.get("outlook_dir", cfg.DEFAULTS["outlook_dir"])
    dummy_path = doc_data.get("файл")
    attach_path = find_msg_by_link(doc_data.get("link"), outlook_dir, dummy_path)
    if attach_path:
        log.info(f"Прикрепляю: {os.path.basename(attach_path)}")
        attach_content(driver, attach_path)
        wait_modal_closed(driver)
        # Реальный (не dummy) файл → переносим в Завершено/
        if attach_path != dummy_path:
            move_to_done(attach_path, outlook_dir)
    else:
        log.info("Нет файла — пропускаю")

    # [7] Регистрация (если ФИО найдено) или черновик
    if doc_data["корр_найден"]:
        register_and_resolve(driver, index, total)
    else:
        log.warning(f"Row {doc_data['row_idx']}: ФИО НЕ найдено — "
                    f"оставляю в ЧЕРНОВИКАХ для ручной доработки "
                    f"(тема: {doc_data['тема'][:60]})")

    close_card_and_wait_main(driver)


def process_auto_create(driver, doc_data, index, total):
    """[AUTO-CREATE] Создаёт один входящий документ:
    корреспондент из Excel, фикс. подтип, dummy .msg, безусловная регистрация."""
    log.info(f"{'='*50}")
    log.info(f"ДОКУМЕНТ {index}/{total}")
    log.info(f"Корреспондент: {doc_data['корреспондент']}")
    log.info(f"Содержание: {doc_data['содержание'][:60]}...")

    # [1] Кнопка создания
    el = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
        EC.presence_of_element_located((By.ID, "mainscreen-create-button")))
    time.sleep(1)
    click(driver, el, "Создать документ")
    time.sleep(3)

    # [2] Входящий документ
    wait_and_click(driver, By.XPATH,
        "//div[contains(text(),'Входящий документ')]", "Входящий документ")
    time.sleep(1)

    # [3] Вид (фиксированный из config)
    subtype = settings.get("doc_subtype", cfg.DEFAULTS["doc_subtype"])
    short = subtype[:30]
    wait_and_click(driver, By.XPATH,
        f"//div[contains(text(),'{short}')] | //td[contains(text(),'{short}')]", subtype)
    time.sleep(0.5)

    wait_and_click(driver, By.XPATH,
        "//button[contains(text(),'Создать документ')] | //div[contains(text(),'Создать документ')]",
        "Создать документ")
    time.sleep(5)

    # [4] Заполнение формы
    fill_text(driver, doc_data["содержание"])
    fill_correspondent_field(driver, doc_data["корреспондент"])
    fill_corr_number(driver, f"б/н ({index})")
    fill_corr_date(driver)

    for addr in settings.get("addressees", cfg.DEFAULTS["addressees"]):
        add_addressee(driver, addr)
        time.sleep(0.5)

    fill_delivery_method(driver)
    time.sleep(0.5)

    # [5] Сохранение
    try:
        save_btn = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            EC.element_to_be_clickable((By.ID, "header-save-btn")))
        click(driver, save_btn, "Сохранить")
        time.sleep(3)
        log.info(f"Документ {index}/{total} сохранён")
    except Exception as e:
        log.error(f"Ошибка сохранения: {e}")

    # [6] Прикрепление (dummy .msg)
    if doc_data.get("файл"):
        attach_content(driver, doc_data["файл"])
        wait_modal_closed(driver)
    else:
        log.info("Нет пустышки .msg — пропускаю вложение")

    # [7] Регистрация безусловная
    register_and_resolve(driver, index, total)
    close_card_and_wait_main(driver)


# ================= MAIN =================

settings = {}


def _choose_xlsx(base_dir):
    """Интерактивно выбирает .xlsx файл рядом с exe."""
    xlsx_files = [f for f in os.listdir(base_dir) if f.lower().endswith('.xlsx')]
    if not xlsx_files:
        log.error(f"Нет .xlsx в {base_dir}")
        input("Enter...")
        sys.exit(1)
    if len(xlsx_files) == 1:
        log.info(f"Файл: {xlsx_files[0]}")
        return os.path.join(base_dir, xlsx_files[0])
    print(f"\nНайдено {len(xlsx_files)} xlsx-файлов:")
    for i, f in enumerate(xlsx_files, 1):
        print(f"  {i}. {f}")
    choice = input("Выбери номер: ").strip()
    try:
        return os.path.join(base_dir, xlsx_files[int(choice) - 1])
    except (ValueError, IndexError):
        log.error("Неверный выбор")
        sys.exit(1)


def _detect_scenario(xlsx_path):
    """Определяет сценарий по наличию колонки 'Link' в xlsx.

    Правило простое: если хотя бы на одном листе в заголовках
    есть колонка 'Link' → сценарий MIX; иначе → AUTO-CREATE.
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        log.warning(f"Не удалось открыть xlsx для авто-детекта: {e}")
        return None

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            except StopIteration:
                continue
            headers = [str(c).strip().lower() for c in row if c is not None]
            if "link" in headers:
                log.info(f"Авто-детект: MIX (колонка 'Link' на листе '{sheet_name}')")
                return "mix"

        log.info("Авто-детект: AUTO-CREATE (колонка 'Link' не найдена ни на одном листе)")
        return "auto-create"
    finally:
        wb.close()


def _prompt_attachment_dir(default_dir, description):
    """Промпт пути к папке с вложениями. Enter = default."""
    print(f"\n{description}")
    print(f"Нажми Enter, чтобы использовать: {default_dir}")
    user_dir = input("Путь: ").strip().strip('"').strip("'")
    chosen = user_dir if user_dir else default_dir
    if chosen and not os.path.isdir(chosen):
        log.warning(f"Папка '{chosen}' не существует")
    else:
        log.info(f"Папка: {chosen}")
    return chosen


def _start_browser(base_dir):
    """Запускает Edge с msedgedriver.exe из base_dir."""
    driver_path = os.path.join(base_dir, "msedgedriver.exe")
    if not os.path.exists(driver_path):
        log.error(f"msedgedriver.exe не найден в {base_dir}")
        input("Enter...")
        sys.exit(1)

    options = EdgeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--auth-server-whitelist=*.interrao.ru")
    options.add_argument("--auth-negotiate-delegate-whitelist=*.interrao.ru")
    options.add_argument("--log-level=3")
    options.add_experimental_option('excludeSwitches', ['enable-logging'])

    service = EdgeService(executable_path=driver_path)
    return webdriver.Edge(service=service, options=options)


def _print_summary(done_count, total, err_count, elapsed_seconds, extra_lines=None):
    """Выводит итоговую плашку с временем."""
    elapsed = timedelta(seconds=int(elapsed_seconds))
    avg = timedelta(seconds=int(elapsed_seconds / done_count)) if done_count else None
    lines = [
        "",
        "=" * 60,
        "ГОТОВО!",
        f"  Обработано: {done_count} / {total}",
        f"  Ошибок:     {err_count}",
    ]
    if extra_lines:
        lines.extend(extra_lines)
    lines.append(f"  Затрачено:  {elapsed}" +
                 (f"  (в среднем {avg}/док)" if avg else ""))
    lines.append("=" * 60)
    for line in lines:
        log.info(line)
        print(line)


# ---------- Сценарий 1: auto-create ----------

def run_auto_create(excel_path=None):
    """Сценарий: B=Содержание, C=Корреспондент, dummy .msg, фикс. подтип,
    безусловная регистрация + На резолюцию."""
    global start_time
    start_time = time.monotonic()

    log.info("=" * 50)
    log.info("АСУД ИК — сценарий AUTO-CREATE")
    log.info("=" * 50)

    base_dir = cfg.get_base_dir()
    if excel_path is None:
        excel_path = _choose_xlsx(base_dir)

    # Папка для dummy .msg (рядом с exe по умолчанию)
    msg_dir = _prompt_attachment_dir(
        base_dir, "Папка с .msg-пустышкой (поиск рекурсивно по подпапкам).")
    msg_path = get_dummy_msg(msg_dir)
    if msg_path:
        log.info(f"Пустышка: {os.path.basename(msg_path)}")
    else:
        log.warning(f"В {msg_dir} не найдено .msg — документы без вложений")

    docs = load_excel_auto_create(excel_path)
    for doc in docs:
        doc["файл"] = msg_path

    if not docs:
        log.error("Нет данных!")
        input("Enter...")
        sys.exit(1)

    print(f"\nПервые 5:")
    for i, d in enumerate(docs[:5], 1):
        print(f"  {i}. {d['корреспондент'][:30]} | {d['содержание'][:50]}...")
    print(f"\nВсего: {len(docs)}")

    if input("Начать? (да/нет): ").strip().lower() not in ("да", "д", "y", "yes", ""):
        print("Отменено.")
        sys.exit(0)

    driver = _start_browser(base_dir)
    try:
        url = settings.get("asud_url", cfg.DEFAULTS["asud_url"])
        log.info(f"Открываю {url}")
        driver.get(url)
        wait_asud_loaded(driver)

        done_count, err_count = 0, 0
        for i, doc in enumerate(docs, 1):
            try:
                process_auto_create(driver, doc, i, len(docs))
                done_count += 1
            except Exception as e:
                log.error(f"ОШИБКА документ {i}: {e}")
                err_count += 1
                driver.get(url)
                wait_asud_loaded(driver)
                continue

        _print_summary(done_count, len(docs), err_count,
                       time.monotonic() - start_time)
        input("\nEnter для закрытия...")

    except Exception as e:
        log.error(f"Ошибка: {e}")
        input("Enter...")
    finally:
        driver.quit()
        log.info("Браузер закрыт")


# ---------- Сценарий 2: mix ----------

def run_mix(excel_path=None):
    """Сценарий: Лист2, ФИО из TextBody, .msg по Link, регистрация
    только если ФИО найдено, resume-state."""
    global start_time
    start_time = time.monotonic()

    log.info("=" * 50)
    log.info("АСУД ИК — сценарий MIX")
    log.info("=" * 50)

    base_dir = cfg.get_base_dir()
    if excel_path is None:
        excel_path = _choose_xlsx(base_dir)

    # Папка outlook_dir (для поиска .msg по Link)
    default_outlook = settings.get("outlook_dir", cfg.DEFAULTS["outlook_dir"])
    outlook_dir = _prompt_attachment_dir(
        default_outlook, "Папка с .msg-файлами (поиск рекурсивно по подпапкам).")
    settings["outlook_dir"] = outlook_dir

    msg_path = get_dummy_msg(base_dir)
    if msg_path:
        log.info(f"Пустышка (fallback): {os.path.basename(msg_path)}")

    docs = load_excel_mix(excel_path)
    for doc in docs:
        doc["файл"] = msg_path

    if not docs:
        log.error("Нет данных!")
        input("Enter...")
        sys.exit(1)

    # Resume: проверяем state-файл
    processed = load_state(excel_path)
    if processed:
        done_in_current = [d for d in docs
                           if _link_key(d.get("link")) in processed]
        if done_in_current:
            print(f"\nВ state-файле {len(done_in_current)} ранее обработанных документов.")
            print("  Enter / 'да'  — ПРОПУСТИТЬ их, продолжить с остальных")
            print("  'нет'         — обработать ВСЁ заново (дубли!)")
            print("  'сброс'       — обнулить state и обработать всё")
            ans = input("Что делаем? [да]: ").strip().lower()
            if ans in ("сброс", "reset"):
                save_state(excel_path, set())
                processed = set()
                log.info("State обнулён — обрабатываю всё заново")
            elif ans in ("нет", "н", "n", "no"):
                log.info("Обрабатываю всё заново (state будет дополнен)")
            else:
                before = len(docs)
                docs = [d for d in docs
                        if _link_key(d.get("link")) not in processed]
                log.info(f"Пропускаю {before - len(docs)} обработанных, "
                         f"осталось {len(docs)}")
                if not docs:
                    log.info("Все строки уже обработаны.")
                    input("Enter...")
                    sys.exit(0)

    known = sum(1 for d in docs if d["корр_найден"])
    unknown = len(docs) - known
    print(f"\nПервые 5:")
    for i, d in enumerate(docs[:5], 1):
        flag = 'OK' if d["корр_найден"] else '!!'
        print(f"  {i}. [{d['тип_индекс']}] {flag} {d['корреспондент'][:30]} | {d['тема'][:50]}")
    print(f"\nВсего: {len(docs)}  (ФИО: {known}, заглушка: {unknown})")

    if input("Начать? (да/нет): ").strip().lower() not in ("да", "д", "y", "yes", ""):
        print("Отменено.")
        sys.exit(0)

    driver = _start_browser(base_dir)
    try:
        url = settings.get("asud_url", cfg.DEFAULTS["asud_url"])
        log.info(f"Открываю {url}")
        driver.get(url)
        wait_asud_loaded(driver)

        done_count, err_count = 0, 0
        for i, doc in enumerate(docs, 1):
            try:
                process_mix(driver, doc, i, len(docs))
                key = _link_key(doc.get("link"))
                if key:
                    processed.add(key)
                    save_state(excel_path, processed)
                done_count += 1
            except Exception as e:
                log.error(f"ОШИБКА документ {i}: {e}")
                err_count += 1
                driver.get(url)
                wait_asud_loaded(driver)
                continue

        _print_summary(done_count, len(docs), err_count,
                       time.monotonic() - start_time,
                       extra_lines=[f"  В черновиках: {unknown}"])
        if unknown:
            log.warning(f"Проверьте {unknown} документов в черновиках — "
                        f"ФИО не извлечено автоматически")
        if err_count:
            log.warning(f"{err_count} документов упали с ошибкой. "
                        f"Перезапуск продолжит с них (state запомнил обработанные).")
        input("\nEnter для закрытия...")

    except Exception as e:
        log.error(f"Ошибка: {e}")
        input("Enter...")
    finally:
        driver.quit()
        log.info("Браузер закрыт")


# ---------- Диспетчер ----------

_SCENARIO_DESCR = {
    "mix":         "Лист2 (Link/Subject/TextBody/Тип) → ФИО из TextBody, "
                   ".msg по Link, черновики при ненайденном ФИО",
    "auto-create": "B=Содержание, C=Корреспондент → корреспондент из Excel, "
                   "dummy .msg, безусловная регистрация",
}


def _ask_scenario_manual(reason=""):
    """Ручной выбор сценария когда авто-детект не сработал или переопределяется."""
    if reason:
        print(f"\n{reason}")
    print("\nВыбери сценарий вручную:")
    print(f"  1. auto-create — {_SCENARIO_DESCR['auto-create']}")
    print(f"  2. mix         — {_SCENARIO_DESCR['mix']}")
    choice = input("Номер: ").strip()
    if choice == "1":
        return "auto-create"
    if choice == "2":
        return "mix"
    log.error(f"Неверный выбор: {choice!r}")
    input("Enter...")
    sys.exit(1)


def main():
    global settings
    settings = cfg.load()

    print("=" * 60)
    print("АСУД ИК — автоматизация Входящих документов")
    print("=" * 60)

    base_dir = cfg.get_base_dir()
    excel_path = _choose_xlsx(base_dir)

    # Авто-детект сценария по структуре xlsx
    scenario = _detect_scenario(excel_path)

    if scenario:
        print(f"\nАвто-детект: {scenario} — {_SCENARIO_DESCR[scenario]}")
        ans = input("Принять? [Enter=да, 'нет' = выбрать вручную]: ").strip().lower()
        if ans in ("нет", "н", "n", "no"):
            scenario = _ask_scenario_manual()
    else:
        scenario = _ask_scenario_manual(
            "Не удалось определить сценарий по структуре файла.")

    if scenario == "auto-create":
        run_auto_create(excel_path)
    elif scenario == "mix":
        run_mix(excel_path)
    else:
        log.error(f"Неизвестный сценарий: {scenario}")
        sys.exit(1)


if __name__ == "__main__":
    main()
