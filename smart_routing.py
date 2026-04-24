"""
smart_routing.py — Пакетное создание Входящих документов в АСУД ИК.

Читает Excel, определяет тип документа, создаёт в АСУД, прикрепляет .msg.

Модули:
  config.py       — настройки (+ config.json)
  ui.py           — Selenium UI-хелперы
  correspondent.py — создание корреспондентов
  attachments.py  — поиск и прикрепление файлов
"""

import os
import re
import sys
import time
import logging
from datetime import date, datetime, timedelta

import openpyxl
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import config as cfg
from ui import (click, wait_and_click, find_input_near_label,
                wait_asud_loaded, wait_modal_closed, close_open_modals, js_set_value)
from correspondent import (fill_correspondent_field, create_correspondent,
                           match_strict, fio_to_initials)
from attachments import find_msg_by_link, get_dummy_msg, attach_content, move_to_done


# ================= LOGGING =================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
    handlers=[
        logging.StreamHandler(),
    ]
)
log = logging.getLogger("asud")
start_time = time.monotonic()


# ================= EXCEL =================

def _clean_body(text):
    """Очищает TextBody от служебных строк."""
    if not text:
        return ""
    t = str(text).replace('_x000D_', '\n')
    lines = t.split('\n')
    cleaned = []
    for line in lines:
        s = line.strip()
        if re.search(r'внимание!?\s*письмо\s+было\s+отправлено\s+внешним', s, re.IGNORECASE):
            continue
        if re.match(r'^-{3,}\s*Original\s*Message\s*-{3,}$', s, re.IGNORECASE):
            continue
        cleaned.append(line)
    t = '\n'.join(cleaned)
    t = re.sub(r'\n\s*\n\s*\n+', '\n\n', t)
    return t.strip()


def load_excel(file_path):
    """Читает Excel. Колонки: A=Link, B=Subject, C=TextBody, D=Тип."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row or len(row) < 4:
            skipped += 1
            continue
        subject = row[1]
        body = row[2]
        type_idx = row[3]

        if not subject:
            skipped += 1
            continue
        try:
            type_idx = int(type_idx) if type_idx is not None else 0
        except (ValueError, TypeError):
            type_idx = 0
        if type_idx == 0 or type_idx not in cfg.DOC_TYPE_MAP:
            skipped += 1
            continue

        clean_subject = re.sub(r'^(FW:|RE:|Fwd:)\s*', '', str(subject).strip(), flags=re.IGNORECASE)
        body_clean = _clean_body(body) if body else clean_subject

        rows.append({
            "содержание": body_clean,
            "корреспондент": settings.get("correspondent", "Неизвестный Неизвестный Неизвестный"),
            "тема": clean_subject,
            "тип_индекс": type_idx,
            "тип_название": cfg.DOC_TYPE_MAP[type_idx],
            "link": row[0],
        })
    wb.close()
    log.info(f"Загружено: {len(rows)}, пропущено: {skipped}")
    return rows


# ================= FORM FILLING =================

def fill_text(driver, text):
    """Заполняет краткое содержание (textarea). JS-ввод (атомарный),
    send_keys как fallback."""
    try:
        areas = driver.find_elements(By.TAG_NAME, "textarea")
        visible = [a for a in areas if a.is_displayed()]
        if not visible:
            log.warning("Textarea не найдена")
            time.sleep(0.5)
            return
        ta = visible[0]
        # JS-ввод: атомарно, без потерь символов на длинных TextBody
        js_ok = False
        try:
            driver.execute_script("""
                var el = arguments[0], value = arguments[1];
                el.focus();
                el.value = value;
                el.dispatchEvent(new Event('input', {bubbles:true}));
                el.dispatchEvent(new Event('change', {bubbles:true}));
            """, ta, text)
            # Верификация: значение реально установилось
            actual = (ta.get_attribute('value') or '')
            if actual.strip() == text.strip():
                log.info(f"Краткое содержание: JS-ввод ({len(text)} символов)")
                js_ok = True
            else:
                log.warning(f"JS-ввод не закрепился "
                            f"(ожидал {len(text)} симв., получил {len(actual)}), "
                            f"падаю в send_keys")
        except Exception as e:
            log.warning(f"JS-ввод textarea упал: {e}, падаю в send_keys")

        if not js_ok:
            # Fallback: замещаем табы пробелами, чтобы они не уводили
            # фокус на следующее поле формы АСУД
            safe_text = text.replace('\t', '    ')
            ta.click()
            time.sleep(0.3)
            ta.clear()
            ta.send_keys(safe_text)
            log.info("Краткое содержание: send_keys fallback (табы заменены)")
    except Exception as e:
        log.error(f"Ошибка заполнения содержания: {e}")
    time.sleep(0.5)


def fill_corr_number(driver, link=None):
    """Заполняет 'Номер у корреспондента' = 'б/н <link>'."""
    if isinstance(link, (datetime, date)):
        link_str = link.strftime("%d.%m.%Y %H-%M-%S")
    elif link:
        link_str = str(link).strip()
    else:
        link_str = ""
    value = f"б/н {link_str}" if link_str else "б/н"

    inp = find_input_near_label(driver, "Номер у корреспондента")
    if not inp:
        log.warning("Поле 'Номер у корреспондента' не найдено")
        return
    try:
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", inp)
        time.sleep(0.3)
        inp.click()
        time.sleep(0.3)
        inp.clear()
        time.sleep(0.2)
        inp.send_keys(value)
        inp.send_keys(Keys.TAB)
        log.info(f"Номер: {value}")
    except Exception:
        js_set_value(driver, inp, value)
        log.info(f"Номер (JS): {value}")


def fill_corr_date(driver):
    """Заполняет 'Дата у корреспондента' = сегодня."""
    today = date.today().strftime("%d.%m.%Y")
    labels = driver.find_elements(By.XPATH,
        "//*[normalize-space(text())='Дата у корреспондента']")
    inp = None
    for label in labels:
        try:
            if not label.is_displayed():
                continue
            for level in range(1, 6):
                parent = label
                for _ in range(level):
                    parent = parent.find_element(By.XPATH, "..")
                inputs = parent.find_elements(By.CSS_SELECTOR, "input[type='text']")
                visible = [i for i in inputs
                           if i.is_displayed() and i.get_attribute("readonly") is None]
                if visible:
                    inp = visible[0]
                    break
            if inp:
                break
        except Exception:
            continue

    if not inp:
        log.warning("Поле 'Дата у корреспондента' не найдено")
        return
    try:
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", inp)
        driver.execute_script("arguments[0].focus(); arguments[0].click();", inp)
        time.sleep(0.3)
        inp.send_keys(Keys.CONTROL + "a")
        time.sleep(0.2)
        inp.send_keys(Keys.DELETE)
        time.sleep(0.2)
        inp.send_keys(today)
        inp.send_keys(Keys.TAB)
        log.info(f"Дата: {today}")
    except Exception:
        js_set_value(driver, inp, today)
        log.info(f"Дата (JS): {today}")


def fill_delivery_method(driver):
    """Выбирает 'Электронная почта' в 'Способ получения'."""
    target_text = settings.get("delivery_method", "Электронная почта")
    labels = driver.find_elements(By.XPATH,
        "//*[normalize-space(text())='Способ получения']")
    trigger = None
    for label in labels:
        try:
            if not label.is_displayed():
                continue
            for level in range(1, 8):
                parent = label
                for _ in range(level):
                    parent = parent.find_element(By.XPATH, "..")
                inputs = parent.find_elements(By.CSS_SELECTOR, "input[type='text']")
                for i in inputs:
                    if i.is_displayed():
                        trigger = i
                        break
                if trigger:
                    break
                for sel in ["div[class*='trigger']", "img[class*='trigger']"]:
                    try:
                        el = parent.find_element(By.CSS_SELECTOR, sel)
                        if el.is_displayed():
                            trigger = el
                            break
                    except Exception:
                        continue
                if trigger:
                    break
            if trigger:
                break
        except Exception:
            continue

    if not trigger:
        log.warning("Поле 'Способ получения' не найдено")
        return

    click(driver, trigger, "Способ получения")
    time.sleep(1.5)

    option = None
    for attempt in range(3):
        candidates = driver.find_elements(By.XPATH,
            f"//*[contains(text(),'{target_text}')]")
        for c in candidates:
            try:
                if c.is_displayed() and c.tag_name.lower() != 'input':
                    option = c
                    break
            except Exception:
                continue
        if option:
            break
        time.sleep(1)

    if option:
        click(driver, option, target_text)
        log.info(f"Способ получения: {target_text}")
    else:
        log.warning(f"'{target_text}' не найдена в списке")


def add_addressee(driver, person_name):
    """Добавляет адресата через combobox."""
    inp = find_input_near_label(driver, "Адресаты")
    if not inp:
        log.warning("Поле адресата не найдено")
        return

    surname = person_name.split()[0]
    inp.click()
    time.sleep(0.5)
    inp.clear()
    time.sleep(0.3)
    for char in surname:
        inp.send_keys(char)
        time.sleep(0.1)
    time.sleep(2)

    from correspondent import match_correspondent
    results = driver.find_elements(By.XPATH, f"//*[contains(text(),'{surname}')]")
    all_results = [r for r in results
                   if r.is_displayed() and r != inp and r.tag_name.lower() != 'input']

    if not all_results:
        inp.send_keys(Keys.ENTER)
        time.sleep(2)
        results = driver.find_elements(By.XPATH, f"//*[contains(text(),'{surname}')]")
        all_results = [r for r in results
                       if r.is_displayed() and r != inp and r.tag_name.lower() != 'input']

    target = None
    for r in all_results:
        try:
            if match_correspondent(r.text, person_name):
                target = r
                break
        except Exception:
            continue
    if not target and all_results:
        target = all_results[0]

    if target:
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", target)
        time.sleep(0.3)
        ActionChains(driver).move_to_element(target).pause(0.3).click().perform()
        time.sleep(1)
        log.info(f"Адресат добавлен: {person_name}")
    else:
        log.warning(f"Адресат не найден: {person_name}")


# ================= DOCUMENT FLOW =================

def create_one_document(driver, doc_data, index, total):
    """Создаёт один входящий документ."""
    log.info(f"{'='*50}")
    log.info(f"ДОКУМЕНТ {index}/{total}: {doc_data['тема'][:60]}")
    log.info(f"Корреспондент: {doc_data['корреспондент']}")
    log.info(f"Тип: {doc_data['тип_название']}")

    # [1/7] Кнопка создания
    el = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
        EC.presence_of_element_located((By.ID, "mainscreen-create-button")))
    time.sleep(1)
    click(driver, el, "Создать документ")
    time.sleep(3)

    # [2/7] Входящий документ
    wait_and_click(driver, By.XPATH,
        "//div[contains(text(),'Входящий документ')]", "Входящий документ")
    time.sleep(1)

    # [3/7] Вид
    subtype = doc_data.get("тип_название", "Письма, заявления и жалобы граждан, акционеров")
    short = subtype[:30]
    wait_and_click(driver, By.XPATH,
        f"//div[contains(text(),'{short}')] | //td[contains(text(),'{short}')]", subtype)
    time.sleep(0.5)

    # Создать документ
    wait_and_click(driver, By.XPATH,
        "//button[contains(text(),'Создать документ')] | //div[contains(text(),'Создать документ')]",
        "Создать документ")
    time.sleep(5)

    # [4/7] Заполнение формы
    fill_text(driver, doc_data["содержание"])
    fill_correspondent_field(driver, doc_data["корреспондент"])
    fill_corr_number(driver, doc_data.get("link"))
    fill_corr_date(driver)
    add_addressee(driver, settings.get("addressee", "Басманов Александр Владимирович"))
    time.sleep(0.5)
    fill_delivery_method(driver)
    time.sleep(0.5)

    # [5/7] Сохранение
    try:
        save_btn = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            EC.element_to_be_clickable((By.ID, "header-save-btn")))
        click(driver, save_btn, "Сохранить")
        time.sleep(3)
        log.info(f"Документ {index}/{total} сохранён")
    except Exception as e:
        log.error(f"Ошибка сохранения: {e}")

    # [6/7] Прикрепление
    outlook_dir = settings.get("outlook_dir", cfg.DEFAULTS["outlook_dir"])
    dummy_path = doc_data.get("файл")
    attach_path = find_msg_by_link(doc_data.get("link"), outlook_dir, dummy_path)
    if attach_path:
        log.info(f"Прикрепляю: {os.path.basename(attach_path)}")
        attach_content(driver, attach_path)
        wait_modal_closed(driver)
        # Реальный (не dummy) файл → переносим в Завершено/
        # В smart-routing все документы в черновиках по дизайну,
        # поэтому критерий переноса — успешный attach (а не регистрация).
        if attach_path != dummy_path:
            move_to_done(attach_path, outlook_dir)
    else:
        log.info("Нет файла — пропускаю")

    # [7/7] Закрытие карточки
    log.info(f"Документ {index}/{total} в черновиках")
    time.sleep(2)
    try:
        close_btn = driver.find_element(By.ID, "header-close-btn")
        if close_btn.is_displayed():
            ActionChains(driver).move_to_element(close_btn).pause(0.3).click().perform()
            time.sleep(2)
            log.info("Карточка закрыта")
        else:
            log.info("Карточка уже закрыта")
    except Exception:
        log.info("Карточка уже закрыта")

    try:
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "mainscreen-create-button")))
    except Exception:
        log.warning("Главная не загрузилась — перезагружаю")
        driver.get(settings.get("asud_url", cfg.DEFAULTS["asud_url"]))
        wait_asud_loaded(driver)


# ================= MAIN =================

settings = {}


def main():
    global settings
    settings = cfg.load()

    log.info("=" * 50)
    log.info("АСУД ИК — Пакетное создание Входящих документов")
    log.info("=" * 50)

    base_dir = cfg.get_base_dir()

    # Excel
    xlsx_files = [f for f in os.listdir(base_dir) if f.lower().endswith('.xlsx')]
    if not xlsx_files:
        log.error(f"Нет .xlsx в {base_dir}")
        input("Enter...")
        sys.exit(1)
    elif len(xlsx_files) == 1:
        excel_path = os.path.join(base_dir, xlsx_files[0])
        log.info(f"Файл: {xlsx_files[0]}")
    else:
        print(f"\nНайдено {len(xlsx_files)} xlsx-файлов:")
        for i, f in enumerate(xlsx_files, 1):
            print(f"  {i}. {f}")
        choice = input("Выбери номер: ").strip()
        try:
            excel_path = os.path.join(base_dir, xlsx_files[int(choice) - 1])
        except (ValueError, IndexError):
            log.error("Неверный выбор")
            sys.exit(1)

    # Папка с .msg — интерактивный ввод (Enter = дефолт из config)
    default_outlook = settings.get("outlook_dir", cfg.DEFAULTS["outlook_dir"])
    print(f"\nПапка с .msg-файлами (поиск рекурсивно по подпапкам).")
    print(f"Нажми Enter, чтобы использовать: {default_outlook}")
    user_dir = input("Путь: ").strip().strip('"').strip("'")
    if user_dir:
        settings["outlook_dir"] = user_dir
    outlook_dir = settings["outlook_dir"]
    if not os.path.isdir(outlook_dir):
        log.warning(f"Папка '{outlook_dir}' не существует — "
                    f"все вложения уйдут как пустышки")
    else:
        log.info(f"Папка вложений: {outlook_dir}")

    # Пустышка
    msg_path = get_dummy_msg(base_dir)
    if msg_path:
        log.info(f"Пустышка: {os.path.basename(msg_path)}")

    # Данные
    docs = load_excel(excel_path)
    for doc in docs:
        doc["файл"] = msg_path

    if not docs:
        log.error("Нет данных!")
        input("Enter...")
        sys.exit(1)

    print(f"\nПервые 5:")
    for i, d in enumerate(docs[:5], 1):
        print(f"  {i}. [{d['тип_индекс']}] {d['тема'][:60]}")
    print(f"\nВсего: {len(docs)}")

    confirm = input("Начать? (да/нет): ").strip().lower()
    if confirm not in ("да", "д", "y", "yes", ""):
        print("Отменено.")
        sys.exit(0)

    # Браузер
    from config import get_base_dir
    driver_path = os.path.join(get_base_dir(), "msedgedriver.exe")
    if not os.path.exists(driver_path):
        log.error(f"msedgedriver.exe не найден в {get_base_dir()}")
        input("Enter...")
        sys.exit(1)

    options = EdgeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--auth-server-whitelist=*.interrao.ru")
    options.add_argument("--auth-negotiate-delegate-whitelist=*.interrao.ru")
    options.add_argument("--log-level=3")
    options.add_experimental_option('excludeSwitches', ['enable-logging'])

    service = EdgeService(executable_path=driver_path)
    driver = webdriver.Edge(service=service, options=options)

    try:
        url = settings.get("asud_url", cfg.DEFAULTS["asud_url"])
        log.info(f"Открываю {url}")
        driver.get(url)
        wait_asud_loaded(driver)

        done_count, err_count = 0, 0
        for i, doc in enumerate(docs, 1):
            try:
                create_one_document(driver, doc, i, len(docs))
                done_count += 1
            except Exception as e:
                log.error(f"ОШИБКА документ {i}: {e}")
                err_count += 1
                driver.get(url)
                wait_asud_loaded(driver)
                continue

        elapsed_seconds = time.monotonic() - start_time
        elapsed = timedelta(seconds=int(elapsed_seconds))
        avg = timedelta(seconds=int(elapsed_seconds / done_count)) if done_count else None

        summary = [
            "",
            "=" * 60,
            f"ГОТОВО!",
            f"  Обработано: {done_count} / {len(docs)}",
            f"  Ошибок:     {err_count}",
            f"  Затрачено:  {elapsed}" + (f"  (в среднем {avg}/док)" if avg else ""),
            "=" * 60,
        ]
        for line in summary:
            log.info(line)
            print(line)

        input("\nEnter для закрытия...")

    except Exception as e:
        log.error(f"Ошибка: {e}")
        input("Enter...")
    finally:
        driver.quit()
        log.info("Браузер закрыт")


if __name__ == "__main__":
    main()
