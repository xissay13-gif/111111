"""
Экспорт писем из Outlook в Excel + .msg файлы.

Подключается к запущенному Outlook Desktop через COM,
скачивает письма из выбранной папки, сохраняет:
  - .msg файлы в D:\OutlookSubjects (имя = дата-время)
  - Почта.xlsx с реестром (Link, Subject, TextBody, ...)

Требования:
  - Windows + Outlook Desktop (не Web)
  - pip install pywin32 openpyxl pyinstaller

Сборка:
  pyinstaller --onefile --name outlook_export outlook_export.py
"""

import os
import sys
import re
from datetime import datetime, timedelta

try:
    import win32com.client
    import pywintypes
except ImportError:
    print("!! Не установлен pywin32. Выполните: pip install pywin32")
    input("Enter...")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("!! Не установлен openpyxl. Выполните: pip install openpyxl")
    input("Enter...")
    sys.exit(1)


OUTPUT_DIR = r"D:\OutlookSubjects"
OL_MSG_FORMAT = 3  # olMSG


def get_output_dir():
    """Создаёт папку вывода если не существует."""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        print(f"  Создана папка: {OUTPUT_DIR}")
    return OUTPUT_DIR


def connect_outlook():
    """Подключается к Outlook через COM."""
    print("Подключение к Outlook...")
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        namespace = outlook.GetNamespace("MAPI")
        print("  ОК Подключено к Outlook")
        return namespace
    except Exception as e:
        print(f"!! Не удалось подключиться к Outlook: {e}")
        print("   Убедитесь что Outlook Desktop запущен.")
        input("Enter...")
        sys.exit(1)


def list_folders(namespace, parent=None, level=0, folders_list=None):
    """Рекурсивно собирает список папок Outlook."""
    if folders_list is None:
        folders_list = []
    if parent is None:
        # Перебираем все аккаунты
        for store in namespace.Stores:
            try:
                root = store.GetRootFolder()
                folders_list.append((level, root.FolderPath, root))
                list_folders(namespace, root, level + 1, folders_list)
            except Exception:
                continue
    else:
        try:
            for folder in parent.Folders:
                folders_list.append((level, folder.FolderPath, folder))
                list_folders(namespace, folder, level + 1, folders_list)
        except Exception:
            pass
    return folders_list


def select_folder(namespace):
    """Позволяет пользователю выбрать папку из списка."""
    print("\nСканирую папки Outlook...")
    all_folders = list_folders(namespace)

    # Фильтруем: показываем только папки с письмами (не корневые)
    mail_folders = [(lvl, path, f) for lvl, path, f in all_folders if lvl >= 1]

    if not mail_folders:
        print("!! Папки не найдены")
        input("Enter...")
        sys.exit(1)

    print(f"\nНайдено {len(mail_folders)} папок:\n")
    for i, (lvl, path, _) in enumerate(mail_folders, 1):
        indent = "  " * (lvl - 1)
        # Показываем только имя папки + количество элементов
        name = path.split("\\")[-1] if "\\" in path else path
        try:
            count = _.Items.Count
            print(f"  {i:3}. {indent}{name} ({count})")
        except Exception:
            print(f"  {i:3}. {indent}{name}")

    print()
    while True:
        choice = input("Выберите номер папки: ").strip()
        try:
            idx = int(choice) - 1
            if 0 <= idx < len(mail_folders):
                selected = mail_folders[idx]
                print(f"\n  Выбрана: {selected[1]}")
                return selected[2]  # folder object
        except (ValueError, IndexError):
            pass
        print("  Неверный номер, попробуйте ещё раз")


def get_filter(folder):
    """Спрашивает у пользователя какие письма брать."""
    try:
        total = folder.Items.Count
    except Exception:
        total = "?"

    print(f"\nВ папке {total} писем.")
    print("Какие письма экспортировать?")
    print("  1. Все")
    print("  2. Только непрочитанные")
    print("  3. За сегодня")
    print("  4. За последние N дней")
    print()

    choice = input("Выбор (1-4): ").strip()

    if choice == "2":
        return "unread"
    elif choice == "3":
        return "today"
    elif choice == "4":
        days = input("За сколько дней? ").strip()
        try:
            return ("days", int(days))
        except ValueError:
            print("  Неверное число, беру за 7 дней")
            return ("days", 7)
    else:
        return "all"


def format_datetime_for_filename(dt):
    """Форматирует datetime в имя файла: DD.MM.YYYY HH-MM-SS."""
    if isinstance(dt, datetime):
        return dt.strftime("%d.%m.%Y %H-%M-%S")
    try:
        # pywintypes.datetime → обычный datetime
        py_dt = datetime(dt.year, dt.month, dt.day, dt.hour, dt.minute, dt.second)
        return py_dt.strftime("%d.%m.%Y %H-%M-%S")
    except Exception:
        return datetime.now().strftime("%d.%m.%Y %H-%M-%S")


def export_messages(folder, filter_type):
    """Экспортирует письма из папки: сохраняет .msg + собирает данные."""
    output_dir = get_output_dir()
    items = folder.Items
    items.Sort("[ReceivedTime]", True)  # Сортировка по дате (новые первые)

    # Применяем фильтр
    if filter_type == "unread":
        restriction = "[Unread] = true"
        items = items.Restrict(restriction)
        print(f"\nФильтр: только непрочитанные")
    elif filter_type == "today":
        today = datetime.now().strftime("%m/%d/%Y")
        restriction = f"[ReceivedTime] >= '{today}'"
        items = items.Restrict(restriction)
        print(f"\nФильтр: за сегодня ({today})")
    elif isinstance(filter_type, tuple) and filter_type[0] == "days":
        days = filter_type[1]
        since = (datetime.now() - timedelta(days=days)).strftime("%m/%d/%Y")
        restriction = f"[ReceivedTime] >= '{since}'"
        items = items.Restrict(restriction)
        print(f"\nФильтр: за последние {days} дней (с {since})")
    else:
        print(f"\nФильтр: все письма")

    try:
        count = items.Count
    except Exception:
        count = "?"
    print(f"Писем для экспорта: {count}")

    if count == 0 or count == "?":
        print("Нет писем для экспорта.")
        return []

    confirm = input(f"\nНачать экспорт {count} писем? (да/нет): ").strip().lower()
    if confirm not in ("да", "д", "y", "yes", ""):
        print("Отменено.")
        return []

    results = []
    exported = 0
    errors = 0

    print()
    for i, msg in enumerate(items, 1):
        try:
            # Получаем данные
            subject = str(msg.Subject or "").strip()
            body = str(msg.Body or "").strip()
            received = msg.ReceivedTime
            has_attach = bool(msg.Attachments.Count > 0)
            is_read = not msg.UnRead

            # Формируем имя файла
            link = format_datetime_for_filename(received)
            filename = f"{link}.msg"
            filepath = os.path.join(output_dir, filename)

            # Сохраняем .msg (если ещё не существует)
            if not os.path.exists(filepath):
                try:
                    msg.SaveAs(filepath, OL_MSG_FORMAT)
                except Exception as e:
                    print(f"  ! Не удалось сохранить {filename}: {e}")
                    errors += 1

            # Собираем в реестр
            results.append({
                "link": link,
                "subject": subject,
                "body": body,
                "has_attachments": has_attach,
                "is_read": is_read,
                "received": received,
            })

            exported += 1
            # Прогресс каждые 10 писем
            if i % 10 == 0 or i == count:
                print(f"  Обработано: {i}/{count}")

        except Exception as e:
            print(f"  ! Ошибка письма #{i}: {e}")
            errors += 1
            continue

    print(f"\nЭкспортировано: {exported}, ошибок: {errors}")
    return results


def save_excel(results, output_dir):
    """Сохраняет реестр в Почта.xlsx."""
    filepath = os.path.join(output_dir, "Почта.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Реестр"

    # Заголовки
    headers = ["Link", "Subject", "TextBody", "HasAttachments",
               "IsRead", "DateTimeReceived", "Тема"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)

    # Данные
    for i, row in enumerate(results, 2):
        ws.cell(row=i, column=1, value=row["link"])
        ws.cell(row=i, column=2, value=f"FW: {row['subject']}")
        ws.cell(row=i, column=3, value=row["body"])
        ws.cell(row=i, column=4, value=row["has_attachments"])
        ws.cell(row=i, column=5, value=row["is_read"])
        try:
            # pywintypes.datetime → строка
            dt = row["received"]
            ws.cell(row=i, column=6, value=datetime(
                dt.year, dt.month, dt.day, dt.hour, dt.minute, dt.second))
        except Exception:
            ws.cell(row=i, column=6, value=str(row["received"]))
        # Колонка G (Тема/Тип) — пустая, заполняется вручную
        ws.cell(row=i, column=7, value="")

    # Автоширина колонок (примерно)
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 8

    wb.save(filepath)
    print(f"\nРеестр сохранён: {filepath}")
    print(f"  Строк данных: {len(results)}")
    return filepath


def main():
    print("=" * 60)
    print("Outlook → Excel + .msg экспорт")
    print("=" * 60)
    print(f"Папка вывода: {OUTPUT_DIR}")
    print()

    # 1. Подключение
    namespace = connect_outlook()

    # 2. Выбор папки
    folder = select_folder(namespace)

    # 3. Фильтр
    filter_type = get_filter(folder)

    # 4. Экспорт
    results = export_messages(folder, filter_type)

    if not results:
        input("\nEnter для выхода...")
        return

    # 5. Сохранение Excel
    save_excel(results, get_output_dir())

    print(f"\n{'=' * 60}")
    print(f"ГОТОВО!")
    print(f"  .msg файлы: {OUTPUT_DIR}")
    print(f"  Реестр: {os.path.join(OUTPUT_DIR, 'Почта.xlsx')}")
    print(f"{'=' * 60}")

    input("\nEnter для выхода...")


if __name__ == "__main__":
    main()
